import sys
import subprocess

# Self-healing dependency check
for mod in ["xlsxwriter", "scipy", "python-calamine", "PyQt6"]:
    try:
        if mod == "xlsxwriter": import xlsxwriter
        if mod == "scipy": import scipy
        if mod == "python-calamine": import python_calamine
        if mod == "PyQt6": import PyQt6
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", mod])

import os
import shutil
import time
import re
from pathlib import Path
import pandas as pd
import numpy as np
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QLabel, QPushButton, QLineEdit, QFileDialog, QProgressBar, 
    QFrame, QComboBox, QMessageBox, QCompleter, QTabWidget, QSplitter,
    QTableWidget, QTableWidgetItem, QHeaderView, QScrollArea, 
    QInputDialog, QAbstractItemView, QStyledItemDelegate, QMenu,
    QGridLayout, QListWidget, QDialog, QFormLayout, QDateEdit, QDoubleSpinBox,
    QSpinBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QEvent, QRect, QDate
from PyQt6.QtGui import QPainter, QPen, QColor, QPainterPath, QStandardItem, QStandardItemModel, QAction, QIcon, QBrush, QFont
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.pyplot as plt
from matplotlib.figure import Figure

import engine
import history

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))

    return os.path.join(base_path, relative_path)


# ============================================================
# MODERN LIGHT MODE QSS STYLING
# ============================================================
STYLESHEET = """
QMainWindow { background-color: #f1f5f9; }
* { outline: none; }
QWidget { background-color: #f1f5f9; color: #1e293b; font-family: 'Segoe UI', sans-serif; }
QFrame#ControlPanel { background-color: #ffffff; border-radius: 12px; border: 1px solid #e2e8f0; }
QLabel#HeaderLabel { font-size: 24px; font-weight: bold; color: #0f172a; margin-bottom: 5px; }
QLabel#SubHeader { color: #64748b; font-size: 13px; }
QPushButton { background-color: transparent; color: #2563eb; border: 2px solid #2563eb; border-radius: 6px; padding: 8px 16px; font-weight: normal; }
QPushButton:hover { background-color: #eff6ff; border-color: #1d4ed8; color: #1d4ed8; }
QPushButton#SecondaryBtn { background-color: transparent; color: #64748b; border: 1px solid #cbd5e1; }
QPushButton#SecondaryBtn:hover { background-color: #f1f5f9; border-color: #94a3b8; }
QPushButton#ExportBtn { background-color: transparent; color: #10b981; border: 2px solid #10b981; }
QPushButton#ExportBtn:hover { background-color: #f0fdf4; border-color: #059669; color: #059669; }
QLineEdit, QComboBox, QDateEdit, QSpinBox, QDoubleSpinBox, QTextEdit { 
    background-color: #ffffff; 
    border: 1px solid #cbd5e1; 
    border-radius: 4px; 
    padding: 4px 8px; 
    min-height: 32px;
    color: #0f172a; 
    selection-background-color: #e2e8f0; 
    selection-color: #000000;
}
QLineEdit:focus, QComboBox:focus, QDateEdit:focus, QSpinBox:focus { border-color: #2563eb; }
QAbstractSpinBox { background-color: #ffffff; min-height: 32px; }
QAbstractSpinBox::up-button, QAbstractSpinBox::down-button { background: #f8fafc; border-left: 1px solid #cbd5e1; }
QProgressBar { border: 1px solid #e2e8f0; border-radius: 10px; text-align: center; background-color: #f1f5f9; height: 16px; font-weight: bold; color: #0f172a; }
QProgressBar::chunk { background-color: #22c55e; border-radius: 9px; }
QTabWidget::pane { border: 1px solid #e2e8f0; border-radius: 8px; top: -1px; background: white; }
QTabBar::tab { background: #f1f5f9; border: 1px solid #e2e8f0; padding: 8px 16px; border-top-left-radius: 8px; border-top-right-radius: 8px; color: #64748b; font-weight: normal; }
QTabBar::tab:selected { background: white; border-bottom-color: white; color: #2563eb; }
QTabBar::tab:disabled { color: rgba(100, 116, 139, 0.35); background: rgba(241, 245, 249, 0.4); border-color: rgba(226, 232, 240, 0.4); }
QTableWidget { gridline-color: #f1f5f9; border: none; alternate-background-color: #f8fafc; selection-background-color: #dbeafe; selection-color: #1e293b; }
QHeaderView::section { background-color: #ffffff; padding: 10px; border: none; border-bottom: 2px solid #f1f5f9; font-weight: bold; color: #0f172a; }
QMenu { background-color: #ffffff; border: 1px solid #cbd5e1; border-radius: 6px; padding: 4px; }
QMenu::item { background-color: transparent; padding: 6px 20px; border-radius: 4px; color: #1e293b; }
QMenu::item:selected { background-color: #eff6ff; color: #2563eb; }

/* DROPDOWN LIST STYLING */
QComboBox QAbstractItemView {
    background-color: #ffffff;
    border: 1px solid #cbd5e1;
    selection-background-color: #eff6ff;
    selection-color: #2563eb;
    outline: none;
    padding: 4px;
}
QComboBox QAbstractItemView::item {
    padding: 6px;
    border-radius: 4px;
}
QComboBox QAbstractItemView::item:selected {
    background-color: #eff6ff;
    color: #2563eb;
}
QComboBox QAbstractItemView::item:disabled {
    color: rgba(148, 163, 184, 0.35);
}

/* MODERN THIN SCROLLBAR */
QScrollBar:vertical {
    border: none;
    background: transparent;
    width: 4px;
    margin: 0px;
}
QScrollBar::handle:vertical {
    background: #cbd5e1;
    min-height: 30px;
    border-radius: 2px;
}
QScrollBar::handle:vertical:hover {
    background: #94a3b8;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    border: none;
    background: none;
    height: 0px;
}
QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {
    border: none;
    background: none;
}
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
    background: none;
}

QScrollBar:horizontal {
    border: none;
    background: transparent;
    height: 4px;
    margin: 0px;
}
QScrollBar::handle:horizontal {
    background: #cbd5e1;
    min-width: 30px;
    border-radius: 2px;
}
QScrollBar::handle:horizontal:hover {
    background: #94a3b8;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    border: none;
    background: none;
    width: 0px;
}
QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
    background: none;
}

/* CALENDAR POPUP STYLING */
QCalendarWidget QWidget { background-color: #ffffff; }
QCalendarWidget QAbstractItemView {
    background-color: #ffffff;
    selection-background-color: #2563eb;
    selection-color: #ffffff;
}
QCalendarWidget QToolButton {
    color: #1e293b;
    background-color: transparent;
    border: none;
    font-weight: bold;
}
QCalendarWidget QToolButton::menu-indicator {
    image: none;
    width: 0px;
}
QCalendarWidget QMenu { background-color: #ffffff; }
QCalendarWidget QSpinBox { 
    background-color: transparent; 
    border: none; 
    min-height: 0px; 
    padding: 0px; 
    color: #1e293b;
    font-weight: bold;
}
QCalendarWidget QSpinBox::up-button, QCalendarWidget QSpinBox::down-button { 
    width: 0px; 
    border: none; 
    image: none; 
}
"""

# ============================================================
# CONFIGURATION
# ============================================================
if getattr(sys, 'frozen', False):
    DEFAULT_DATA_PATH = os.path.join(os.path.dirname(sys.executable), "Data")
else:
    DEFAULT_DATA_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data")

# ============================================================
# WORKERS
# ============================================================
class ScanWorker(QThread):
    scan_finished = pyqtSignal(list)
    error = pyqtSignal(str)
    def __init__(self, path):
        super().__init__()
        self.path = path
    def run(self):
        try:
            groups = engine.get_available_groups(self.path)
            if not groups: self.error.emit("No departments found. Check column 'M_STORE_DEPARTMENT'.")
            else: self.scan_finished.emit(groups)
        except Exception as e: self.error.emit(str(e))

# ============================================================
# CUSTOM UI DELEGATES
# ============================================================
class SparklineDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        data = index.data(Qt.ItemDataRole.UserRole)
        if not data or not isinstance(data, list):
            super().paint(painter, option, index)
            return

        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        rect = option.rect.adjusted(4, 4, -4, -4)
        if not data: return
        
        max_val = max(data) if max(data) > 0 else 1
        n = len(data)
        
        path = QPainterPath()
        x_step = rect.width() / (n - 1) if n > 1 else rect.width()
        
        for i, val in enumerate(data):
            x = rect.left() + i * x_step
            y = rect.bottom() - (val / max_val) * rect.height()
            if i == 0: path.moveTo(x, y)
            else: path.lineTo(x, y)
            
        painter.setPen(QPen(QColor("#2563eb"), 1.5))
        painter.drawPath(path)
        painter.restore()

class DataBarDelegate(QStyledItemDelegate):
    def __init__(self, color_hex, max_val=None):
        super().__init__()
        self.color = QColor(color_hex)
        self.max_val = max_val

    def paint(self, painter, option, index):
        try:
            text = index.data(Qt.ItemDataRole.DisplayRole)
            if not text or text == "-":
                super().paint(painter, option, index)
                return

            ratio = 0.0
            import re
            
            # Check if this cell is part of the Grand Total row
            is_grand_total = False
            tb = option.widget
            if isinstance(tb, QTableWidget):
                item_dept = tb.item(index.row(), 0)
                if item_dept and item_dept.text() == "Grand Total":
                    is_grand_total = True

            # If it's the Grand Total row, we do NOT draw a data bar
            if not is_grand_total:
                pct_match = re.search(r"([\d\.]+)\s*%", text)
                if pct_match:
                    try:
                        pct_val = float(pct_match.group(1))
                        
                        # Dynamically find the min and max percentages in this column (excluding Grand Total)
                        pct_vals = []
                        if isinstance(tb, QTableWidget):
                            col = index.column()
                            for r in range(tb.rowCount()):
                                r_dept = tb.item(r, 0)
                                if r_dept and r_dept.text() == "Grand Total":
                                    continue
                                cell_item = tb.item(r, col)
                                if cell_item:
                                    cell_text = cell_item.text()
                                    m = re.search(r"([\d\.]+)\s*%", cell_text)
                                    if m:
                                        try:
                                            pct_vals.append(float(m.group(1)))
                                        except ValueError:
                                            pass
                        
                        if pct_vals:
                            min_pct = min(pct_vals)
                            max_pct = max(pct_vals)
                            if max_pct > min_pct:
                                # Scale between 0.05 (5% minimum sliver) and 1.0 (100% width)
                                ratio = 0.05 + 0.95 * ((pct_val - min_pct) / (max_pct - min_pct))
                            else:
                                ratio = 1.0
                        else:
                            ratio = min(pct_val / 100.0, 1.0)
                    except ValueError:
                        ratio = 0.0
                elif self.max_val and self.max_val > 0:
                    # Fallback to standard ratio relative to max_val
                    try:
                        clean_text = text.replace("%", "").replace(",", "")
                        val = float(clean_text)
                        
                        # Dynamically find the min and max values in this column for raw numbers too
                        vals = []
                        if isinstance(tb, QTableWidget):
                            col = index.column()
                            for r in range(tb.rowCount()):
                                r_dept = tb.item(r, 0)
                                if r_dept and r_dept.text() == "Grand Total":
                                    continue
                                cell_item = tb.item(r, col)
                                if cell_item:
                                    try:
                                        c_val = float(cell_item.text().replace("%", "").replace(",", ""))
                                        vals.append(c_val)
                                    except ValueError:
                                        pass
                        
                        if vals:
                            min_v = min(vals)
                            max_v = max(vals)
                            if max_v > min_v:
                                ratio = 0.05 + 0.95 * ((val - min_v) / (max_v - min_v))
                            else:
                                ratio = 1.0
                        else:
                            ratio = min(val / self.max_val, 1.0)
                    except ValueError:
                        ratio = 0.0

            # --- Step 1: Draw the background and bar ---
            painter.save()
            
            # Draw default item panel (handles hover, selection, alternating colors, etc.)
            style = option.widget.style() if option.widget else QApplication.style()
            style.drawPrimitive(style.PrimitiveElement.PE_PanelItemViewItem, option, painter, option.widget)
            
            if ratio > 0.0 and not is_grand_total:
                # Full width of cell minus some padding
                bar_w = int(option.rect.width() * ratio)
                if bar_w < 3: bar_w = 3 # Ensure it's visible
                
                # Draw bar with padding for a modern pill look
                bar_rect = QRect(option.rect.left() + 2, option.rect.top() + 3, bar_w - 4 if bar_w > 4 else bar_w, option.rect.height() - 6)

                # Use a beautiful pastel/soft alpha of the color
                fill_color = QColor(self.color)
                fill_color.setAlpha(70) # Semi-transparent for maximum readability of text on top
                
                # Draw rounded rectangle
                painter.setPen(Qt.PenStyle.NoPen)
                painter.setBrush(QBrush(fill_color))
                painter.drawRoundedRect(bar_rect, 4, 4)

            painter.restore()

            # --- Step 2: Draw the text on top ---
            painter.save()
            option.backgroundBrush = QBrush(Qt.BrushStyle.NoBrush)
            super().paint(painter, option, index)
            painter.restore()

        except Exception:
            super().paint(painter, option, index)

def calculate_balance_score(y):
    """
    Balance Score:
    1 = perfectly balanced (left sum == right sum)
    0 = extremely imbalanced (all mass on one side)
    """
    total = sum(y)
    if total == 0:
        return 0.0
    n = len(y)
    mid = n // 2
    left_sum = sum(y[:mid])
    right_sum = total - left_sum
    bi = (left_sum - right_sum) / total
    score = 1.0 - abs(bi)
    return round(max(0.0, min(1.0, score)), 6)


def calculate_symmetry_score(y):
    """
    Symmetry Score:
    1 = perfectly symmetric
    0 = highly asymmetric
    """
    total = sum(y)
    if total == 0:
        return 0.0
    n = len(y)
    mid = n // 2
    left = y[:mid]
    if n % 2 == 0:
        right = y[mid:]
    else:
        right = y[mid + 1:]
    right_rev = right[::-1]
    m = min(len(left), len(right_rev))
    abs_diff = 0
    for i in range(m):
        abs_diff += abs((left[i] / total) - (right_rev[i] / total))
    symmetry = max(0.0, min(1.0, 1.0 - abs_diff))
    return round(symmetry, 6)


def calculate_center_score(y):
    """
    Center Score:
    1 = concentrated at center
    0 = concentrated at edges
    """
    total = sum(y)
    if total == 0:
        return 0.0
    n = len(y)
    center_idx = (n - 1) / 2.0
    max_dist = max(center_idx, (n - 1) - center_idx)
    weighted_sum = 0
    for idx, value in enumerate(y):
        weight = 1.0 - abs(idx - center_idx) / max_dist
        weighted_sum += value * weight
    center_score = weighted_sum / total
    return round(center_score, 6)


def calculate_final_score(balance_score, symmetry_score, center_score):
    """
    Final Score:
    Average of Balance Score, Symmetry Score, Center Score
    """
    final_score = (
        balance_score +
        symmetry_score +
        center_score
    ) / 3
    return round(final_score, 6)

class ScoreBarDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        try:
            text = index.data(Qt.ItemDataRole.DisplayRole)
            if not text or text == "-":
                super().paint(painter, option, index)
                return

            # Skip drawing background bar if it's the Grand Total or Total row
            is_grand_total = False
            tb = option.widget
            if isinstance(tb, QTableWidget):
                item_dept = tb.item(index.row(), 0)
                if item_dept and ("total" in item_dept.text().lower()):
                    is_grand_total = True

            if is_grand_total:
                super().paint(painter, option, index)
                return

            is_percentage = "%" in str(text)

            try:
                # Strip commas and percentage signs for float conversion
                val = float(str(text).replace(",", "").replace("%", ""))
            except ValueError:
                super().paint(painter, option, index)
                return

            painter.save()
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            
            # Draw standard background (handles selection, hover, etc.)
            style = option.widget.style() if option.widget else QApplication.style()
            style.drawPrimitive(style.PrimitiveElement.PE_PanelItemViewItem, option, painter, option.widget)

            # Determine normalized value (norm_val) between 0.0 and 1.0
            if is_percentage:
                # Dynamically find the min and max percentages in this column (excluding Grand Total)
                pct_vals = []
                if isinstance(tb, QTableWidget):
                    col = index.column()
                    for r in range(tb.rowCount()):
                        r_dept = tb.item(r, 0)
                        if r_dept and ("total" in r_dept.text().lower()):
                            continue
                        cell_item = tb.item(r, col)
                        if cell_item:
                            cell_text = cell_item.text()
                            if "%" in cell_text:
                                try:
                                    pct_vals.append(float(cell_text.replace(",", "").replace("%", "")))
                                except ValueError:
                                    pass
                if pct_vals:
                    max_pct = max(pct_vals)
                    min_pct = min(pct_vals)
                    if max_pct > min_pct:
                        norm_val = (val - min_pct) / (max_pct - min_pct)
                    else:
                        norm_val = 1.0
                else:
                    norm_val = val / 100.0
            else:
                # Standard score column scaling logic
                norm_val = val
                if norm_val > 1.0:
                    if norm_val <= 2.5:
                        norm_val = norm_val / 2.0  # Scale score out of 2.0
                    elif norm_val <= 100.0:
                        norm_val = norm_val / 100.0  # Scale percentage out of 100
                    else:
                        norm_val = 1.0

            norm_val = max(0.0, min(1.0, norm_val))

            if norm_val < 0.5:
                color = QColor("#fee2e2") # light red
                text_color = QColor("#991b1b") # dark red
            elif norm_val < 0.75:
                color = QColor("#fef3c7") # light amber
                text_color = QColor("#92400e") # dark amber
            else:
                color = QColor("#d1fae5") # light emerald
                text_color = QColor("#065f46") # dark emerald

            # Draw a pill bar inside the cell
            margin = 3
            bar_w = int((option.rect.width() - 2 * margin) * norm_val)
            if bar_w < 5: bar_w = 5
            
            bar_rect = QRect(
                option.rect.left() + margin, 
                option.rect.top() + margin + 1, 
                bar_w, 
                option.rect.height() - 2 * margin - 2
            )
            
            # Fill with color
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QBrush(color))
            painter.drawRoundedRect(bar_rect, 4, 4)
            painter.restore()

            # Paint standard text on top with high-contrast text color
            painter.save()
            option.palette.setColor(option.palette.ColorGroup.All, option.palette.ColorRole.Text, text_color)
            option.displayAlignment = Qt.AlignmentFlag.AlignCenter
            super().paint(painter, option, index)
            painter.restore()

        except Exception:
            super().paint(painter, option, index)

class NumericTableItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            tb = self.tableWidget()
            if tb:
                r1 = self.row()
                r2 = other.row()
                item1_dept = tb.item(r1, 0)
                item2_dept = tb.item(r2, 0)
                is_gt1 = item1_dept and item1_dept.text() == "Grand Total"
                is_gt2 = item2_dept and item2_dept.text() == "Grand Total"
                
                if is_gt1 and not is_gt2:
                    order = tb.horizontalHeader().sortIndicatorOrder()
                    return False if order == Qt.SortOrder.AscendingOrder else True
                if is_gt2 and not is_gt1:
                    order = tb.horizontalHeader().sortIndicatorOrder()
                    return True if order == Qt.SortOrder.AscendingOrder else False
            
            t1 = re.sub(r'[^0-9.-]', '', self.text()).strip()
            t2 = re.sub(r'[^0-9.-]', '', other.text()).strip()
            return float(t1) < float(t2)
        except:
            return super().__lt__(other)

class CheckableComboBox(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.view().viewport().installEventFilter(self)
        self.setModel(QStandardItemModel(self))
        self._changed = False

    def eventFilter(self, widget, event):
        if widget == self.view().viewport() and event.type() == QEvent.Type.MouseButtonRelease:
            index = self.view().indexAt(event.position().toPoint())
            item = self.model().itemFromIndex(index)
            if item:
                if item.checkState() == Qt.CheckState.Checked:
                    item.setCheckState(Qt.CheckState.Unchecked)
                else:
                    item.setCheckState(Qt.CheckState.Checked)
                self._changed = True
            return True
        return super().eventFilter(widget, event)

    def hidePopup(self):
        if self._changed:
            self.currentIndexChanged.emit(self.currentIndex())
            self._changed = False
        super().hidePopup()

    def add_checkable_item(self, text, checked=False):
        item = QStandardItem(text)
        item.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
        item.setData(Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked, Qt.ItemDataRole.CheckStateRole)
        self.model().appendRow(item)

    def checked_items(self):
        try:
            checked = []
            for i in range(self.count()):
                item = self.model().item(i)
                if item.data(Qt.ItemDataRole.CheckStateRole) == Qt.CheckState.Checked:
                    checked.append(item.text())
            return checked
        except RuntimeError:
            return []

class LimitedCheckableComboBox(QComboBox):
    def __init__(self, max_choices=3, parent=None):
        super().__init__(parent)
        self.max_choices = max_choices
        self.view().viewport().installEventFilter(self)
        self.setModel(QStandardItemModel(self))
        self._changed = False

    def eventFilter(self, widget, event):
        if widget == self.view().viewport() and event.type() == QEvent.Type.MouseButtonRelease:
            from PyQt6.QtCore import Qt
            index = self.view().indexAt(event.position().toPoint())
            item = self.model().itemFromIndex(index)
            if item:
                if not item.isEnabled():
                    return True
                currently_checked = len(self.checked_items())
                current_state = item.data(Qt.ItemDataRole.CheckStateRole)
                
                if current_state == Qt.CheckState.Checked or current_state == Qt.CheckState.Checked.value:
                    if currently_checked <= 1:
                        from PyQt6.QtWidgets import QMessageBox
                        QMessageBox.warning(self, "Selection Required", "At least one reporting set must remain selected.")
                        return True
                    item.setData(Qt.CheckState.Unchecked, Qt.ItemDataRole.CheckStateRole)
                    self._changed = True
                    self.update()
                else:
                    if currently_checked >= self.max_choices:
                        from PyQt6.QtWidgets import QMessageBox
                        QMessageBox.warning(self, "Selection Limit", f"You can select a maximum of {self.max_choices} sets.")
                    else:
                        item.setData(Qt.CheckState.Checked, Qt.ItemDataRole.CheckStateRole)
                        self._changed = True
                        self.update()
            return True
        return super().eventFilter(widget, event)

    def checked_labels(self):
        try:
            from PyQt6.QtCore import Qt
            checked = []
            for i in range(self.count()):
                item = self.model().item(i)
                current_state = item.data(Qt.ItemDataRole.CheckStateRole)
                if current_state == Qt.CheckState.Checked or current_state == Qt.CheckState.Checked.value:
                    checked.append(item.text())
            return checked
        except RuntimeError:
            return []

    def paintEvent(self, event):
        from PyQt6.QtWidgets import QStylePainter, QStyleOptionComboBox
        painter = QStylePainter(self)
        opt = QStyleOptionComboBox()
        self.initStyleOption(opt)
        
        checked = self.checked_labels()
        if checked:
            opt.currentText = ", ".join(checked)
        else:
            opt.currentText = "None"
            
        from PyQt6.QtWidgets import QStyle
        painter.drawComplexControl(QStyle.ComplexControl.CC_ComboBox, opt)
        painter.drawControl(QStyle.ControlElement.CE_ComboBoxLabel, opt)

    def hidePopup(self):
        if self._changed:
            self.currentIndexChanged.emit(self.currentIndex())
            self._changed = False
        super().hidePopup()

    def hide_item_by_data(self, user_data):
        from PyQt6.QtCore import Qt
        for i in range(self.count()):
            item = self.model().item(i)
            if item.data(Qt.ItemDataRole.UserRole) == user_data:
                self.view().setRowHidden(i, True)
                item.setData(Qt.CheckState.Unchecked, Qt.ItemDataRole.CheckStateRole)
            else:
                self.view().setRowHidden(i, False)
        self.update()

    def add_checkable_item(self, text, user_data, checked=False):
        from PyQt6.QtGui import QStandardItem
        from PyQt6.QtCore import Qt
        item = QStandardItem(text)
        item.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
        item.setData(Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked, Qt.ItemDataRole.CheckStateRole)
        item.setData(user_data, Qt.ItemDataRole.UserRole)
        self.model().appendRow(item)

    def checked_items(self):
        try:
            from PyQt6.QtCore import Qt
            checked = []
            for i in range(self.count()):
                item = self.model().item(i)
                current_state = item.data(Qt.ItemDataRole.CheckStateRole)
                if current_state == Qt.CheckState.Checked or current_state == Qt.CheckState.Checked.value:
                    checked.append(item.data(Qt.ItemDataRole.UserRole))
            return checked
        except RuntimeError:
            return []

class MultiSelectFilterDialog(QDialog):
    def __init__(self, field_name, unique_values, parent=None):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QListWidget, QListWidgetItem, QLabel, QCheckBox
        from PyQt6.QtCore import Qt
        super().__init__(parent)
        self.setWindowTitle(f"Filter {field_name}")
        self.setMinimumSize(300, 400)
        self.setStyleSheet("QDialog { background-color: #f3f4f6; } QLabel { border: none; background: transparent; font-size: 12px; font-weight: bold; }")
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"Select multiple items for {field_name}:"))
        
        self.chk_all = QCheckBox("(All)")
        self.chk_all.setChecked(True)
        self.chk_all.stateChanged.connect(self.toggle_all)
        layout.addWidget(self.chk_all)
        
        self.list_widget = QListWidget()
        for val in unique_values:
            item = QListWidgetItem(val)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked)
            self.list_widget.addItem(item)
        layout.addWidget(self.list_widget)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        self.btn_ok = QPushButton("OK")
        self.btn_ok.clicked.connect(self.accept)
        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(self.btn_ok)
        btn_layout.addWidget(self.btn_cancel)
        layout.addLayout(btn_layout)
        
    def toggle_all(self, state):
        from PyQt6.QtCore import Qt
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(Qt.CheckState.Checked if state == 2 else Qt.CheckState.Unchecked)
            
    def get_selected_values(self):
        from PyQt6.QtCore import Qt
        selected = []
        for i in range(self.list_widget.count()):
            if self.list_widget.item(i).checkState() == Qt.CheckState.Checked:
                selected.append(self.list_widget.item(i).text())
        return selected

class CountryExportDialog(QDialog):
    def __init__(self, countries, country_names, active_country, parent=None):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QCheckBox, QScrollArea, QWidget, QMessageBox
        from PyQt6.QtCore import Qt
        super().__init__(parent)
        self.setWindowTitle("Export Charts - Select Countries")
        self.setMinimumSize(450, 450)
        self.setStyleSheet("""
            QDialog { background-color: #f8fafc; }
            QLabel { font-size: 13px; font-weight: 600; color: #1e293b; border: none; background: transparent; }
            QCheckBox { font-size: 13px; color: #334155; padding: 4px; }
            QCheckBox:disabled { color: rgba(51, 65, 85, 0.4); }
            QCheckBox::indicator { width: 18px; height: 18px; }
            QPushButton { 
                padding: 8px 16px; 
                font-size: 12px; 
                font-weight: bold; 
                border-radius: 6px; 
            }
            QPushButton#ExportBtn { 
                background-color: #2563eb; 
                color: white; 
                border: 1px solid #2563eb; 
            }
            QPushButton#ExportBtn:hover { 
                background-color: #1d4ed8; 
            }
            QPushButton#AllBtn { 
                background-color: #10b981; 
                color: white; 
                border: 1px solid #10b981; 
            }
            QPushButton#AllBtn:hover { 
                background-color: #059669; 
            }
            QPushButton#OneBtn { 
                background-color: #f59e0b; 
                color: white; 
                border: 1px solid #f59e0b; 
            }
            QPushButton#OneBtn:hover { 
                background-color: #d97706; 
            }
            QPushButton#CancelBtn { 
                background-color: #cbd5e1; 
                color: #334155; 
                border: 1px solid #cbd5e1; 
            }
            QPushButton#CancelBtn:hover { 
                background-color: #94a3b8; 
            }
        """)
        
        self.countries = countries
        self.country_names = country_names
        self.active_country = active_country
        self.result_selection = []
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        lbl_info = QLabel("Choose the countries to export charts:")
        layout.addWidget(lbl_info)
        
        # Scroll area for checkbox list
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: 1px solid #e2e8f0; border-radius: 6px; background-color: #ffffff; }")
        
        scroll_widget = QWidget()
        scroll_widget.setStyleSheet("background-color: #ffffff;")
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setContentsMargins(15, 15, 15, 15)
        scroll_layout.setSpacing(10)
        
        self.checkboxes = {}
        allowed_countries = ["ID", "TH"]
        
        for code in self.countries:
            full_name = "Brunei" if code == "BR" else self.country_names.get(code, code)
            chk = QCheckBox(f"{full_name} ({code})")
            
            if code in allowed_countries:
                if code == self.active_country:
                    chk.setChecked(True)
            else:
                chk.setEnabled(False)
                chk.setStyleSheet("color: rgba(51, 65, 85, 0.4);")
                chk.setToolTip(f"{full_name} is currently under development.")
                
            scroll_layout.addWidget(chk)
            self.checkboxes[code] = chk
            
        scroll_layout.addStretch()
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)
        
        # Buttons layout
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)
        
        btn_export = QPushButton("EXPORT")
        btn_export.setObjectName("ExportBtn")
        btn_export.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_export.clicked.connect(self.on_export_clicked)
        btn_layout.addWidget(btn_export)
        
        btn_all = QPushButton("ALL")
        btn_all.setObjectName("AllBtn")
        btn_all.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_all.clicked.connect(self.on_all_clicked)
        btn_layout.addWidget(btn_all)
        
        btn_one = QPushButton("ONE")
        btn_one.setObjectName("OneBtn")
        btn_one.setCursor(Qt.CursorShape.PointingHandCursor)
        active_name = self.country_names.get(self.active_country, self.active_country)
        btn_one.setToolTip(f"Export only the currently selected country: {active_name}")
        btn_one.clicked.connect(self.on_one_clicked)
        btn_layout.addWidget(btn_one)
        
        btn_layout.addStretch()
        
        btn_cancel = QPushButton("Cancel")
        btn_cancel.setObjectName("CancelBtn")
        btn_cancel.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancel)
        
        layout.addLayout(btn_layout)
        
    def get_selected_countries(self):
        selected = []
        for code, chk in self.checkboxes.items():
            if chk.isChecked():
                selected.append(code)
        return selected
        
    def confirm_and_accept(self, chosen_codes):
        from PyQt6.QtWidgets import QMessageBox
        if not chosen_codes:
            QMessageBox.warning(self, "No Selection", "Please select at least one country to export.")
            return
            
        chosen_names = [self.country_names.get(code, code) for code in chosen_codes]
        list_str = ", ".join(chosen_names)
        
        reply = QMessageBox.question(
            self, 
            "Confirm Export", 
            f"Confirm export for ({list_str})?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.result_selection = chosen_codes
            self.accept()
            
    def on_export_clicked(self):
        codes = self.get_selected_countries()
        self.confirm_and_accept(codes)
        
    def on_all_clicked(self):
        allowed_codes = ["ID", "TH"]
        # Check only allowed ones in UI
        for code, chk in self.checkboxes.items():
            if code in allowed_codes:
                chk.setChecked(True)
            else:
                chk.setChecked(False)
        self.confirm_and_accept(allowed_codes)
        
    def on_one_clicked(self):
        self.confirm_and_accept([self.active_country])

class FieldSettingsDialog(QDialog):
    def __init__(self, source_name, custom_name, parent=None):
        from PyQt6.QtWidgets import QDialog, QLabel, QLineEdit, QTabWidget, QRadioButton, QPushButton, QHBoxLayout, QVBoxLayout, QFormLayout, QWidget, QButtonGroup
        super().__init__(parent)
        self.setWindowTitle("Field Settings")
        self.setMinimumSize(400, 300)
        self.setStyleSheet("QDialog { background-color: #f3f4f6; } QLabel { border: none; background: transparent; font-size: 12px; }")
        
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        form_layout = QFormLayout()
        form_layout.setSpacing(8)
        
        lbl_src_title = QLabel("Source Name:")
        lbl_src_title.setStyleSheet("font-weight: bold; color: #be3519;")
        self.lbl_src = QLabel(source_name)
        self.lbl_src.setStyleSheet("color: #374151;")
        form_layout.addRow(lbl_src_title, self.lbl_src)
        
        lbl_cust_title = QLabel("Custom Name:")
        lbl_cust_title.setStyleSheet("font-weight: bold; color: #be3519;")
        self.edit_custom = QLineEdit(custom_name)
        self.edit_custom.setStyleSheet("padding: 4px; border: 1px solid #cbd5e1; border-radius: 4px; background-color: white;")
        form_layout.addRow(lbl_cust_title, self.edit_custom)
        
        main_layout.addLayout(form_layout)
        
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane { border: 1px solid #cbd5e1; background: white; border-radius: 4px; }
            QTabBar::tab { background: #e5e7eb; padding: 6px 12px; border-top-left-radius: 4px; border-top-right-radius: 4px; }
            QTabBar::tab:selected { background: white; border: 1px solid #cbd5e1; border-bottom: none; }
        """)
        
        tab_sub = QWidget()
        sub_vbox = QVBoxLayout(tab_sub)
        sub_vbox.setContentsMargins(10, 10, 10, 10)
        sub_vbox.setSpacing(8)
        
        lbl_sub = QLabel("Subtotals")
        lbl_sub.setStyleSheet("font-weight: bold; color: #be3519;")
        sub_vbox.addWidget(lbl_sub)
        
        self.bg = QButtonGroup(self)
        self.rad_auto = QRadioButton("Automatic")
        self.rad_none = QRadioButton("None")
        self.rad_custom = QRadioButton("Custom")
        self.rad_auto.setChecked(True)
        
        self.bg.addButton(self.rad_auto)
        self.bg.addButton(self.rad_none)
        self.bg.addButton(self.rad_custom)
        
        sub_vbox.addWidget(self.rad_auto)
        sub_vbox.addWidget(self.rad_none)
        sub_vbox.addWidget(self.rad_custom)
        
        self.tabs.addTab(tab_sub, "Subtotals & Filters")
        
        tab_layout = QWidget()
        layout_vbox = QVBoxLayout(tab_layout)
        lbl_layout = QLabel("Layout options:")
        lbl_layout.setStyleSheet("color: #374151;")
        layout_vbox.addWidget(lbl_layout)
        self.tabs.addTab(tab_layout, "Layout & Print")
        
        main_layout.addWidget(self.tabs)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.btn_ok = QPushButton("OK")
        self.btn_ok.setStyleSheet("padding: 6px 18px; background-color: #2563eb; color: white; font-weight: bold; border-radius: 4px;")
        self.btn_ok.clicked.connect(self.accept)
        
        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.setStyleSheet("padding: 6px 18px; border: 1px solid #cbd5e1; border-radius: 4px; background-color: white;")
        self.btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addWidget(self.btn_ok)
        btn_layout.addWidget(self.btn_cancel)
        
        main_layout.addLayout(btn_layout)
        
    def get_settings(self):
        return self.edit_custom.text()

class ValueFieldSettingsDialog(QDialog):
    def __init__(self, source_name, custom_name, current_func="Count", parent=None):
        from PyQt6.QtWidgets import QDialog, QLabel, QLineEdit, QTabWidget, QListWidget, QPushButton, QHBoxLayout, QVBoxLayout, QFormLayout, QWidget
        super().__init__(parent)
        self.setWindowTitle("Value Field Settings")
        self.setMinimumSize(400, 350)
        self.setStyleSheet("QDialog { background-color: #f3f4f6; } QLabel { border: none; background: transparent; font-size: 12px; }")
        
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        form_layout = QFormLayout()
        form_layout.setSpacing(8)
        
        lbl_src_title = QLabel("Source Name:")
        lbl_src_title.setStyleSheet("font-weight: bold; color: #374151;")
        self.lbl_src = QLabel(source_name)
        self.lbl_src.setStyleSheet("color: #374151;")
        form_layout.addRow(lbl_src_title, self.lbl_src)
        
        lbl_cust_title = QLabel("Custom Name:")
        lbl_cust_title.setStyleSheet("font-weight: bold; color: #374151;")
        self.edit_custom = QLineEdit(custom_name)
        self.edit_custom.setStyleSheet("padding: 4px; border: 1px solid #cbd5e1; border-radius: 4px; background-color: white;")
        form_layout.addRow(lbl_cust_title, self.edit_custom)
        
        main_layout.addLayout(form_layout)
        
        # Tabs
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane { border: 1px solid #cbd5e1; background: white; border-radius: 4px; }
            QTabBar::tab { background: #e5e7eb; padding: 6px 12px; border-top-left-radius: 4px; border-top-right-radius: 4px; }
            QTabBar::tab:selected { background: white; border: 1px solid #cbd5e1; border-bottom: none; }
        """)
        
        # Tab 1: Summarize Values By
        tab_summ = QWidget()
        summ_vbox = QVBoxLayout(tab_summ)
        summ_vbox.setContentsMargins(10, 10, 10, 10)
        summ_vbox.setSpacing(6)
        
        lbl_summ_bold = QLabel("Summarize value field by")
        lbl_summ_bold.setStyleSheet("font-weight: bold; color: #be3519;")
        summ_vbox.addWidget(lbl_summ_bold)
        
        lbl_summ_desc = QLabel("Choose the type of calculation that you want to use to summarize data from the selected field")
        lbl_summ_desc.setWordWrap(True)
        lbl_summ_desc.setStyleSheet("color: #4b5563;")
        summ_vbox.addWidget(lbl_summ_desc)
        
        self.func_list = QListWidget()
        self.func_list.setStyleSheet("border: 1px solid #cbd5e1; border-radius: 4px; background-color: white;")
        funcs = ["Sum", "Count", "Average", "Max", "Min", "Product", "Count Numbers", "StdDev", "StdDevp", "Var", "Varp"]
        self.func_list.addItems(funcs)
        
        # Select current func
        for i in range(self.func_list.count()):
            if self.func_list.item(i).text() == current_func:
                self.func_list.setCurrentRow(i)
                break
        
        summ_vbox.addWidget(self.func_list)
        self.func_list.currentItemChanged.connect(self.update_custom_name)
        
        self.tabs.addTab(tab_summ, "Summarize Values By")
        
        # Tab 2: Show Values As
        tab_show = QWidget()
        show_vbox = QVBoxLayout(tab_show)
        show_vbox.setContentsMargins(10, 10, 10, 10)
        show_vbox.setSpacing(6)
        
        lbl_show_bold = QLabel("Show values as")
        lbl_show_bold.setStyleSheet("font-weight: bold; color: #be3519;")
        show_vbox.addWidget(lbl_show_bold)
        
        self.combo_calc = QComboBox()
        self.combo_calc.setStyleSheet("padding: 4px; border: 1px solid #cbd5e1; border-radius: 4px; background-color: white;")
        calcs = ["No Calculation", "% of Grand Total", "% of Column Total", "% of Row Total", "% of", "% of Parent Row Total", "% of Parent Column Total", "% of Parent Total", "Difference From", "% Difference From", "Running Total In", "% Running Total In", "Rank Smallest to Largest", "Rank Largest to Smallest", "Index"]
        self.combo_calc.addItems(calcs)
        show_vbox.addWidget(self.combo_calc)
        
        lists_hbox = QHBoxLayout()
        lists_hbox.setSpacing(10)
        
        vbox_base_f = QVBoxLayout()
        lbl_base_f = QLabel("Base field:")
        lbl_base_f.setStyleSheet("color: #4b5563;")
        self.list_base_f = QListWidget()
        self.list_base_f.setStyleSheet("border: 1px solid #cbd5e1; border-radius: 4px; background-color: white;")
        
        default_fields = ["Store_Floor_Area", "Store_Display_Area", "Set 1", "Set 2", "Set 3", "Set 4"]
        self.list_base_f.addItems(default_fields)
        vbox_base_f.addWidget(lbl_base_f)
        vbox_base_f.addWidget(self.list_base_f)
        lists_hbox.addLayout(vbox_base_f)
        
        vbox_base_i = QVBoxLayout()
        lbl_base_i = QLabel("Base item:")
        lbl_base_i.setStyleSheet("color: #4b5563;")
        self.list_base_i = QListWidget()
        self.list_base_i.setStyleSheet("border: 1px solid #cbd5e1; border-radius: 4px; background-color: white;")
        vbox_base_i.addWidget(lbl_base_i)
        vbox_base_i.addWidget(self.list_base_i)
        lists_hbox.addLayout(vbox_base_i)
        
        show_vbox.addLayout(lists_hbox)
        
        self.tabs.addTab(tab_show, "Show Values As")
        
        main_layout.addWidget(self.tabs)
        
        # Bottom Buttons
        btn_layout = QHBoxLayout()
        
        self.btn_num_format = QPushButton("Number Format")
        self.btn_num_format.setStyleSheet("padding: 6px 12px; border: 1px solid #cbd5e1; border-radius: 4px; background-color: white;")
        btn_layout.addWidget(self.btn_num_format)
        
        btn_layout.addStretch()
        
        self.btn_ok = QPushButton("OK")
        self.btn_ok.setStyleSheet("padding: 6px 18px; background-color: #2563eb; color: white; font-weight: bold; border-radius: 4px;")
        self.btn_ok.clicked.connect(self.accept)
        
        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.setStyleSheet("padding: 6px 18px; border: 1px solid #cbd5e1; border-radius: 4px; background-color: white;")
        self.btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addWidget(self.btn_ok)
        btn_layout.addWidget(self.btn_cancel)
        
        main_layout.addLayout(btn_layout)
        
    def get_settings(self):
        func = "Count"
        if self.func_list.currentItem():
            func = self.func_list.currentItem().text()
        return func, self.edit_custom.text()
        
    def update_custom_name(self, current, previous):
        if current:
            func = current.text()
            src_name = self.lbl_src.text()
            self.edit_custom.setText(f"{func} of {src_name}")

class PivotTableCard(QFrame):
    def __init__(self, title="Pivot Table", parent_tab=None):
        from PyQt6.QtWidgets import QFrame, QVBoxLayout, QLineEdit, QTableWidget
        from PyQt6.QtCore import Qt
        super().__init__()
        self.title = title
        self.parent_tab = parent_tab
        self.filt_c = []
        self.row_c = []
        self.col_c = []
        self.val_items = []
        self.filter_values = {} # {col_name: [list_of_selected_values]}
        
        self.setStyleSheet("""
            QFrame { 
                background-color: #ffffff; 
                border: 1px solid #e2e8f0; 
                border-radius: 6px; 
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)
        
        from PyQt6.QtWidgets import QHBoxLayout, QPushButton
        header_h = QHBoxLayout()
        self.edit_title = QLineEdit(title)
        self.edit_title.setStyleSheet("font-weight: bold; color: #0f172a; font-size: 13px; border: 1px solid transparent; padding: 2px;")
        self.edit_title.setPlaceholderText("Pivot Title")
        header_h.addWidget(self.edit_title, 1)
        
        self.btn_delete = QPushButton("x")
        self.btn_delete.setObjectName("DeleteCardBtn")
        self.btn_delete.setFixedSize(28, 28)
        self.btn_delete.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_delete.setStyleSheet("""
            QPushButton#DeleteCardBtn { 
                border: none; 
                background-color: #e2e8f0; 
                color: #475569; 
                font-size: 14px; 
                font-weight: normal;
                border-radius: 14px;
                padding: 0px;
            } 
            QPushButton#DeleteCardBtn:hover { 
                background-color: #ef4444; 
                color: #ffffff; 
            }
        """)
        self.btn_delete.clicked.connect(self.delete_card)
        header_h.addWidget(self.btn_delete)
        
        layout.addLayout(header_h)
        
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("background-color: white; border: 1px solid #cbd5e1; border-radius: 4px;")
        layout.addWidget(self.table)
        
        self.setMinimumSize(350, 300)
        
        self.edit_title.installEventFilter(self)
        self.table.installEventFilter(self)
        
    def delete_card(self):
        if self.parent_tab:
            # Skip confirmation if the pivot is completely empty (no fields assigned)
            is_empty = not (self.filt_c or self.row_c or self.col_c or self.val_items)
            
            if is_empty:
                self.perform_delete()
                return

            from PyQt6.QtWidgets import QMessageBox
            msg = QMessageBox(self)
            msg.setWindowTitle("Delete Pivot")
            msg.setText(f"Are you sure you want to delete '{self.edit_title.text()}'?")
            msg.setIcon(QMessageBox.Icon.Question)
            msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            
            no_btn = msg.button(QMessageBox.StandardButton.No)
            if no_btn:
                no_btn.setStyleSheet("border: 2px solid #ef4444; color: #ef4444; font-weight: bold;")
                
            msg.setStyleSheet("QLabel { border: none; outline: none; background: transparent; } QMessageBox { outline: none; }")
                
            reply = msg.exec()
            if reply == QMessageBox.StandardButton.Yes:
                self.perform_delete()

    def perform_delete(self):
        if self in self.parent_tab.cards:
            self.parent_tab.cards.remove(self)
        
        if self.parent_tab.active_card == self:
            self.parent_tab.active_card = None
        
        self.setParent(None)
        self.deleteLater()
        
        if hasattr(self.parent_tab, 'rearrange_cards'):
            self.parent_tab.rearrange_cards()

    def eventFilter(self, obj, event):
        from PyQt6.QtCore import QEvent
        if event.type() in [QEvent.Type.MouseButtonPress, QEvent.Type.FocusIn]:
            self.activate()
        return super().eventFilter(obj, event)
        
    def mousePressEvent(self, event):
        super().mousePressEvent(event)
        self.activate()
        
    def activate(self):
        if self.parent_tab:
            self.parent_tab.set_active_pivot(self)

class PivotFieldCard(QFrame):
    def __init__(self, text, parent_list, list_item, is_value=False):
        super().__init__()
        self.parent_list = parent_list
        self.list_item = list_item
        self.field_text = text
        self.is_value = is_value
        
        self.setStyleSheet("""
            QFrame#card { 
                background-color: #f4f4f5; 
                border: 1px solid #cbd5e1; 
                border-radius: 2px; 
            }
        """)
        self.setObjectName("card")
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(6, 2, 26, 2)
        layout.setSpacing(4)
        
        display_text = f"Count of {text}" if is_value else text
        self.lbl = QLabel(display_text)
        self.lbl.setStyleSheet("font-size: 12px; color: #18181b; border: none; background: transparent;")
        layout.addWidget(self.lbl, 1)
        
        self.btn_arrow = QPushButton("v")
        self.btn_arrow.setFixedWidth(30)
        self.btn_arrow.setFixedHeight(30)
        self.btn_arrow.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_arrow.setStyleSheet("""
            QPushButton { 
                border: none; 
                color: #000000; 
                font-size: 13px; 
                font-weight: bold;
                background: transparent; 
                padding-right: 4px;
            } 
            QPushButton:hover { 
                color: #374151; 
            }
        """)
        layout.addWidget(self.btn_arrow)
        self.btn_arrow.clicked.connect(self.show_menu)
        
    def _move_item(self, new_row):
        """Helper to move this card to a new position in the same list."""
        from PyQt6.QtWidgets import QListWidgetItem
        from PyQt6.QtCore import QSize, Qt
        old_row = -1
        for i in range(self.parent_list.count()):
            if self.parent_list.itemWidget(self.parent_list.item(i)) is self:
                old_row = i
                break
        if old_row == -1 or old_row == new_row:
            return
        self.parent_list.takeItem(old_row)
        new_item = QListWidgetItem()
        new_item.setData(Qt.ItemDataRole.UserRole, self.field_text)
        new_item.setSizeHint(QSize(0, 32))
        self.parent_list.insertItem(new_row, new_item)
        new_card = PivotFieldCard(self.field_text, self.parent_list, new_item, self.is_value)
        self.parent_list.setItemWidget(new_item, new_card)
        self.parent_list.setCurrentItem(new_item)

    def show_menu(self):
        from PyQt6.QtWidgets import QMenu
        from PyQt6.QtCore import Qt
        menu = QMenu(self)
        menu.setStyleSheet("QMenu { background-color: #f8fafc; border: 1px solid #cbd5e1; font-size: 12px; } QMenu::item:selected { background-color: #e2e8f0; color: #0f172a; } QMenu::item:disabled { color: #94a3b8; }")
        
        # Find exact row index dynamically to avoid stale references
        current_row = -1
        total_count = self.parent_list.count()
        for i in range(total_count):
            if self.parent_list.itemWidget(self.parent_list.item(i)) is self:
                current_row = i
                break
                
        is_top = (current_row == 0)
        is_bottom = (current_row >= total_count - 1 or current_row == -1)
        
        act_up = menu.addAction("Move Up")
        act_up.setEnabled(not is_top)
        act_down = menu.addAction("Move Down")
        act_down.setEnabled(not is_bottom)
        act_beg = menu.addAction("Move to Beginning")
        act_beg.setEnabled(not is_top)
        act_end = menu.addAction("Move to End")
        act_end.setEnabled(not is_bottom)
        menu.addSeparator()
        
        act_rem = menu.addAction("Remove Field")
        
        act_filter = None
        if not self.is_value:
            act_filter = menu.addAction("Filter...")
            
        act_settings = menu.addAction("Field Settings...")
        
        action = menu.exec(self.btn_arrow.mapToGlobal(self.btn_arrow.rect().bottomLeft()))
        
        if not action or current_row == -1:
            return
            
        # Refresh current row index in case of mid-execution state shift
        for i in range(self.parent_list.count()):
            if self.parent_list.itemWidget(self.parent_list.item(i)) is self:
                current_row = i
                break
        
        if action == act_up and current_row > 0:
            self._move_item(current_row - 1)
        elif action == act_down and current_row < self.parent_list.count() - 1:
            self._move_item(current_row + 1)
        elif action == act_beg and current_row > 0:
            self._move_item(0)
        elif action == act_end and current_row < self.parent_list.count() - 1:
            self._move_item(self.parent_list.count() - 1)
        elif action == act_rem:
            self.parent_list.takeItem(current_row)
        elif action == act_settings:
            if self.is_value:
                current_display_text = self.lbl.text()
                current_func = "Count"
                for f in ["Sum", "Count", "Average", "Max", "Min", "Product", "Count Numbers", "StdDev", "StdDevp", "Var", "Varp"]:
                    if current_display_text.startswith(f):
                        current_func = f
                        break
                        
                custom_name = current_display_text
                
                dialog = ValueFieldSettingsDialog(self.field_text, custom_name, current_func, self)
                if dialog.exec():
                    func, new_name = dialog.get_settings()
                    self.lbl.setText(new_name)
            else:
                custom_name = self.lbl.text()
                # Generic settings for Rows/Columns/Filters could go here
        elif action == act_filter and act_filter:
            # Trigger filtering
            if self.parent_list and hasattr(self.parent_list, 'parent_tab'):
                tab = self.parent_list.parent_tab
                if tab and tab.source_df is not None and tab.active_card:
                    df = tab.source_df
                    col = self.field_text
                    unique_vals = [str(x) for x in df[col].dropna().unique()]
                    unique_vals.sort()
                    
                    from PyQt6.QtWidgets import QDialog
                    # Need to reach MultiSelectFilterDialog - assuming it's available in global scope or via parent
                    # For now, let's assume apply_pivot_logic will handle it if we trigger it
                    # But we need to pass the selected values
                    
                    current_selected = tab.active_card.filter_values.get(col, [])
                    dialog = MultiSelectFilterDialog(col, unique_vals, self)
                    # We should probably pre-select values in dialog if possible
                    
                    if dialog.exec():
                        selected = dialog.get_selected_values()
                        tab.active_card.filter_values[col] = selected
                        # Trigger auto-update
                        if hasattr(tab, 'apply_pivot_logic'):
                            tab.apply_pivot_logic()
                dialog = FieldSettingsDialog(self.field_text, custom_name, self)
                if dialog.exec():
                    new_name = dialog.get_settings()
                    self.lbl.setText(new_name)

class DropListWidget(QListWidget):
    itemsChanged = pyqtSignal()
    
    def __init__(self, is_value_area=False, parent=None):
        super().__init__(parent)
        self.is_value_area = is_value_area
        self.setAcceptDrops(True)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragDropMode.DragDrop)
        self.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.model().rowsInserted.connect(lambda: self.itemsChanged.emit())
        self.model().rowsRemoved.connect(lambda: self.itemsChanged.emit())
        
    def add_card_item(self, text, display_name=None):
        from PyQt6.QtWidgets import QListWidgetItem
        from PyQt6.QtCore import QSize, Qt
        list_item = QListWidgetItem()
        list_item.setData(Qt.ItemDataRole.UserRole, text)
        list_item.setSizeHint(QSize(0, 32))
        self.addItem(list_item)
        
        card = PivotFieldCard(text, self, list_item, self.is_value_area)
        if display_name:
            card.lbl.setText(display_name)
        self.setItemWidget(list_item, card)
        return list_item
        
    def dragEnterEvent(self, event):
        event.acceptProposedAction()
            
    def dragMoveEvent(self, event):
        event.acceptProposedAction()
        
    def dropEvent(self, event):
        source_list = event.source()
        if source_list == self:
            # Internal Reordering logic for drag & drop inside the same box
            current_item = self.currentItem()
            if current_item:
                text = current_item.data(Qt.ItemDataRole.UserRole)
                if not text:
                    text = current_item.text()
                    
                # Determine dropped index based on physical cursor coordinates
                drop_row = self.row(self.itemAt(event.position().toPoint()))
                if drop_row == -1:
                    drop_row = self.count() - 1
                    
                old_row = self.row(current_item)
                if old_row != drop_row and old_row != -1:
                    self.takeItem(old_row)
                    
                    from PyQt6.QtWidgets import QListWidgetItem
                    from PyQt6.QtCore import QSize
                    list_item = QListWidgetItem()
                    list_item.setData(Qt.ItemDataRole.UserRole, text)
                    list_item.setSizeHint(QSize(0, 32))
                    self.insertItem(drop_row, list_item)
                    
                    card = PivotFieldCard(text, self, list_item, self.is_value_area)
                    self.setItemWidget(list_item, card)
                    self.setCurrentItem(list_item)
            event.acceptProposedAction()
            return
            
        if isinstance(source_list, QListWidget):
            for item in source_list.selectedItems():
                text = item.data(Qt.ItemDataRole.UserRole)
                if not text:
                    text = item.text()
                    
                is_move = (source_list != self) and isinstance(source_list, DropListWidget)
                
                exists = False
                for i in range(self.count()):
                    item_data = self.item(i).data(Qt.ItemDataRole.UserRole)
                    if not item_data:
                        item_data = self.item(i).text()
                    if item_data == text:
                        exists = True; break
                        
                if not exists:
                    from PyQt6.QtWidgets import QListWidgetItem
                    from PyQt6.QtCore import QSize
                    list_item = QListWidgetItem()
                    list_item.setData(Qt.ItemDataRole.UserRole, text)
                    list_item.setSizeHint(QSize(0, 32))
                    self.addItem(list_item)
                    card = PivotFieldCard(text, self, list_item, self.is_value_area)
                    self.setItemWidget(list_item, card)
                    
                    if is_move:
                        source_list.takeItem(source_list.row(item))
        event.acceptProposedAction()

class ChartWorker(QThread):
    chart_ready = pyqtSignal(tuple) # (fig, fig_dist, preview_df, overall_df, quarter_df, total_unique, sales_df, full_df, config)
    error = pyqtSignal(str)

    def __init__(self, config):
        super().__init__()
        self.config = config
        self.sales_path = config.get('sales_path', "")
        self.details_path = config.get('details_path', "")
        self.country = config.get('country', "ID")

    def run(self):
        try:
            # Note: We Load & Merge data here to keep it fresh
            df, excluded_df = engine.load_and_merge(self.config['base_folder'], self.config['possys_path'], self.config['store_path'])
            self.excluded_df = excluded_df # Carry to result
            
            set5_mall = int(self.config.get('set5_mall', 10000))
            set5_sa = int(self.config.get('set5_sa', 8000))
            
            # 1. Full Summary (affected by DA/SA but NOT by Dept selection)
            overall_df = engine.generate_consolidated_data(df, int(self.config['da_low']), int(self.config['da_high']), int(self.config['sa_split']), int(self.config.get('da_slicer', 8000)), set5_mall_split=set5_mall, set5_sa_split=set5_sa)
            
            # 2. Filtered Preview (specific to chosen Dept and Set)
            defs = engine.get_defs(df, int(self.config['da_low']), int(self.config['da_high']), int(self.config['sa_split']), int(self.config.get('da_slicer', 8000)), set5_mall_split=set5_mall, set5_sa_split=set5_sa)
            target_set = self.config['target_set']
            target_group = self.config['target_group']
            
            v_depts = sorted(df["M_STORE_DEPARTMENT"].dropna().unique().tolist())
            dg = {}
            for d in v_depts:
                base = engine.get_base_id(d)
                if base not in dg: dg[base] = []
                dg[base].append(d)
            
            # Use case-insensitive lookup to support typed search (e.g. "l881" matches "L881")
            target_group_clean = str(target_group).strip().upper()
            dg_upper = {k.upper(): v for k, v in dg.items()}
            groups_list = [dg_upper[target_group_clean]] if target_group_clean in dg_upper else []
            
            target_set_compare = self.config.get('target_set_compare', [])
            if isinstance(target_set_compare, str):
                target_set_compare = [target_set_compare] if target_set_compare != "None" else []

            # Build ordered list: primary set first, then all valid compare sets
            all_sets = [target_set]
            for cs in target_set_compare:
                if cs != target_set and cs in defs and cs not in all_sets:
                    all_sets.append(cs)

            sets_info = [(s, defs[s]) for s in all_sets if s in defs]
            fig, preview_df = engine.plot_multi_targeted_sets(sets_info, groups_list, df)
            
            # 3. Quarter Analysis Summary
            quarter_df = engine.generate_quarter_data(overall_df)
            
            # 4. Distribution Chart (Now respects selected Report Set)
            fig_dist = engine.plot_display_area_distribution(df, target_set, 
                                                            int(self.config['da_low']), 
                                                            int(self.config['da_high']), 
                                                            int(self.config['sa_split']),
                                                            int(self.config.get('da_slicer', 8000)),
                                                            target_set_compare,
                                                            set5_mall_split=set5_mall,
                                                            set5_sa_split=set5_sa)
            
            # 5. Total Unique Stores (Including Injected Express)
            # Use matched active stores from the Store List sheet if available to align with the active region
            if "Store_Code" in df.columns:
                total_unique = df["Store_Code"].dropna().nunique()
            else:
                total_unique = df["M_STORE"].nunique()
            
            
            # 6. Sales Performance Data (Load & Merge Both Files)
            sales_df = pd.DataFrame()
            s_paths = [p for p in [self.sales_path, self.config.get('sales_new_path')] if p]
            
            s_dfs = []
            for sp in s_paths:
                if os.path.exists(sp):
                    try:
                        # Try finding the right sheet
                        # Target "Sales & Balance (No N Dept)" specifically, fallback to 2nd sheet (index 1)
                        excel_file = pd.ExcelFile(sp, engine='calamine')
                        s_name = "Sales & Balance (No N Dept)"
                        if s_name not in excel_file.sheet_names:
                            if len(excel_file.sheet_names) > 1:
                                s_name = excel_file.sheet_names[1]
                            else:
                                s_name = excel_file.sheet_names[0]
                        
                        temp_df = pd.read_excel(sp, sheet_name=s_name, engine='calamine', dtype=str)
                        temp_df.columns = [str(c).strip() for c in temp_df.columns]
                        
                        # Normalize columns before concat so Existing and New align
                        rename_map = {}
                        for c in temp_df.columns:
                            c_up = c.upper()
                            if c_up in ["SDTL_STORE", "M_STORE", "STORE_CODE"]:
                                rename_map[c] = "STORE"
                            elif c_up in ["SALES_AMT_PER_30DAYS", "SALES_AMT"]:
                                rename_map[c] = "AVG_TOP2_TOTAL_AMT_SALES"
                            elif c_up in ["BALANCE_AMT"]:
                                rename_map[c] = "AVG_TOP2_TOTAL_AMT_BALANCE"
                        if rename_map:
                            temp_df = temp_df.rename(columns=rename_map)
                            
                        # Drop DATECOUNT if it exists (from new store file)
                        if "DATECOUNT" in temp_df.columns:
                            temp_df = temp_df.drop(columns=["DATECOUNT"])
                            
                        # Also drop case-insensitive variations just to be safe
                        cols_to_drop = [c for c in temp_df.columns if c.upper() == "DATECOUNT"]
                        if cols_to_drop:
                            temp_df = temp_df.drop(columns=cols_to_drop)
                            
                        s_dfs.append(temp_df)
                    except Exception as se:
                        print(f"Error loading sales file {sp}: {se}")

            if s_dfs:
                sales_df = pd.concat(s_dfs, ignore_index=True)
                # Clean column names to be safe
                sales_df.columns = [str(c).strip() for c in sales_df.columns]
                
                # 7-8. Enrich sales data (dept type, details, merge, turnover)
                dept_info = df[['M_STORE', 'M_STORE_DEPARTMENT']].drop_duplicates('M_STORE')
                sales_df = engine.enrich_sales_df(sales_df, dept_info, self.details_path, self.country)
                
                # 9. Display Area, Store Type, Set1–Set4 Lookup
                try:
                    net_fallback = os.path.join(DEFAULT_DATA_PATH, "ID store list.xlsx")
                    lookup_path = self.config.get("store_path") if self.config.get("store_path") else net_fallback
                    sales_df = engine.lookup_store_sets(sales_df, lookup_path, self.config)
                except Exception as dae:
                    print(f"Data Lookup Error: {dae}")
                    for c in ["Display Area", "Store Type", "Set1_DA", "Set2_StoreType", "Set3_Type_DA", "Set4_CurrentSetting", "Set5_Mall_SA_Split"]:
                        sales_df[c] = "Error"
                        
                except Exception as e:
                    print(f"Sales Data Processing Error: {e}")

            self.chart_ready.emit((fig, fig_dist, preview_df, overall_df, quarter_df, total_unique, sales_df, df, self.config, self.excluded_df))
        except Exception as e: self.error.emit(str(e))

class ExportWorker(QThread):
    export_finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, config, save_path):
        super().__init__()
        self.config = config
        self.save_path = save_path

    def run(self):
        try:
            df, excluded_df = engine.load_and_merge(self.config['base_folder'], self.config['possys_path'], self.config['store_path'])
            set5_mall = int(self.config.get('set5_mall', 10000))
            set5_sa = int(self.config.get('set5_sa', 8000))
            out_df = engine.generate_consolidated_data(df, int(self.config['da_low']), int(self.config['da_high']), int(self.config['sa_split']), int(self.config.get('da_slicer', 8000)), set5_mall_split=set5_mall, set5_sa_split=set5_sa)
            engine.save_consolidated_excel(out_df, self.save_path)
            self.export_finished.emit(self.save_path)
        except Exception as e: self.error.emit(str(e))

class ChartExportWorker(QThread):
    progress = pyqtSignal(int, int) # (current, total)
    export_finished = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, config, output_dir, export_tasks):
        super().__init__()
        self.config = config
        self.output_dir = output_dir
        self.export_tasks = export_tasks  # list of dicts: [{'code': 'ID', 'name': 'Indonesia', 'possys': '...', 'store': '...'}]

    def run(self):
        try:
            import os
            
            # Count the total number of charts to export across all valid tasks
            total_charts = 0
            task_details = []
            
            for task in self.export_tasks:
                code = task['code']
                name = task['name']
                possys_path = task['possys']
                store_path = task['store']
                
                if not possys_path or not store_path or not os.path.exists(possys_path) or not os.path.exists(store_path):
                    continue
                    
                df, _ = engine.load_and_merge(os.path.dirname(possys_path), possys_path, store_path)
                if df.empty:
                    continue
                    
                v_depts = sorted(df["M_STORE_DEPARTMENT"].dropna().unique().tolist())
                dg = {}
                for d in v_depts:
                    base = engine.get_base_id(d)
                    if base not in dg: dg[base] = []
                    dg[base].append(d)
                    
                groups = sorted(list(dg.keys()))
                
                sets_mapping = {
                    "Set1": "Set1_DA",
                    "Set2": "Set2_StoreType",
                    "Set3": "Set3_Type_DA",
                    "Set4": "Set4_CurrentSetting",
                    "Set5": "Set5_Mall_SA_Split"
                }
                
                task_total = len(sets_mapping) * len(groups)
                total_charts += task_total
                task_details.append({
                    'code': code,
                    'name': name,
                    'df': df,
                    'dg': dg,
                    'groups': groups,
                    'sets_mapping': sets_mapping,
                    'possys': possys_path,
                    'store': store_path
                })
                
            if total_charts == 0:
                self.error.emit("No valid data or files found to export for the chosen countries.")
                return
                
            current = 0
            
            for task in task_details:
                name = task['name']
                df = task['df']
                dg = task['dg']
                groups = task['groups']
                sets_mapping = task['sets_mapping']
                
                set5_mall = int(self.config.get('set5_mall', 10000))
                set5_sa = int(self.config.get('set5_sa', 8000))
                
                defs = engine.get_defs(
                    df, 
                    int(self.config['da_low']), 
                    int(self.config['da_high']), 
                    int(self.config['sa_split']), 
                    int(self.config.get('da_slicer', 8000)),
                    set5_mall_split=set5_mall,
                    set5_sa_split=set5_sa
                )
                
                # Base output dir for this country: <selected_dir>/Output/<CountryName>
                country_dir = os.path.join(self.output_dir, "Output", name)
                os.makedirs(country_dir, exist_ok=True)
                
                for set_folder, set_name in sets_mapping.items():
                    set_path = os.path.join(country_dir, set_folder)
                    os.makedirs(set_path, exist_ok=True)
                    
                    for group_base in groups:
                        filename = f"{set_folder}_{group_base}.jpg"
                        file_path = os.path.join(set_path, filename)
                        
                        sets_info = [(set_name, defs[set_name])]
                        groups_list = [dg[group_base]]
                        
                        fig, _ = engine.plot_multi_targeted_sets(sets_info, groups_list, df)
                        fig.savefig(file_path, format='jpg', dpi=120)
                        fig.clear()
                        
                        current += 1
                        self.progress.emit(current, total_charts)
                        
            self.export_finished.emit()
        except Exception as e:
            self.error.emit(str(e))

# ============================================================
# MAIN WINDOW
# ============================================================
class StoreChartApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"Store Splitting Tool {history.VERSION_HISTORY[0][0]}")
        self.resize(1600, 900)
        self.setMinimumSize(1200, 800)
        self.setStyleSheet(STYLESHEET)
        self.setWindowIcon(QIcon(resource_path("coa_team_logo.ico")))
        self.possys_path = ""
        self.store_path = ""; self.sales_path = ""; self.sales_new_path = ""
        
        # Determine country and set Google Sheet URL for Department Details
        self.country = "ID"
        workspace_path = os.path.abspath(__file__).upper()
        # Split path by folder separators to check for exact country code matches
        path_parts = re.split(r'[\\/]', workspace_path)
        if "THAILAND" in path_parts or "TH" in path_parts:
            self.country = "TH"

        g_sheets = {
            "BR": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=856910923",
            "ID": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=1592927868",
            "IN": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=0", # Placeholder GID for India (data given later)
            "MY": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=1298692130",
            "SG": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=856910923",
            "TH": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=1506559356",
        }
        self.details_url = g_sheets.get(self.country, g_sheets["ID"])
        self.details_path = f"{self.details_url}|Y:\\R&D\\Category-based Start Up - Split DB Analysis\\Tool Data\\Department Details.xlsx"
        
        self.canvas = None
        self.scanner = None
        self.worker = None
        self.exporter = None
        self._retiring_threads = []  # Threads being retired — kept alive until finished
        
        # --- DEEP HOT-RELOAD WATCHER ---
        from PyQt6.QtCore import QFileSystemWatcher
        self.style_watcher = QFileSystemWatcher([os.path.abspath(__file__)])
        self.style_watcher.fileChanged.connect(self.hot_reload_full)
        # ------------------------------
        # Hot-reload triggered comment - Small dots in DA Analysis
        self.active_filters = {} # Dictionary mapping column index to filter text
        
        try:
            self.auto_locate_network_files(start_scanner=False)
        except Exception as e:
            print(f"[Network/Startup Warning] Failed to auto locate network files: {e}")
        self.init_ui()
        
        # Start scanning if file was auto-detected
        try:
            if self.possys_path:
                # Wrap os.path.exists in case it throws a network error
                if os.path.exists(self.possys_path):
                    self._discard_thread('scanner')
                    self.status_msg.setText(f"Scanning POSSYS departments for {self.country}...")
                    self.scanner = ScanWorker(self.possys_path)
                    self.scanner.scan_finished.connect(self.on_scan_finished)
                    self.scanner.error.connect(self.on_error)
                    self.scanner.start()
        except Exception as e:
            print(f"[Network/Startup Warning] Failed to start scanner at startup: {e}")

    def auto_locate_network_files(self, start_scanner=False):
        import glob
        try:
            # Specific lookup for POSSYS / Grouping File on network drive (Dynamic Network Priority)
            drives = [r"Z:\\", r"Y:\\", r"X:\\", r"U:\\", r"T:\\"]
            rd_root = None
            for drive in drives:
                try:
                    candidate = os.path.join(drive, "R&D", "Category-based Start Up - Split DB Analysis")
                    if os.path.exists(candidate):
                        rd_root = candidate
                        break
                except Exception as e:
                    print(f"[Network/IO Warning] Could not scan drive {drive}: {e}")
                    continue

            paths_assigned = False
            tool_data_dir = None
            if rd_root and os.path.exists(rd_root):
                tool_data_dir = os.path.join(rd_root, "Tool Data")
            


            if tool_data_dir and os.path.exists(tool_data_dir):
                if True:
                    def find_latest_in_subfolder(subfolder, pattern):
                        try:
                            folder_path = os.path.join(tool_data_dir, subfolder)
                            if os.path.exists(folder_path):
                                matches = glob.glob(os.path.join(folder_path, pattern))
                                if matches:
                                    # Filter out any files that fail with OSError (e.g. WinError 59)
                                    valid_matches = []
                                    for m in matches:
                                        try:
                                            # Trigger getmtime to verify accessibility
                                            os.path.getmtime(m)
                                            valid_matches.append(m)
                                        except Exception as ex:
                                            print(f"[Network/IO Warning] Skipping file {m} due to error: {ex}")
                                    if valid_matches:
                                        return max(valid_matches, key=os.path.getmtime)
                        except Exception as e:
                            print(f"[Network/IO Warning] Error searching subfolder '{subfolder}' with pattern '{pattern}': {e}")
                        return None

                    p_store = find_latest_in_subfolder("Store List", f"{self.country} - Store List*.xlsx")
                    if not p_store:
                        p_store = find_latest_in_subfolder("Store List", f"*{self.country}*Store List*.xlsx")

                    p_sb_existing = find_latest_in_subfolder("Sales & Balance", f"{self.country} - S&B*Existing*.xlsx")
                    if not p_sb_existing:
                        p_sb_existing = find_latest_in_subfolder("Sales & Balance", f"*{self.country}*S&B*Existing*.xlsx")

                    p_sb_new = find_latest_in_subfolder("Sales & Balance", f"{self.country} - S&B*New*.xlsx")
                    if not p_sb_new:
                        p_sb_new = find_latest_in_subfolder("Sales & Balance", f"*{self.country}*S&B*New*.xlsx")

                    p_grouping = find_latest_in_subfolder("Grouping", f"{self.country} - Grouping*.xlsx")
                    if not p_grouping:
                        p_grouping = find_latest_in_subfolder("Grouping", f"*{self.country}*Grouping*.xlsx")
                    
                    if p_store and p_sb_existing and p_sb_new and p_grouping:
                        self.store_path = p_store
                        self.sales_path = p_sb_existing
                        self.sales_new_path = p_sb_new
                        self.possys_path = p_grouping
                        paths_assigned = True
                        
                        p_details = find_latest_in_subfolder("Grouping", "Department Details*.xlsx")
                        if not p_details:
                            p_details = find_latest_in_subfolder("Store List", "Department Details*.xlsx")
                        if not p_details:
                            downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
                            if os.path.exists(downloads_dir):
                                import glob
                                dl_files = glob.glob(os.path.join(downloads_dir, "Department Details*.xlsx"))
                                if dl_files:
                                    dl_files.sort(key=os.path.getmtime, reverse=True)
                                    p_details = dl_files[0]
                        if p_details:
                            self.details_path = f"{self.details_url}|{p_details}|Y:\\R&D\\Category-based Start Up - Split DB Analysis\\Tool Data\\Department Details.xlsx"
                        else:
                            self.details_path = f"{self.details_url}|Y:\\R&D\\Category-based Start Up - Split DB Analysis\\Tool Data\\Department Details.xlsx"

            if start_scanner and self.possys_path:
                try:
                    if os.path.exists(self.possys_path):
                        self._discard_thread('scanner')
                        self.status_msg.setText(f"Scanning POSSYS departments for {self.country}...")
                        self.scanner = ScanWorker(self.possys_path)
                        self.scanner.scan_finished.connect(self.on_scan_finished)
                        self.scanner.error.connect(self.on_error)
                        self.scanner.start()
                except Exception as e:
                    print(f"[Network/IO Warning] Could not start scanner for {self.possys_path}: {e}")
        except Exception as e:
            print(f"[Network/IO Error] auto_locate_network_files failed completely: {e}")
        
    def _discover_and_normalize_store_sheet(self, path):
        import pandas as pd
        import numpy as np
        
        xl = pd.ExcelFile(path, engine='calamine')
        sidf = None
        
        # Look for the sheet that actually contains store list data
        for s_name in xl.sheet_names:
            # Try the first 2 rows as potential headers
            for h_test in [0, 1]:
                temp_df = xl.parse(s_name, header=h_test, nrows=5, dtype=str)
                temp_df.columns = [str(c).lower() for c in temp_df.columns]
                # Match "store_code", "store code", "display area", "sqm", "da" in lower case
                if any(k in " ".join(temp_df.columns) for k in ["store_code", "store code", "display area", "sqm", "da"]):
                    sidf = xl.parse(s_name, header=h_test, dtype=str)
                    print(f"[UI] Found Store Data in sheet: '{s_name}' at Row {h_test+1}")
                    break
            if sidf is not None:
                break
                
        # Fallback 1: sheet named ID STORES SHORT NAME (backward compatibility)
        if sidf is None:
            if "ID STORES SHORT NAME" in xl.sheet_names:
                sidf = xl.parse("ID STORES SHORT NAME", header=1, dtype=str)
                
        # Fallback 2: first sheet if nothing found
        if sidf is None:
            sidf = xl.parse(xl.sheet_names[0], dtype=str)
            
        sidf.columns = [str(c).strip() for c in sidf.columns]
        
        def find_sid(kw, default):
            for c in sidf.columns:
                if all(k.lower() in c.lower() for k in kw):
                    return c
            return default
            
        st_code_col    = find_sid(["store", "code"],    None) or find_sid(["code"], "Store_Code")
        st_name_col    = find_sid(["store", "name"],    None) or find_sid(["system", "name"], None) or find_sid(["name"], "Store_Name")
        st_type_col    = find_sid(["store", "type"],    None) or find_sid(["lot", "type"], None) or "Store_Type"
        st_da_col      = find_sid(["display", "area"], None) or find_sid(["sqm"], None) or find_sid(["da"], "Store_Display_Area")
        st_concept_col = find_sid(["concept"],          "Store_Concept")
        st_island_col  = find_sid(["store", "island", "status"], None) or find_sid(["island", "status"], None) or find_sid(["island"], "Store_Island_Status")
        st_date_col    = find_sid(["opening", "date"],  None) or find_sid(["start", "business"], None) or find_sid(["date"], "Start_Business_Date")
        st_country_col = find_sid(["country"],          "Country_Code")
        
        rename_dict = {}
        if st_code_col and st_code_col in sidf.columns:
            rename_dict[st_code_col] = "Store_Code"
        if st_name_col and st_name_col in sidf.columns:
            rename_dict[st_name_col] = "Store_Name"
        if st_da_col and st_da_col in sidf.columns:
            rename_dict[st_da_col] = "Store_Display_Area"
        if st_type_col and st_type_col in sidf.columns:
            rename_dict[st_type_col] = "Store_Type"
        if st_concept_col and st_concept_col in sidf.columns:
            rename_dict[st_concept_col] = "Store_Concept"
        if st_island_col and st_island_col in sidf.columns:
            rename_dict[st_island_col] = "Store_Island_Status"
        if st_date_col and st_date_col in sidf.columns:
            rename_dict[st_date_col] = "Start_Business_Date"
        if st_country_col and st_country_col in sidf.columns:
            rename_dict[st_country_col] = "Country_Code"
            
        sidf = sidf.rename(columns=rename_dict)
        
        # De-duplicate any duplicate columns to ensure 100% robustness (e.g. against multiple 'Store_Island_Status' fields)
        cols = []
        counts = {}
        for col in sidf.columns:
            if col in counts:
                counts[col] += 1
                cols.append(f"{col}_{counts[col]}")
            else:
                counts[col] = 0
                cols.append(col)
        sidf.columns = cols
        
        # Ensure standard columns are present
        if "Store_Type" not in sidf.columns:
            sidf["Store_Type"] = "Mall"
        else:
            is_sa = sidf["Store_Type"].astype(str).str.strip().str.upper().isin(["SALONE", "STANDALONE", "SA"])
            sidf["Store_Type"] = np.where(is_sa, "Standalone", "Mall")
            
        if "Store_Concept" not in sidf.columns:
            sidf["Store_Concept"] = "Mr_DIY"
            
        if "Store_Island_Status" not in sidf.columns:
            sidf["Store_Island_Status"] = "Unknown"
        else:
            sidf["Store_Island_Status"] = sidf["Store_Island_Status"].fillna("Unknown")
            
        if "Country_Code" not in sidf.columns:
            sidf["Country_Code"] = self.country
            
        if "Start_Business_Date" not in sidf.columns:
            sidf["Start_Business_Date"] = "NEW STORE"
            
        is_th = self.country == "TH" or "TH" in str(path).upper()
        if "Store _Dummy_Type" not in sidf.columns:
            sidf["Store _Dummy_Type"] = "Department based" if is_th else "Unknown"
        else:
            if is_th:
                sidf["Store _Dummy_Type"] = sidf["Store _Dummy_Type"].fillna("Department based")
                sidf.loc[sidf["Store _Dummy_Type"].astype(str).str.strip().str.lower().isin(["nan", "", "unknown", "none", "n/a"]), "Store _Dummy_Type"] = "Department based"
                
        for col in ["Store_Code", "Store_Name", "Store_Display_Area", "Store_Type", "Country_Code"]:
            if col in sidf.columns:
                sidf[col] = sidf[col].fillna("")
                
        # Remove "NO" column if it exists
        cols_to_drop = [c for c in sidf.columns if str(c).strip().upper() in ["NO", "NO.", "NO "]]
        if cols_to_drop:
            sidf = sidf.drop(columns=cols_to_drop)
            
        # Reorder: Move Country_Code to the right of Store_Name
        cols = list(sidf.columns)
        if "Country_Code" in cols and "Store_Name" in cols:
            cols.remove("Country_Code")
            idx = cols.index("Store_Name")
            cols.insert(idx + 1, "Country_Code")
            sidf = sidf[cols]
            
        return sidf

    def _discard_thread(self, attr):
        """
        Safely retire a running QThread without blocking the UI or letting it get
        garbage-collected while still running (which triggers 'QThread: Destroyed
        while thread is still running' and crashes the process).

        Strategy:
          - Disconnect all signals so stale results don't fire on the new UI.
          - Park the thread in a retirement pool (self._retiring_threads) so
            Python keeps a reference to it until the C++ thread finishes.
          - Connect finished -> _on_thread_retired to clean up when it's done.
          - Call quit() as a polite hint (only useful if the thread has an event loop).
        """
        t = getattr(self, attr, None)
        setattr(self, attr, None)
        if t is None:
            return
        try:
            running = t.isRunning()
        except RuntimeError:
            return  # C++ object already gone — nothing to do

        if not running:
            try:
                t.deleteLater()
            except RuntimeError:
                pass
            return

        # Disconnect all signals to silence any pending results
        for sig_name in ('finished', 'error'):
            try:
                sig = getattr(t, sig_name, None)
                if sig is not None:
                    sig.disconnect()
            except (RuntimeError, TypeError):
                pass

        # Park in retirement pool — keeps Python ref alive until thread is done
        if not hasattr(self, '_retiring_threads'):
            self._retiring_threads = []
        self._retiring_threads.append(t)
        try:
            t.finished.connect(lambda *args, th=t: self._on_thread_retired(th))
        except RuntimeError:
            pass
        try:
            t.quit()  # Polite stop (no-op for threads without an event loop)
        except RuntimeError:
            pass

    def _on_thread_retired(self, t):
        """Called when a retired thread finishes — remove from pool and delete safely."""
        try:
            self._retiring_threads.remove(t)
        except (ValueError, AttributeError):
            pass
        try:
            t.deleteLater()
        except RuntimeError:
            pass

    def hot_reload_full(self):
        """Re-reads the entire file and re-builds the UI inside the same window."""
        try:
            import re
            # 1. Read the fresh code
            with open(__file__, 'r', encoding='utf-8') as f:
                code = f.read()

            # 2. Update styles instantly
            match = re.search(r'STYLESHEET = """(.*?)"""', code, re.DOTALL)
            if match: self.setStyleSheet(match.group(1))

            # 3. Retire all running threads BEFORE tearing down the UI.
            #    This disconnects their signals so they can't fire on stale widgets.
            for attr in ('scanner', 'worker', 'exporter'):
                self._discard_thread(attr)

            # 4. Wipe and Rebuild UI
            # We use a temporary namespace to extract the new init_ui method
            namespace = {}
            # Prevent the re-executed code from starting a second QApplication
            # We use concatenation to prevent this line from matching itself during the split
            split_trigger = 'if __name__ == ' + '"__main__":'
            safe_code = code.split(split_trigger)[0]
            exec(safe_code, globals(), namespace)
            
            new_class = namespace.get('StoreChartApp')
            if new_class:
                # Replace ALL methods from the new class definition to ensure signatures match
                for name, attr in new_class.__dict__.items():
                    if callable(attr) and not name.startswith("__"):
                        setattr(self, name, attr.__get__(self, self.__class__))
                
                # Clear and Re-init
                if self.centralWidget():
                    self.centralWidget().setParent(None)
                self.init_ui()
                print(f"🚀 Deep Reload Sync: {time.strftime('%H:%M:%S')}")
                
        except Exception as e:
            print(f"⚠️ Reload Error (Fix your syntax!): {e}")

    def reload_styles(self):
        # Kept for backward compatibility if needed
        self.hot_reload_full()

    def closeEvent(self, event):
        # Retire all active workers (non-blocking)
        for attr in ('scanner', 'worker', 'exporter'):
            self._discard_thread(attr)
        # Give retiring threads a brief moment to finish (max 500ms each)
        for t in list(getattr(self, '_retiring_threads', [])):
            try:
                t.wait(500)
            except RuntimeError:
                pass
        event.accept()

    def on_country_changed(self):
        selected = self.country_combo.currentText().strip()
        if selected not in ["ID", "TH"]:
            # If somehow a disabled/unsupported country is selected (e.g. keyboard navigation), revert to previous valid selection or "ID"
            valid_fallback = self.country if self.country in ["ID", "TH"] else "ID"
            idx = self.country_combo.findText(valid_fallback)
            if idx >= 0:
                self.country_combo.blockSignals(True)
                self.country_combo.setCurrentIndex(idx)
                self.country_combo.blockSignals(False)
            return
            
        self.country = selected
        
        if hasattr(self, 'combo_dept') and self.combo_dept:
            self.previous_selected_dept = self.combo_dept.currentText().strip()
            
        g_sheets = {
            "BR": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=856910923",
            "ID": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=1592927868",
            "IN": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=0", # Placeholder GID for India
            "MY": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=1298692130",
            "SG": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=856910923",
            "TH": "https://docs.google.com/spreadsheets/d/18x0wd4tVRPwpUbjpLog0VeNaQRm6ZyEzCmEjFUxeplE/export?format=csv&gid=1506559356",
        }
        self.details_url = g_sheets.get(self.country, g_sheets["ID"])
        
        # Re-locate the network files using the new country selection
        self.auto_locate_network_files(start_scanner=True)
        
        if hasattr(self, 'readme_lbl') and self.readme_lbl:
            import re
            html = self.readme_lbl.text()
            country_names = {
                "BR": "Brunei",
                "ID": "Indonesia",
                "IN": "India",
                "MY": "Malaysia",
                "SG": "Singapore",
                "TH": "Thailand"
            }
            c_name = country_names.get(self.country, self.country)
            html = re.sub(r'<b>Active Analysis:</b> .*?</div>', f'<b>Active Analysis:</b> {c_name}</div>', html)
            self.readme_lbl.setText(html)
            
        print(f"[UI] Country changed to {self.country}. Details URL updated to: {self.details_url}")
        self.status_msg.setText(f"Country switched to {self.country}. Auto-updating data...")

    def init_ui(self):
        central = QWidget(); self.setCentralWidget(central)
        layout = QVBoxLayout(central); layout.setContentsMargins(30, 20, 30, 20); layout.setSpacing(15)

        # HEADER
        header_h = QHBoxLayout()
        header_v = QVBoxLayout()
        lbl_title = QLabel("Store Splitting Tool"); lbl_title.setObjectName("HeaderLabel")
        header_v.addWidget(lbl_title)
        header_h.addLayout(header_v)

        # Progress bar in the middle (Long bar)
        self.progress = QProgressBar(); self.progress.setVisible(False)
        self.progress.setTextVisible(True)
        header_h.addSpacing(30)
        header_h.addWidget(self.progress, 1)
        header_h.addSpacing(30)

        # Actions in Header
        actions_h = QHBoxLayout()
        
        # Country Dropdown Selector (Sorted alphabetically by 2-letter acronym)
        self.country_combo = QComboBox()
        self.country_combo.setObjectName("CountryCombo")
        countries = ["ID", "BR", "IN", "MY", "SG", "TH"]
        self.country_combo.addItems(countries)
        self.country_combo.setFixedSize(65, 40)
        
        # Only enable active profiles (ID, TH). Disable and fade (reduce opacity) the rest.
        for i in range(self.country_combo.count()):
            val = self.country_combo.itemText(i)
            if val not in ["ID", "TH"]:
                item = self.country_combo.model().item(i)
                if item:
                    item.setEnabled(False)
                    item.setForeground(QBrush(QColor("#cbd5e1")))
        
        default_index = countries.index(self.country) if self.country in countries else 0
        self.country_combo.setCurrentIndex(default_index)
        self.country_combo.currentIndexChanged.connect(self.on_country_changed)
        
        self.btn_run = QPushButton("UPDATE"); self.btn_run.setFixedSize(150, 38)
        self.btn_run.clicked.connect(self.update_chart)
        
        self.btn_export = QPushButton("EXPORT"); self.btn_export.setObjectName("ExportBtn")
        self.btn_export.setFixedSize(160, 40)
        
        self.export_menu = QMenu(self)
        self.action_master = QAction("Master.xlsx", self)
        self.action_charts = QAction("Charts.jpg", self)
        self.action_overall = QAction("Overall Summary.xlsx", self)
        self.action_all_summary = QAction("All Summary.xlsx", self)
        self.action_perf = QAction("Performance Analysis.xlsx", self)
        self.action_sales = QAction("Sales Performance.xlsx", self)
        self.action_store_list = QAction("Store List.xlsx", self)
        self.action_grouping_analysis = QAction("Grouping Analysis.xlsx", self)
        self.action_grouping_summary = QAction("Grouping Summary.xlsx", self)
        
        self.export_menu.addAction(self.action_master)
        self.export_menu.addAction(self.action_charts)
        # self.export_menu.addAction(self.action_overall)
        # self.export_menu.addAction(self.action_all_summary)
        # self.export_menu.addAction(self.action_perf)
        # self.export_menu.addAction(self.action_sales)
        # self.export_menu.addAction(self.action_store_list)
        # self.export_menu.addAction(self.action_grouping_analysis)
        # self.export_menu.addAction(self.action_grouping_summary)
        self.btn_export.setMenu(self.export_menu)
        
        self.action_master.triggered.connect(self.export_master)
        self.action_charts.triggered.connect(self.export_charts)
        self.action_overall.triggered.connect(lambda: self.export_csv("Overall Summary"))
        self.action_all_summary.triggered.connect(lambda: self.export_csv("All Summary"))
        self.action_perf.triggered.connect(lambda: self.export_csv("Performance Analysis"))
        self.action_sales.triggered.connect(lambda: self.export_csv("Sales Performance"))
        self.action_store_list.triggered.connect(lambda: self.export_csv("Store List"))
        self.action_grouping_analysis.triggered.connect(self.export_grouping_analysis)
        self.action_grouping_summary.triggered.connect(self.export_grouping_summary)
        
        actions_h.addWidget(self.country_combo)
        actions_h.addWidget(self.btn_run)
        actions_h.addWidget(self.btn_export)
        header_h.addLayout(actions_h)
        layout.addLayout(header_h)

        # CONFIG PANEL
        config_frame = QFrame(); config_frame.setObjectName("ControlPanel")
        config_layout = QVBoxLayout(config_frame); config_layout.setContentsMargins(20, 20, 20, 20); config_layout.setSpacing(15)
        
        # Configuration Bar
        params_h = QHBoxLayout()
        
        h_set = QHBoxLayout(); lbl_set = QLabel("Reporting Sets"); lbl_set.setFixedWidth(110)
        lbl_set.setStyleSheet("font-size: 11px; color: #64748b;")
        self.combo_set = LimitedCheckableComboBox(max_choices=4)
        self.combo_set.add_checkable_item("Set1_Display Area", "Set1_DA", checked=True)
        self.combo_set.add_checkable_item("Set2_Store Type", "Set2_StoreType", checked=False)
        self.combo_set.add_checkable_item("Set3_SA Split", "Set3_Type_DA", checked=False)
        self.combo_set.add_checkable_item("Set4_DA Split", "Set4_CurrentSetting", checked=False)
        self.combo_set.add_checkable_item("Set5_Mall & SA Split", "Set5_Mall_SA_Split", checked=False)
        h_set.addWidget(lbl_set); h_set.addWidget(self.combo_set)
        params_h.addLayout(h_set, 1)

        h_dept = QHBoxLayout(); lbl_dept = QLabel("Department"); lbl_dept.setFixedWidth(130)
        lbl_dept.setStyleSheet("font-size: 11px; color: #64748b;")
        self.combo_dept = QComboBox()
        self.combo_dept.setEditable(True)
        self.combo_dept.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.combo_dept.completer().setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        self.combo_dept.completer().setFilterMode(Qt.MatchFlag.MatchContains)
        # FORCE LIGHT THEME ON POPUP
        self.combo_dept.view().setStyleSheet("background-color: white; color: black; selection-background-color: #eff6ff; selection-color: #2563eb;")
        self.combo_dept.completer().popup().setStyleSheet("background-color: white; color: black; selection-background-color: #eff6ff; selection-color: #2563eb;")
        h_dept.addWidget(lbl_dept); h_dept.addWidget(self.combo_dept)
        params_h.addLayout(h_dept, 1)

        self.v_da_low = self.create_input("[Set1] DA Lower", "7500")
        self.v_da_high = self.create_input("[Set1] DA Upper", "9500")
        self.v_sa_split = self.create_input("[Set3] SA Slicer", "8000")
        self.v_da_slicer = self.create_input("[Set4] DA Slicer", "8000")
        self.v_set5_mall = self.create_input("[Set5] Mall Slicer", "10000")
        self.v_set5_sa = self.create_input("[Set5] SA Slicer", "8000")

        # ENTER-KEY SHORTCUTS
        self.v_da_low.input.returnPressed.connect(self.update_chart)
        self.v_da_high.input.returnPressed.connect(self.update_chart)
        self.v_sa_split.input.returnPressed.connect(self.update_chart)
        self.v_da_slicer.input.returnPressed.connect(self.update_chart)
        self.v_set5_mall.input.returnPressed.connect(self.update_chart)
        self.v_set5_sa.input.returnPressed.connect(self.update_chart)
        self.combo_dept.lineEdit().returnPressed.connect(self.update_chart)
        
        params_h.addLayout(self.v_da_low, 1)
        params_h.addLayout(self.v_da_high, 1)
        params_h.addLayout(self.v_sa_split, 1)
        params_h.addLayout(self.v_da_slicer, 1)
        params_h.addLayout(self.v_set5_mall, 1)
        params_h.addLayout(self.v_set5_sa, 1)
        
        # Connect visibility updates AFTER all widgets exist
        self.combo_set.currentIndexChanged.connect(self.on_primary_set_changed)

        config_layout.addLayout(params_h)
        layout.addWidget(config_frame)

        # PROGRESS AREA placeholder (empty)
        action_h = QHBoxLayout()
        layout.addLayout(action_h)

        # TABS AREA
        from PyQt6.QtWidgets import QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView, QScrollArea
        self.tabs = QTabWidget()
        
        # TAB 0: README
        self.readme_container = QWidget()
        readme_main_layout = QHBoxLayout(self.readme_container)
        readme_main_layout.setContentsMargins(0, 0, 0, 0)
        readme_main_layout.setSpacing(0)
        
        # --- LEFT PANEL: Guide & Overview ---
        left_panel = QWidget()
        left_panel.setStyleSheet("background-color: white; border-right: 1px solid #e2e8f0;")
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(30, 30, 30, 30)
        left_layout.setSpacing(15)
        
        readme_title = QLabel("Store Splitting Tool")
        readme_title.setStyleSheet("font-size: 22px; font-weight: bold; color: #0f172a;")
        left_layout.addWidget(readme_title)
        
        country_names = {
            "ID": "Indonesia",
            "BR": "Brunei",
            "IN": "India",
            "MY": "Malaysia",
            "SG": "Singapore",
            "TH": "Thailand"
        }
        active_country_name = country_names.get(self.country, self.country)

        decision_html = ""
        if self.country == "ID":
            decision_html = """
            <div style='background-color: #f0fdf4; border: 1px solid #bbf7d0; border-left: 4px solid #22c55e; padding: 12px 15px; border-radius: 6px; margin-bottom: 20px;'>
                <strong style='color: #166534; font-size: 14px;'>DECISION</strong>
                <p style='color: #15803d; margin: 6px 0; line-height: 1.4; font-size: 13px;'>
                    For ID, Set 4 is preferrable to use due to these reasons:
                </p>
                <ol style='color: #15803d; margin-top: 6px; font-size: 13px; padding-left: 20px;'>
                    <li>Good grouping distribution</li>
                    <li>Display area range is good to cover for each split.</li>
                </ol>
            </div>
            """

        readme_text = f"""
        <div style='color: #334155;'>
            {decision_html}
            <!-- Country Compatibility Announcement -->
            <div style='background-color: #eff6ff; border: 1px solid #dbeafe; border-left: 4px solid #3b82f6; padding: 12px 15px; border-radius: 6px; margin-bottom: 20px;'>
                <strong style='color: #1e40af; font-size: 14px;'>IMPORTANT</strong>
                <p style='color: #1e3a8a; margin: 6px 0; line-height: 1.4; font-size: 13px;'>
                    Currently, Indonesia (ID) and Thailand (TH) analysis are available. Other countries are still under development.
                </p>
            </div>

            <h3 style='color: #0f172a;'>Overview</h3>
            <p>Welcome to the Store Splitting Tool! This tool analyzes and classifies retail locations into distinct sets for optimized logistical tracking.</p>
            
            <h3 style='color: #0f172a;'>Quick Guide to Reporting Sets</h3>
            <p>Use the upper bounds configuration widgets to change criteria on the fly:</p>
            <ul>
                <li><b>Set1_Display Area:</b> Segments stores by size (e.g., under or over default DA bounds).</li>
                <li><b>Set2_Store Type:</b> Segments by operational format (Mall vs Standalone).</li>
                <li><b>Set3_SA Split:</b> Adds an intermediate breakdown for Standalone locations.</li>
                <li><b>Set4_DA Split:</b> Splits standard distributions aggressively via DA Slicers.</li>
            </ul>
            
            <h3 style='color: #0f172a;'>Features</h3>
            <ul>
                <li><b>Dynamic Comparison:</b> Compare groupings directly side by side.</li>
                <li><b>Analytics Visuals:</b> Real-time cross-quarter trend visualization.</li>
                <li><b>Performance Analysis:</b> High-performance vectorized computations.</li>
                <li><b>Custom Workspaces:</b> Drag-and-drop pivot tables with save/load capability.</li>
            </ul>
        </div>
        """
        self.readme_lbl = QLabel(readme_text)
        self.readme_lbl.setWordWrap(True)
        self.readme_lbl.setStyleSheet("font-size: 13px; line-height: 1.6;")
        left_layout.addWidget(self.readme_lbl)
        left_layout.addStretch()
        
        # --- RIGHT PANEL: Version History (Scrollable) ---
        right_panel_container = QWidget()
        right_panel_container.setStyleSheet("background-color: #f8fafc;")
        right_panel_layout = QVBoxLayout(right_panel_container)
        right_panel_layout.setContentsMargins(25, 30, 25, 30)
        right_panel_layout.setSpacing(10)
        
        version_title_container = QWidget()
        version_title_layout = QVBoxLayout(version_title_container)
        version_title_layout.setContentsMargins(0, 0, 0, 10)
        
        version_title = QLabel("Version History")
        version_title.setStyleSheet("font-size: 18px; font-weight: bold; color: #0f172a;")
        version_title_layout.addWidget(version_title)
        
        divider = QFrame()
        divider.setFrameShape(QFrame.Shape.HLine)
        divider.setStyleSheet("background-color: #e2e8f0; max-height: 1px; border: none;")
        version_title_layout.addWidget(divider)
        right_panel_layout.addWidget(version_title_container)
        
        # Independent Scroll Area for Version Entries
        history_scroll = QScrollArea()
        history_scroll.setWidgetResizable(True)
        history_scroll.setStyleSheet("QScrollArea { border: none; background-color: transparent; }")
        
        history_content = QWidget()
        history_content.setStyleSheet("background-color: transparent;")
        history_layout = QVBoxLayout(history_content)
        history_layout.setContentsMargins(0, 0, 10, 0)
        history_layout.setSpacing(0)
        
        version_history = history.VERSION_HISTORY
        
        for version, date, changes in version_history:
            entry_frame = QFrame()
            entry_frame.setStyleSheet("QFrame { border: none; border-bottom: 1px solid #e2e8f0; background-color: transparent; }")
            entry_layout = QVBoxLayout(entry_frame)
            entry_layout.setContentsMargins(0, 12, 0, 12)
            entry_layout.setSpacing(6)
            
            header_row = QHBoxLayout()
            ver_lbl = QLabel(version)
            ver_lbl.setStyleSheet("font-size: 14px; font-weight: bold; color: #0f172a;")
            date_lbl = QLabel(date)
            date_lbl.setStyleSheet("font-size: 12px; color: #64748b;")
            header_row.addWidget(ver_lbl)
            header_row.addStretch()
            header_row.addWidget(date_lbl)
            entry_layout.addLayout(header_row)
            
            for change in changes:
                c_lbl = QLabel(f"• {change}")
                c_lbl.setWordWrap(True)
                c_lbl.setStyleSheet("font-size: 12px; color: #475569; padding-left: 5px;")
                entry_layout.addWidget(c_lbl)
            
            history_layout.addWidget(entry_frame)
            
        history_layout.addStretch()
        history_scroll.setWidget(history_content)
        right_panel_layout.addWidget(history_scroll)
        
        readme_main_layout.addWidget(left_panel, 6)   # Left takes 60%
        readme_main_layout.addWidget(right_panel_container, 4) # Right takes 40%
        
        readme_scroll = QScrollArea()
        readme_scroll.setWidgetResizable(True)
        readme_scroll.setWidget(self.readme_container)
        readme_scroll.setStyleSheet("border: none;")
        



        # TAB 1: CHART
        self.chart_container = QWidget()
        self.chart_layout = QVBoxLayout(self.chart_container)
        self.chart_layout.setContentsMargins(10, 10, 10, 40)
        self.chart_layout.addWidget(QLabel("Chart preview will appear here...", alignment=Qt.AlignmentFlag.AlignCenter))

        chart_scroll = QScrollArea()
        chart_scroll.setWidgetResizable(True)
        chart_scroll.setWidget(self.chart_container)
        chart_scroll.setStyleSheet("border: none;")
        
        # TAB 2: DEPT SUMMARY
        self.summary_container = QWidget()
        summary_vbox = QVBoxLayout(self.summary_container)
        summary_vbox.setContentsMargins(15, 15, 15, 15); summary_vbox.setSpacing(10)
        
        self.table_summary = QTableWidget()
        self.table_summary.setAlternatingRowColors(True)
        self.table_summary.setSortingEnabled(False)
        self.table_summary.verticalHeader().setDefaultSectionSize(30)
        self.table_summary.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_summary.cellClicked.connect(lambda r, c: self.on_cell_clicked(self.table_summary, r, c))
        summary_vbox.addWidget(self.table_summary)
        
        # TAB 3: OVERALL SUMMARY
        self.overall_container = QWidget()
        overall_vbox = QVBoxLayout(self.overall_container)
        overall_vbox.setContentsMargins(15, 15, 15, 15); overall_vbox.setSpacing(10)
        
        self.table_overall = QTableWidget()
        self.table_overall.setAlternatingRowColors(True)
        self.table_overall.setSortingEnabled(False)
        self.table_overall.verticalHeader().setDefaultSectionSize(30)
        self.table_overall.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_overall.horizontalHeader().sectionClicked.connect(lambda idx: self.on_header_clicked(self.table_overall, idx))
        self.table_overall.cellClicked.connect(lambda r, c: self.on_cell_clicked(self.table_overall, r, c))
        overall_vbox.addWidget(self.table_overall)
        
        # TAB 4: QUARTER ANALYSIS (DISTRIBUTION)
        self.quarter_container = QWidget()
        self.quarter_layout = QVBoxLayout(self.quarter_container)
        self.quarter_layout.setContentsMargins(10, 10, 10, 10)
        self.quarter_layout.addWidget(QLabel("Distribution chart will appear here...", alignment=Qt.AlignmentFlag.AlignCenter))
        
        quarter_scroll = QScrollArea()
        quarter_scroll.setWidgetResizable(True)
        quarter_scroll.setWidget(self.quarter_container)
        quarter_scroll.setStyleSheet("border: none;")
        
        # TAB: HEALTHY MINMAX
        self.healthy_minmax_container = QWidget()
        healthy_minmax_vbox = QVBoxLayout(self.healthy_minmax_container)
        healthy_minmax_vbox.setContentsMargins(15, 15, 15, 15); healthy_minmax_vbox.setSpacing(10)
               
        self.table_healthy = QTableWidget()
        self.table_healthy.setAlternatingRowColors(True)
        self.table_healthy.setStyleSheet("background-color: white; border: 1px solid #e2e8f0; border-radius: 4px;")
        healthy_minmax_vbox.addWidget(self.table_healthy)
        
        # Load Healthy MinMax Data
        healthy_path = os.path.join(DEFAULT_DATA_PATH, "Healthy Min Max Stock Level.xlsx")
        if os.path.exists(healthy_path):
            try:
                h_df = pd.read_excel(healthy_path, header=None)
                h_data = h_df.iloc[3:].copy()
                id_rows = h_data[h_data[2].astype(str).str.contains("ID -", na=False)]
                
                headers = ["Lead Day", "Country's Area", "Forecast Turnover", "Non-Festival Min", "Non-Festival Max", "Festival Min", "Festival Max", "Buffer (week)"]
                self.table_healthy.setRowCount(len(id_rows) + 2)
                self.table_healthy.setColumnCount(len(headers))
                self.table_healthy.horizontalHeader().hide()
                
                # Set spans for visual multi-level headers
                self.table_healthy.setSpan(0, 0, 2, 1) # Lead Day
                self.table_healthy.setSpan(0, 1, 2, 1) # Area
                self.table_healthy.setSpan(0, 2, 2, 1) # Forecast
                self.table_healthy.setSpan(0, 3, 1, 2) # Non-Festival
                self.table_healthy.setSpan(0, 5, 1, 2) # Festival
                self.table_healthy.setSpan(0, 7, 2, 1) # Buffer
                
                hdr_defs = [
                    (0, 0, "Lead Day", "#ffffff", "#1e293b"),
                    (0, 1, "Area", "#ffffff", "#1e293b"),
                    (0, 2, "Forecast Turnover", "#ffffff", "#1e293b"),
                    (0, 3, "Non-Festival", "#e2efda", "#375623"),
                    (0, 5, "Festival", "#fce4d6", "#c65911"),
                    (0, 7, "Buffer (week)", "#ffffff", "#1e293b"),
                    (1, 3, "Min", "#e2efda", "#375623"),
                    (1, 4, "Max", "#e2efda", "#375623"),
                    (1, 5, "Min", "#fce4d6", "#c65911"),
                    (1, 6, "Max", "#fce4d6", "#c65911")
                ]
                
                for r, c, text, bg, fg in hdr_defs:
                    item = QTableWidgetItem(text)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setBackground(QBrush(QColor(bg)))
                    item.setForeground(QBrush(QColor(fg)))
                    font = QFont()
                    font.setBold(True)
                    item.setFont(font)
                    item.setFlags(Qt.ItemFlag.ItemIsEnabled)
                    self.table_healthy.setItem(r, c, item)
                    
                empty_coords = [(1, 0), (1, 1), (1, 2), (0, 4), (0, 6), (1, 7)]
                for r, c in empty_coords:
                    item = QTableWidgetItem("")
                    item.setBackground(QBrush(QColor("#ffffff" if r == 1 and c in [0, 1, 2, 7] else "#e2efda" if c == 4 else "#fce4d6")))
                    item.setFlags(Qt.ItemFlag.ItemIsEnabled)
                    self.table_healthy.setItem(r, c, item)

                col_indices = [1, 2, 3, 5, 6, 7, 8, 9]
                
                for r_idx, (_, row) in enumerate(id_rows.iterrows()):
                    for c_idx, src_col in enumerate(col_indices):
                        val = row.iloc[src_col]
                        if src_col in [5, 6, 7, 8]:
                            try:
                                text = f"{float(val) * 100:.0f}%"
                            except (ValueError, TypeError):
                                text = str(val) if pd.notnull(val) else ""
                        elif isinstance(val, float) and pd.notnull(val):
                            text = f"{val:.1f}".rstrip('0').rstrip('.')
                        else:
                            text = str(val) if pd.notnull(val) else ""
                            
                        item = QTableWidgetItem(text)
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        self.table_healthy.setItem(r_idx + 2, c_idx, item)
                        
                self.table_healthy.resizeColumnsToContents()
                for c_idx in range(len(headers)):
                    curr_w = self.table_healthy.columnWidth(c_idx)
                    self.table_healthy.setColumnWidth(c_idx, max(curr_w + 40, 150))
            except Exception as e:
                lbl_err = QLabel(f"Failed to load Healthy Stock Level data: {e}")
                lbl_err.setStyleSheet("color: #ef4444;")
                healthy_minmax_vbox.addWidget(lbl_err)
        else:
            lbl_err = QLabel("File not found: Healthy Min Max Stock Level.xlsx")
            lbl_err.setStyleSheet("color: #ef4444;")
            healthy_minmax_vbox.addWidget(lbl_err)
        # Attach Sparkline Delegate to all
        self.sparkline_delegate = SparklineDelegate()
        
        # TAB: STORE LIST
        self.store_list_container = QWidget()
        store_list_vbox = QVBoxLayout(self.store_list_container)
        store_list_vbox.setContentsMargins(15, 15, 15, 15); store_list_vbox.setSpacing(10)
        
        # State for column filters
        self.sl_active_filters = {} 
        
        # Active Filter Banner from All Summary (hidden by default)
        self.sl_filter_banner = QFrame()
        self.sl_filter_banner.setObjectName("SLFilterBanner")
        self.sl_filter_banner.setStyleSheet("""
            QFrame#SLFilterBanner {
                background-color: #eff6ff;
                border: 1px solid #bfdbfe;
                border-radius: 6px;
            }
            QLabel {
                font-family: 'Segoe UI', Arial;
                font-size: 12px;
                color: #1e3a8a;
            }
            QPushButton {
                background-color: #3b82f6;
                color: white;
                font-weight: bold;
                border: none;
                border-radius: 4px;
                padding: 4px 10px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #2563eb;
            }
        """)
        banner_layout = QHBoxLayout(self.sl_filter_banner)
        banner_layout.setContentsMargins(10, 6, 10, 6)
        
        self.sl_filter_banner_lbl = QLabel("")
        banner_layout.addWidget(self.sl_filter_banner_lbl)
        banner_layout.addStretch()
        
        sl_clear_filter_btn = QPushButton("Clear Filter")
        sl_clear_filter_btn.clicked.connect(self.clear_all_store_list_filters)
        banner_layout.addWidget(sl_clear_filter_btn)
        
        self.sl_filter_banner.setVisible(False)
        store_list_vbox.addWidget(self.sl_filter_banner) 
        
        # Static layout for frozen columns (no splitter)
        self.store_list_layout = QHBoxLayout()
        self.store_list_layout.setSpacing(0)
        
        self.table_store_list_frozen = QTableWidget()
        self.table_store_list_frozen.setAlternatingRowColors(True)
        self.table_store_list_frozen.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_store_list_frozen.verticalHeader().setDefaultSectionSize(30)
        self.table_store_list_frozen.verticalScrollBar().setVisible(False)
        self.table_store_list_frozen.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.table_store_list_frozen.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_store_list_frozen.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_store_list_frozen.setStyleSheet("QTableWidget { border: 1px solid #e2e8f0; border-right: none; border-top-right-radius: 0; border-bottom-right-radius: 0; }")
        
        self.table_store_list = QTableWidget()
        self.table_store_list.setAlternatingRowColors(True)
        self.table_store_list.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_store_list.verticalHeader().setDefaultSectionSize(30)
        self.table_store_list.verticalHeader().setVisible(False)
        self.table_store_list.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_store_list.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_store_list.setStyleSheet("QTableWidget { border: 1px solid #e2e8f0; border-left: none; border-top-left-radius: 0; border-bottom-left-radius: 0; }")
        
        # Link headers to filter logic
        self.table_store_list_frozen.horizontalHeader().sectionClicked.connect(lambda idx: self.on_store_list_header_clicked(idx, table=self.table_store_list_frozen))
        self.table_store_list.horizontalHeader().sectionClicked.connect(lambda idx: self.on_store_list_header_clicked(idx, table=self.table_store_list))
        
        # Link them
        self.table_store_list.frozen_side = self.table_store_list_frozen
        self.table_store_list.verticalScrollBar().valueChanged.connect(self.table_store_list_frozen.verticalScrollBar().setValue)
        self.table_store_list_frozen.verticalScrollBar().valueChanged.connect(self.table_store_list.verticalScrollBar().setValue)
        
        # Sync Selection
        self.table_store_list.itemSelectionChanged.connect(lambda: self.sync_table_selection(self.table_store_list, self.table_store_list_frozen))
        self.table_store_list_frozen.itemSelectionChanged.connect(lambda: self.sync_table_selection(self.table_store_list_frozen, self.table_store_list))
        
        self.store_list_layout.addWidget(self.table_store_list_frozen)
        self.store_list_layout.addWidget(self.table_store_list)
        
        store_list_vbox.addLayout(self.store_list_layout)
        
        # Load Store List Data
        if self.store_path and os.path.exists(self.store_path):
            try:
                sl_df = self._discover_and_normalize_store_sheet(self.store_path)
                sl_df["Sales Amount"] = 0
                sl_df = sl_df.fillna("")
                self.full_store_list_df = sl_df.copy() # Store full data for filtering
                
                # Initial population
                self.apply_store_list_filters()
            except Exception as e:
                lbl_err = QLabel(f"Failed to load Store List data: {e}")
                lbl_err.setStyleSheet("color: #ef4444;")
                store_list_vbox.addWidget(lbl_err)

        self.sales_container = QWidget()
        sales_vbox = QVBoxLayout(self.sales_container)
        
        self.table_sales = QTableWidget()
        self.table_sales.setAlternatingRowColors(True)
        self.table_sales.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_sales.horizontalHeader().sectionClicked.connect(self.on_sales_header_clicked)
        sales_vbox.addWidget(self.table_sales)
        
        # TAB 6: PERFORMANCE ANALYSIS
        self.perf_container = QWidget()
        perf_vbox = QVBoxLayout(self.perf_container)
        perf_vbox.setContentsMargins(15, 15, 15, 15); perf_vbox.setSpacing(10)

        # Filter Section
        f_layout = QHBoxLayout()
        f_layout.setContentsMargins(0, 0, 0, 10)
        
        # Dept Type (Single Select)
        l1 = QLabel("Dept Type:"); l1.setStyleSheet("font-weight: bold; color: #1e293b; font-size: 11px;")
        self.perf_filter_type = QComboBox(); self.perf_filter_type.setFixedWidth(200); self.perf_filter_type.addItems(["ALL"])
        f_layout.addWidget(l1); f_layout.addWidget(self.perf_filter_type)

        # Dept Details (Multi Select)
        l2 = QLabel("Dept Details:"); l2.setStyleSheet("font-weight: bold; color: #1e293b; font-size: 11px;")
        self.perf_filter_details = CheckableComboBox(); self.perf_filter_details.setFixedWidth(200)
        f_layout.addWidget(l2); f_layout.addWidget(self.perf_filter_details)

        f_layout.addStretch()
        perf_vbox.addLayout(f_layout)

        self.table_perf_splitter = QSplitter(Qt.Orientation.Horizontal)
        self._perf_panels = []  # Will be populated dynamically
        perf_vbox.addWidget(self.table_perf_splitter)

        # Connect filters
        self.perf_filter_type.currentIndexChanged.connect(self.on_perf_filter_changed)
        self.perf_filter_details.currentIndexChanged.connect(self.on_perf_filter_changed)
        
        # TAB 7: ALL SUMMARY
        self.all_summary_container = QWidget()
        all_summary_vbox = QVBoxLayout(self.all_summary_container)
        all_summary_vbox.setContentsMargins(15, 15, 15, 15); all_summary_vbox.setSpacing(10)
        
        self.table_all_summary = QTableWidget()
        self.table_all_summary.setAlternatingRowColors(True)
        self.table_all_summary.verticalHeader().setDefaultSectionSize(30)
        self.table_all_summary.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_all_summary.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_all_summary.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        
        all_summary_vbox.addWidget(self.table_all_summary)
        
        # TAB 8: GROUPING ANALYSIS
        self.grouping_analysis_container = QWidget()
        grouping_vbox = QVBoxLayout(self.grouping_analysis_container)
        grouping_vbox.setContentsMargins(15, 15, 15, 15); grouping_vbox.setSpacing(10)
        
        self.grouping_search = None
        self.grouping_set_combo = None
        
        # Layout container for header and table
        self.grouping_table_layout = QVBoxLayout()
        self.grouping_table_layout.setSpacing(0)
        self.grouping_table_layout.setContentsMargins(0, 0, 0, 0)
        
        # Header Table for Grouping Analysis (multi-level)
        self.table_grouping_header = QTableWidget()
        self.table_grouping_header.setObjectName("GroupingHeader")
        self.table_grouping_header.setRowCount(3)
        self.table_grouping_header.verticalHeader().setVisible(False)
        self.table_grouping_header.horizontalHeader().setVisible(False)
        self.table_grouping_header.setFixedHeight(94)  # 3 rows of ~30px plus borders
        self.table_grouping_header.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table_grouping_header.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table_grouping_header.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_grouping_header.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_grouping_header.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_grouping_header.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table_grouping_header.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.table_grouping_header.setStyleSheet("""
            QTableWidget#GroupingHeader {
                gridline-color: #cbd5e1;
                border: 1px solid #cbd5e1;
                background-color: #ffffff;
            }
            QTableWidget#GroupingHeader QScrollBar:vertical {
                background: transparent;
                border: none;
            }
            QTableWidget#GroupingHeader QScrollBar::handle:vertical {
                background: transparent;
                border: none;
            }
            QTableWidget#GroupingHeader QScrollBar::add-line:vertical,
            QTableWidget#GroupingHeader QScrollBar::sub-line:vertical {
                border: none;
                background: none;
                height: 0px;
            }
            QTableWidget#GroupingHeader QScrollBar::add-page:vertical,
            QTableWidget#GroupingHeader QScrollBar::sub-page:vertical {
                background: none;
            }
            QTableWidget#GroupingHeader QScrollBar:horizontal {
                height: 0px;
                background: transparent;
                border: none;
            }
            QTableWidget#GroupingHeader QScrollBar::handle:horizontal {
                background: transparent;
                border: none;
            }
            QTableWidget#GroupingHeader QScrollBar::add-line:horizontal,
            QTableWidget#GroupingHeader QScrollBar::sub-line:horizontal {
                border: none;
                background: none;
                width: 0px;
            }
            QTableWidget#GroupingHeader QScrollBar::add-page:horizontal,
            QTableWidget#GroupingHeader QScrollBar::sub-page:horizontal {
                background: none;
            }
        """)
        
        # Main Table for Grouping Analysis
        self.table_grouping_analysis = QTableWidget()
        self.table_grouping_analysis.setAlternatingRowColors(True)
        self.table_grouping_analysis.verticalHeader().setVisible(False)
        self.table_grouping_analysis.horizontalHeader().setVisible(False)
        self.table_grouping_analysis.verticalHeader().setDefaultSectionSize(30)
        self.table_grouping_analysis.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_grouping_analysis.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_grouping_analysis.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_grouping_analysis.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table_grouping_analysis.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.table_grouping_analysis.setStyleSheet("""
            QTableWidget {
                gridline-color: #f1f5f9;
                border: 1px solid #cbd5e1;
                border-top: none;
                alternate-background-color: #f8fafc;
                selection-background-color: #dbeafe;
                selection-color: #1e293b;
            }
        """)
        
        # Attach the ScoreBarDelegate to columns
        self.score_bar_delegate = ScoreBarDelegate()
        
        self.grouping_table_layout.addWidget(self.table_grouping_header)
        self.grouping_table_layout.addWidget(self.table_grouping_analysis)
        
        grouping_vbox.addLayout(self.grouping_table_layout)
        
        # TAB 9: GROUPING SUMMARY
        self.grouping_summary_container = QWidget()
        summary_vbox = QVBoxLayout(self.grouping_summary_container)
        summary_vbox.setContentsMargins(15, 15, 15, 15); summary_vbox.setSpacing(10)
        
        # Legend (Horizontal Alignment)
        legend_layout = QHBoxLayout()
        legend_layout.setSpacing(10)
        legend_layout.setContentsMargins(5, 0, 5, 5)
        
        lbl_legend_title = QLabel("Rankings Legend:")
        lbl_legend_title.setStyleSheet("font-weight: bold; font-size: 12px; color: #475569;")
        legend_layout.addWidget(lbl_legend_title)
        
        # 1st Place Item
        item_1st = QFrame()
        item_1st.setStyleSheet("background-color: #778873; border-radius: 4px;")
        layout_1st = QHBoxLayout(item_1st)
        layout_1st.setContentsMargins(10, 4, 10, 4)
        lbl_1st = QLabel("Highest")
        lbl_1st.setStyleSheet("color: #ffffff; font-weight: bold; font-size: 11px; background: transparent; border: none;")
        layout_1st.addWidget(lbl_1st)
        legend_layout.addWidget(item_1st)
        
        # 2nd Place Item
        item_2nd = QFrame()
        item_2nd.setStyleSheet("background-color: #A1BC98; border-radius: 4px;")
        layout_2nd = QHBoxLayout(item_2nd)
        layout_2nd.setContentsMargins(10, 4, 10, 4)
        lbl_2nd = QLabel("2nd Highest")
        lbl_2nd.setStyleSheet("color: #000000; font-weight: bold; font-size: 11px; background: transparent; border: none;")
        layout_2nd.addWidget(lbl_2nd)
        legend_layout.addWidget(item_2nd)
        
        # 3rd Place Item
        item_3rd = QFrame()
        item_3rd.setStyleSheet("background-color: #D2DCB6; border-radius: 4px;")
        layout_3rd = QHBoxLayout(item_3rd)
        layout_3rd.setContentsMargins(10, 4, 10, 4)
        lbl_3rd = QLabel("3rd Highest")
        lbl_3rd.setStyleSheet("color: #000000; font-weight: bold; font-size: 11px; background: transparent; border: none;")
        layout_3rd.addWidget(lbl_3rd)
        legend_layout.addWidget(item_3rd)
        
        legend_layout.addStretch()
        summary_vbox.addLayout(legend_layout)
        
        # Table Layout
        self.summary_table_layout = QVBoxLayout()
        self.summary_table_layout.setSpacing(0)
        self.summary_table_layout.setContentsMargins(0, 0, 0, 0)
        
        # Header Table for Grouping Summary (multi-level)
        self.table_summary_header = QTableWidget()
        self.table_summary_header.setObjectName("GroupingSummaryHeader")
        self.table_summary_header.setRowCount(2)
        self.table_summary_header.verticalHeader().setVisible(False)
        self.table_summary_header.horizontalHeader().setVisible(False)
        self.table_summary_header.setFixedHeight(96)  # Increased from 66 to fit 32+60 row heights
        self.table_summary_header.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.table_summary_header.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.table_summary_header.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_summary_header.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table_summary_header.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.table_summary_header.setStyleSheet("""
            QTableWidget#GroupingSummaryHeader {
                border: 1px solid #cbd5e1;
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                background-color: #f8fafc;
            }
            QTableWidget#GroupingSummaryHeader::item {
                border-right: 1px solid #cbd5e1;
                border-bottom: 1px solid #cbd5e1;
                padding: 4px;
            }
        """)
        
        # Main Table for Grouping Summary
        self.table_summary_analysis = QTableWidget()
        self.table_summary_analysis.setAlternatingRowColors(True)
        self.table_summary_analysis.verticalHeader().setVisible(False)
        self.table_summary_analysis.horizontalHeader().setVisible(False)
        self.table_summary_analysis.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_summary_analysis.setWordWrap(True)
        self.table_summary_analysis.setStyleSheet("""
            QTableWidget {
                gridline-color: #cbd5e1;
                border: 1px solid #cbd5e1;
                border-top: none;
                border-bottom-left-radius: 8px;
                border-bottom-right-radius: 8px;
                alternate-background-color: #f8fafc;
                selection-background-color: #dbeafe;
                selection-color: #1e293b;
            }
        """)
        
        self.summary_table_layout.addWidget(self.table_summary_header)
        self.summary_table_layout.addWidget(self.table_summary_analysis)
        summary_vbox.addLayout(self.summary_table_layout)
        
        # Sync scrolling horizontally bidirectionally for Grouping Summary
        self.table_summary_analysis.horizontalScrollBar().valueChanged.connect(
            lambda val: self.table_summary_header.horizontalScrollBar().setValue(val)
        )
        self.table_summary_header.horizontalScrollBar().valueChanged.connect(
            lambda val: self.table_summary_analysis.horizontalScrollBar().setValue(val)
        )
        # Sync scrolling horizontally bidirectionally
        self.table_grouping_analysis.horizontalScrollBar().valueChanged.connect(
            self.sync_grouping_horizontal_analysis
        )
        self.table_grouping_header.horizontalScrollBar().valueChanged.connect(
            self.sync_grouping_horizontal_header
        )
        
        # Drill-down interaction on double clicking table cells
        self.table_all_summary.cellDoubleClicked.connect(self.on_all_summary_cell_clicked)
        
        self.tabs.addTab(readme_scroll, " Readme ")
        self.tabs.addTab(chart_scroll, " Chart ")
        self.tabs.addTab(self.summary_container, " Dept Summary ")
        # self.tabs.addTab(self.overall_container, " Overall Summary ")
        self.tabs.addTab(quarter_scroll, " DA Analysis ")
        # self.tabs.addTab(self.healthy_minmax_container, " Healthy MinMax ")
        self.tabs.addTab(self.store_list_container, " Store List ")
        # self.tabs.addTab(self.sales_container, " Sales Performance ")
        # self.tabs.addTab(self.perf_container, " Performance Analysis ")
        self.tabs.addTab(self.all_summary_container, " All Summary ")
        self.tabs.addTab(self.grouping_analysis_container, " Grouping Analysis ")
        self.tabs.addTab(self.grouping_summary_container, " Grouping Summary ")

        self.tabs.setCurrentIndex(0)
        self.tabs.currentChanged.connect(self.on_tab_changed)
        layout.addWidget(self.tabs, 1)
        
        # Keyboard Shortcuts
        self.action_new_table = QAction("New Table", self)
        self.action_new_table.setShortcut("Ctrl+N")
        self.action_new_table.triggered.connect(self.create_new_custom_tab)
        self.addAction(self.action_new_table)
        
        self.action_close_tab = QAction("Close Tab", self)
        self.action_close_tab.setShortcut("Ctrl+W")
        self.action_close_tab.triggered.connect(self.close_current_custom_tab)
        self.addAction(self.action_close_tab)

        self.on_primary_set_changed()
        
        # STATUS BAR
        self.status_bar = QHBoxLayout()
        self.status_msg = QLabel("Ready")
        self.status_msg.setStyleSheet("color: #64748b; font-size: 11px;")
        self.status_store_count = QLabel("Total Stores: -")
        self.status_store_count.setStyleSheet("color: #1e293b; font-weight: bold; font-size: 11px;")
        
        self.status_bar.addWidget(self.status_msg)
        self.status_bar.addStretch()
        self.status_bar.addWidget(self.status_store_count)
        layout.addLayout(self.status_bar)
        
        self.on_primary_set_changed()

    def on_header_clicked(self, table, index):
        """Show an input dialog to filter by Department (0), Report Set (1), Category (2), or Dept Type (16)."""
        filterable_cols = {0: "Department", 1: "Report Set", 2: "Category", 16: "Dept Type"}
        if index not in filterable_cols: return
        
        try:
            col_name = filterable_cols[index]
            current = self.active_filters.get(index, "")
            from PyQt6.QtWidgets import QInputDialog
            
            if index == 1: # Report Set gets a dropdown
                options_map = {
                    "All Sets": "All Sets",
                    "Set1_Display Area": "Set1_DA",
                    "Set2_Store Type": "Set2_StoreType",
                    "Set3_SA Split": "Set3_Type_DA",
                    "Set4_DA Split": "Set4_CurrentSetting",
                    "Set5_Mall & SA Split": "Set5_Mall_SA_Split"
                }
                options = list(options_map.keys())
                cur_idx = 0
                if current:
                    rev_map = {v: k for k, v in options_map.items()}
                    cur_idx = options.index(rev_map.get(current, "All Sets"))
                    
                text, ok = QInputDialog.getItem(self, f"Select {col_name}", f"Select {col_name}", options, cur_idx, False,
                                               flags=Qt.WindowType.WindowTitleHint | Qt.WindowType.WindowSystemMenuHint | Qt.WindowType.WindowCloseButtonHint)
                if ok:
                    self.active_filters[index] = "" if text == "All Sets" else options_map[text]
                    self.apply_multi_filter()
            elif index in [2, 16]: # Category and Dept Type now both get dropdowns
                # Get unique values from the column
                unique_vals = set()
                for r in range(self.table_overall.rowCount()):
                    item = self.table_overall.item(r, index)
                    if item and item.text().strip():
                        unique_vals.add(item.text().strip())
                
                reset_text = "All Categories" if index == 2 else "All Types"
                options = [reset_text] + sorted(list(unique_vals))
                cur_idx = options.index(current) if current in options else 0
                
                text, ok = QInputDialog.getItem(self, f"Select {col_name}", f"Select {col_name}", options, cur_idx, False,
                                               flags=Qt.WindowType.WindowTitleHint | Qt.WindowType.WindowSystemMenuHint | Qt.WindowType.WindowCloseButtonHint)
                if ok:
                    self.active_filters[index] = "" if text == reset_text else text
                    self.apply_multi_filter()
            else: # Department gets text input
                text, ok = QInputDialog.getText(self, f"Filter {col_name}", f"Enter {col_name}", text=current,
                                                flags=Qt.WindowType.WindowTitleHint | Qt.WindowType.WindowSystemMenuHint | Qt.WindowType.WindowCloseButtonHint)
                if ok:
                    self.active_filters[index] = text.strip()
                    self.apply_multi_filter()
                
        except Exception as e:
            print(f"Filter error: {e}")
 
    def apply_multi_filter(self):
        """Apply all active filters across Overall Summary table."""
        table = self.table_overall
        for r in range(table.rowCount()):
            match = True
            for col_idx, filter_text in self.active_filters.items():
                if not filter_text: continue
                item = table.item(r, col_idx)
                if not item or filter_text.lower() not in item.text().lower():
                    match = False; break
            table.setRowHidden(r, not match)
        
        # Update header labels for all filterable columns
        filterable_cols = {0: "Department", 1: "Report Set", 16: "Dept Type"}
        for idx, base_name in filterable_cols.items():
            f_text = self.active_filters.get(idx, "")
            label = base_name if not f_text else f"{f_text}"
            if table.columnCount() > idx:
                table.model().setHeaderData(idx, Qt.Orientation.Horizontal, label)

    def apply_filter_to_overall(self):
        """Refreshes all active filters (Department and Dept Type)."""
        self.apply_multi_filter()

    def create_input(self, label, default):
        h = QHBoxLayout(); lbl = QLabel(label); lbl.setStyleSheet("font-size: 11px; color: #64748b;")
        lbl.setFixedWidth(120) # Ensure labels have consistent width
        edit = QLineEdit(default); h.addWidget(lbl); h.addWidget(edit)
        h.input = edit; h.lbl = lbl; return h

    def on_primary_set_changed(self):
        self.update_ui_visibility()
        self.update_chart()
            
    def on_tab_changed(self, index):
        tab_text = self.tabs.tabText(index).strip()
        if tab_text in ["All Summary", "Sales Performance", "Store List"]:
            for layout in [self.v_sa_split, self.v_da_low, self.v_da_high, self.v_da_slicer, self.v_set5_mall, self.v_set5_sa]:
                for i in range(layout.count()):
                    w = layout.itemAt(i).widget()
                    if w: w.setVisible(True)
        elif tab_text == "Grouping Analysis":
            self.populate_grouping_analysis_table()
        elif tab_text == "Grouping Summary":
            self.populate_grouping_summary_table()
        elif tab_text == "Group":
            self.populate_group_tab()
        else:
            self.update_ui_visibility()

    def update_ui_visibility(self):
        # Reporting Sets involved
        try:
            checked = self.combo_set.checked_items() if hasattr(self, 'combo_set') and hasattr(self.combo_set, 'checked_items') else []
        except RuntimeError:
            checked = []
        t1 = checked[0] if checked else "Set1_DA"
        checked_compare = checked[1:] if len(checked) > 1 else []
        
        # SA Slicer logic (Only for Set 3)
        show_sa = (t1 == "Set3_Type_DA" or "Set3_Type_DA" in checked_compare)
        for i in range(self.v_sa_split.count()):
            w = self.v_sa_split.itemAt(i).widget()
            if w: w.setVisible(show_sa)

        # DA Lower and Upper logic (Only for Set 1)
        show_da_bounds = (t1 == "Set1_DA" or "Set1_DA" in checked_compare)
        for i in range(self.v_da_low.count()):
            w = self.v_da_low.itemAt(i).widget()
            if w: w.setVisible(show_da_bounds)
        for i in range(self.v_da_high.count()):
            w = self.v_da_high.itemAt(i).widget()
            if w: w.setVisible(show_da_bounds)

        # DA Slicer logic (Only for Set 4)
        show_da_slicer = (t1 == "Set4_CurrentSetting" or "Set4_CurrentSetting" in checked_compare)
        for i in range(self.v_da_slicer.count()):
            w = self.v_da_slicer.itemAt(i).widget()
            if w: w.setVisible(show_da_slicer)

        # Set 5 Slicer logic (Only for Set 5)
        show_set5 = (t1 == "Set5_Mall_SA_Split" or "Set5_Mall_SA_Split" in checked_compare)
        for i in range(self.v_set5_mall.count()):
            w = self.v_set5_mall.itemAt(i).widget()
            if w: w.setVisible(show_set5)
        for i in range(self.v_set5_sa.count()):
            w = self.v_set5_sa.itemAt(i).widget()
            if w: w.setVisible(show_set5)

        # Update All Summary tab tables visibility dynamically
        self.update_all_summary_visibility()

    def update_all_summary_visibility(self):
        """Update which pivot tables are visible in the All Summary tab based on selection."""
        if not hasattr(self, 'all_summary_grid'):
            return
            
        try:
            checked = self.combo_set.checked_items() if hasattr(self, 'combo_set') and hasattr(self.combo_set, 'checked_items') else []
        except RuntimeError:
            checked = []
        primary_set = checked[0] if checked else "Set1_DA"
        compare_sets = checked[1:] if len(checked) > 1 else []
        
        set_mapping = {
            "Set1_DA": self.table_all_set1,
            "Set2_StoreType": self.table_all_set2,
            "Set3_Type_DA": self.table_all_set3,
            "Set4_CurrentSetting": self.table_all_set4,
            "Set5_Mall_SA_Split": self.table_all_set5
        }
        
        # Build ordered list of active sets (primary first, then compared ones)
        active_keys = []
        if primary_set:
            active_keys.append(primary_set)
        for cs in compare_sets:
            if cs and cs not in active_keys:
                active_keys.append(cs)
                
        active_tables = [set_mapping[k] for k in active_keys if k in set_mapping]
        
        # Hide all tables and remove them from grid layout
        for table in set_mapping.values():
            table.setVisible(False)
            self.all_summary_grid.removeWidget(table)
            
        # Dynamically place active tables in grid layout to fill space
        num_active = len(active_tables)
        if num_active == 1:
            self.all_summary_grid.addWidget(active_tables[0], 0, 0, 1, 1)
        elif num_active == 2:
            self.all_summary_grid.addWidget(active_tables[0], 0, 0, 1, 1)
            self.all_summary_grid.addWidget(active_tables[1], 0, 1, 1, 1)
        elif num_active == 3:
            self.all_summary_grid.addWidget(active_tables[0], 0, 0, 1, 1)
            self.all_summary_grid.addWidget(active_tables[1], 0, 1, 1, 1)
            self.all_summary_grid.addWidget(active_tables[2], 1, 0, 1, 2)  # spans both columns
        elif num_active == 4:
            self.all_summary_grid.addWidget(active_tables[0], 0, 0, 1, 1)
            self.all_summary_grid.addWidget(active_tables[1], 0, 1, 1, 1)
            self.all_summary_grid.addWidget(active_tables[2], 1, 0, 1, 1)
            self.all_summary_grid.addWidget(active_tables[3], 1, 1, 1, 1)
        elif num_active >= 5:
            self.all_summary_grid.addWidget(active_tables[0], 0, 0, 1, 1)
            self.all_summary_grid.addWidget(active_tables[1], 0, 1, 1, 1)
            self.all_summary_grid.addWidget(active_tables[2], 1, 0, 1, 1)
            self.all_summary_grid.addWidget(active_tables[3], 1, 1, 1, 1)
            self.all_summary_grid.addWidget(active_tables[4], 2, 0, 1, 2)
            
        # Make selected tables visible
        for table in active_tables:
            table.setVisible(True)

    def on_scan_finished(self, groups):
        prev = getattr(self, 'previous_selected_dept', None) or self.combo_dept.currentText().strip()
        self.combo_dept.clear()
        self.combo_dept.addItems(groups)
        if prev:
            idx = self.combo_dept.findText(prev)
            if idx >= 0:
                self.combo_dept.setCurrentIndex(idx)
        self.status_msg.setText(f"POSSYS Loaded: {Path(self.possys_path).name}")
        # Auto-run analysis once departments are ready
        self.update_chart()

    def update_chart(self):
        if not self.possys_path or not self.store_path: return
        try:
            checked = self.combo_set.checked_items() if hasattr(self, 'combo_set') and hasattr(self.combo_set, 'checked_items') else []
        except RuntimeError:
            checked = []
        target_set = checked[0] if checked else "Set1_DA"
        target_set_compare = checked[1:] if len(checked) > 1 else []
        config = {
            'base_folder': os.path.dirname(self.possys_path), 
            'possys_path': self.possys_path, 
            'store_path': self.store_path,
            'sales_path': self.sales_path,
            'sales_new_path': self.sales_new_path,
            'details_path': self.details_path,
            'country': self.country,
            'da_low': self.v_da_low.input.text(), 
            'da_high': self.v_da_high.input.text(), 
            'sa_split': self.v_sa_split.input.text(),
            'da_slicer': self.v_da_slicer.input.text(),
            'set5_mall': self.v_set5_mall.input.text(),
            'set5_sa': self.v_set5_sa.input.text(),
            'target_group': self.combo_dept.currentText(), 
            'target_set': target_set,
            'target_set_compare': target_set_compare
        }
        self.btn_run.setEnabled(False); self.progress.setVisible(True)
        self.progress.setRange(0, 0)
        self.progress.setFormat("Generating Chart... %p%")
        self._discard_thread('worker')
        self.worker = ChartWorker(config)
        self.worker.chart_ready.connect(self.on_chart_ready); self.worker.error.connect(self.on_error); self.worker.start()

    def on_chart_ready(self, data):
        import time
        try:
            fig, fig_dist, preview_df, overall_df, quarter_df, total_unique, sales_df, full_df, config, excluded_df = data
            self.full_df = full_df
            self.excluded_df = excluded_df
            self.overall_df = overall_df
            self.sales_df = sales_df
            self.update_readme_with_excluded(excluded_df)
            self.last_run_config = config
            t0 = time.time()
            
            # Update Main Chart
            for i in reversed(range(self.chart_layout.count())): self.chart_layout.itemAt(i).widget().setParent(None)
            self.canvas = FigureCanvas(fig)
            self.chart_layout.addWidget(self.canvas)
            print(f"[TIMING] Chart canvas: {time.time()-t0:.2f}s"); t0=time.time()
            
            # Update Distribution Chart (Quarter Tab)
            for i in reversed(range(self.quarter_layout.count())): self.quarter_layout.itemAt(i).widget().setParent(None)
            self.canvas_dist = FigureCanvas(fig_dist)
            self.quarter_layout.addWidget(self.canvas_dist)
            print(f"[TIMING] Dist canvas: {time.time()-t0:.2f}s"); t0=time.time()
            
            # Update Tables
            set_names = {
                "Set1_DA": "Set 1",
                "Set2_StoreType": "Set 2",
                "Set3_Type_DA": "Set 3",
                "Set4_CurrentSetting": "Set 4",
                "Set5_Mall_SA_Split": "Set 5"
            }
            
            if "Report Set" in preview_df.columns:
                preview_df["Report Set"] = preview_df["Report Set"].map(set_names).fillna(preview_df["Report Set"])
                
            if "Report Set" in overall_df.columns:
                overall_df["Report Set"] = overall_df["Report Set"].map(set_names).fillna(overall_df["Report Set"])

            self.populate_summary_table(self.table_summary, preview_df, heatmap=True)
            print(f"[TIMING] table_summary ({len(preview_df)} rows): {time.time()-t0:.2f}s"); t0=time.time()
            self.populate_summary_table(self.table_overall, overall_df, heatmap=True)
            print(f"[TIMING] table_overall ({len(overall_df)} rows): {time.time()-t0:.2f}s"); t0=time.time()
            
            # Populate All Summary Pivot Tables
            try:
                sidf = self._discover_and_normalize_store_sheet(self.store_path)
                
                sidf["Store_Code"] = sidf["Store_Code"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
                
                # Only keep stores that were actually analyzed
                if hasattr(self, "full_df") and self.full_df is not None and not self.full_df.empty:
                    analyzed_codes = set(self.full_df["Store_Code"].dropna().unique())
                    sidf = sidf[sidf["Store_Code"].isin(analyzed_codes)].copy()
                
                sidf["Store_Display_Area_num"] = pd.to_numeric(sidf["Store_Display_Area"].astype(str).str.replace(",", ""), errors="coerce")
                
                d_low = pd.to_numeric(self.v_da_low.input.text().replace(",", ""), errors='coerce')
                d_high = pd.to_numeric(self.v_da_high.input.text().replace(",", ""), errors='coerce')
                s_split = pd.to_numeric(self.v_sa_split.input.text().replace(",", ""), errors='coerce')
                d_slicer = pd.to_numeric(self.v_da_slicer.input.text().replace(",", ""), errors='coerce')
                set5_mall = pd.to_numeric(self.v_set5_mall.input.text().replace(",", ""), errors='coerce')
                set5_sa = pd.to_numeric(self.v_set5_sa.input.text().replace(",", ""), errors='coerce')
                
                if pd.isna(d_low): d_low = 7500
                if pd.isna(d_high): d_high = 9500
                if pd.isna(s_split): s_split = 8000
                if pd.isna(d_slicer): d_slicer = 8000
                if pd.isna(set5_mall): set5_mall = 10000
                if pd.isna(set5_sa): set5_sa = 8000
                
                # --- Map Sales Amount into sidf first ---
                sidf["Sales Amount"] = 0.0
                try:
                    if not sales_df.empty:
                        s_col = "STORE" if "STORE" in sales_df.columns else "M_STORE"
                        v_col = "AVG_TOP2_TOTAL_AMT_SALES"
                        if s_col in sales_df.columns and v_col in sales_df.columns:
                            clean_s_col = sales_df[s_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
                            clean_v_col = pd.to_numeric(sales_df[v_col].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
                            
                            temp_df = pd.DataFrame({s_col: clean_s_col, v_col: clean_v_col})
                            store_sales = temp_df.groupby(s_col)[v_col].sum().reset_index()
                            
                            sidf["Sales Amount"] = sidf["Store_Code"].map(store_sales.set_index(s_col)[v_col]).fillna(0)
                except Exception as sle:
                    print(f"Store List Sales Pre-Update Error: {sle}")

                # Set 1
                sidf["Set 1"] = np.select(
                    [sidf["Store_Display_Area_num"] < d_low, 
                     (sidf["Store_Display_Area_num"] >= d_low) & (sidf["Store_Display_Area_num"] < d_high),
                     sidf["Store_Display_Area_num"] >= d_high],
                    [f"<{d_low:,.0f}", f">={d_low:,.0f}, <{d_high:,.0f}", f">={d_high:,.0f}"],
                    default="Not Found"
                )
                
                # Set 3
                sidf["Set 3"] = np.select(
                    [sidf["Store_Type"].str.lower() == "mall",
                     (sidf["Store_Type"].str.lower() == "standalone") & (sidf["Store_Display_Area_num"] < s_split),
                     (sidf["Store_Type"].str.lower() == "standalone") & (sidf["Store_Display_Area_num"] >= s_split)],
                    ["Mall", f"SA < {s_split:,.0f}", f"SA >= {s_split:,.0f}"],
                    default="Not Found"
                )
                
                # Set 4
                sidf["Set 4"] = np.select(
                    [sidf["Store_Display_Area_num"] < d_slicer,
                     (sidf["Store_Type"].str.lower() == "mall") & (sidf["Store_Display_Area_num"] >= d_slicer),
                     (sidf["Store_Type"].str.lower() == "standalone") & (sidf["Store_Display_Area_num"] >= d_slicer)],
                    [f"<{d_slicer:,.0f}", f"Mall >= {d_slicer:,.0f}", f"SA >= {d_slicer:,.0f}"],
                    default="Not Found"
                )
                
                # Set 2
                sidf["Set 2"] = np.select(
                    [sidf["Store_Type"].str.lower() == "mall",
                     sidf["Store_Type"].str.lower() == "standalone"],
                    ["Mall", "SA"],
                    default="Not Found"
                )

                # Set 5
                sidf["Set 5"] = np.select(
                    [
                        (sidf["Store_Type"].str.lower() == "standalone") & (sidf["Store_Display_Area_num"] < set5_sa),
                        (sidf["Store_Type"].str.lower() == "standalone") & (sidf["Store_Display_Area_num"] >= set5_sa),
                        (sidf["Store_Type"].str.lower() == "mall") & (sidf["Store_Display_Area_num"] < set5_mall),
                        (sidf["Store_Type"].str.lower() == "mall") & (sidf["Store_Display_Area_num"] >= set5_mall)
                    ],
                    [
                        f"SA < {set5_sa:,.0f}",
                        f"SA >= {set5_sa:,.0f}",
                        f"Mall < {set5_mall:,.0f}",
                        f"Mall >= {set5_mall:,.0f}"
                    ],
                    default="Not Found"
                )
                
                # Setup delegates for All Summary data bars
                if not hasattr(self, "_all_summary_delegates"):
                    self._all_summary_delegates = [
                        DataBarDelegate("#f59e0b"), # Amber/Orange for Store Count / Subtotals
                        DataBarDelegate("#10b981")  # Teal Green for Sales / Contributions
                    ]
                orange_del, green_del = self._all_summary_delegates
                
                table = self.table_all_summary
                table.setRowCount(0); table.setColumnCount(9)
                table.clearSpans()
                headers = ["Set", "Split", "Store Subtotal (%)", "Sales Subtotal (%)", "Area", "Min DA", "Max DA", "Store Count (%)", "Sales Contribution (%)"]
                table.setHorizontalHeaderLabels(headers)
                
                # Reset old delegates for non-bar columns
                table.setItemDelegateForColumn(4, None)  # Area
                table.setItemDelegateForColumn(5, None)  # Min DA
                table.setItemDelegateForColumn(6, None)  # Max DA
                
                # Apply data bar delegates to specific columns (rearranged)
                table.setItemDelegateForColumn(2, orange_del)  # Store Subtotal (%)
                table.setItemDelegateForColumn(3, green_del)   # Sales Subtotal (%)
                table.setItemDelegateForColumn(7, orange_del)  # Store Count (%)
                table.setItemDelegateForColumn(8, green_del)   # Sales Contribution (%)
                
                sidf["Store_Island_Status"] = sidf["Store_Island_Status"].fillna("Unknown")
                
                sets = ["Set 1", "Set 2", "Set 3", "Set 4", "Set 5"]
                curr_row = 0
                
                for set_idx, set_key in enumerate(sets):
                    # Group by splitting category and status, aggregating store count, total sales, and min/max Display Area
                    grouped = sidf.groupby([set_key, "Store_Island_Status"]).agg(
                        Store_Count=("Store_Code", "nunique"),
                        Sales_Sum=("Sales Amount", "sum"),
                        Min_DA=("Store_Display_Area_num", "min"),
                        Max_DA=("Store_Display_Area_num", "max")
                    ).reset_index()
                    
                    # Calculate subtotals for each splitting category
                    subtotals_count = sidf.groupby(set_key)["Store_Code"].nunique().to_dict()
                    subtotals_sales = sidf.groupby(set_key)["Sales Amount"].sum().to_dict()
                    total_count = sidf["Store_Code"].nunique()
                    total_sales = sidf["Sales Amount"].sum()
                    
                    # Add rows for data + grand total
                    table.setRowCount(table.rowCount() + len(grouped) + 1)
                    start_row = curr_row
                    
                    categories = grouped[set_key].unique()
                    
                    # Populate Set column (Col 0) with the set name
                    set_item = QTableWidgetItem(str(set_key))
                    font = set_item.font(); font.setBold(True); set_item.setFont(font)
                    set_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(curr_row, 0, set_item)
                    
                    for cat in categories:
                        cat_rows = grouped[grouped[set_key] == cat]
                        cat_len = len(cat_rows)
                        cat_subtotal_cnt = subtotals_count.get(cat, 0)
                        cat_subtotal_sales = subtotals_sales.get(cat, 0)
                        
                        # Add Split category name -> Column 1 (displays `<7,500` etc.)
                        cat_item = QTableWidgetItem(str(cat))
                        font = cat_item.font(); font.setBold(True); cat_item.setFont(font)
                        cat_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        table.setItem(curr_row, 1, cat_item)
                        
                        # Store Subtotal (%) -> Column 2
                        pct = (cat_subtotal_cnt / total_count * 100) if total_count > 0 else 0
                        sub_item = QTableWidgetItem(f"{cat_subtotal_cnt:,d} ({pct:.1f}%)")
                        font = sub_item.font(); font.setBold(True); sub_item.setFont(font)
                        sub_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        table.setItem(curr_row, 2, sub_item)
                        
                        # Sales Subtotal (%) -> Column 3
                        sales_sub_pct = (cat_subtotal_sales / total_sales * 100) if total_sales > 0 else 0
                        sales_sub_item = QTableWidgetItem(f"{cat_subtotal_sales:,.0f} ({sales_sub_pct:.1f}%)")
                        font = sales_sub_item.font(); font.setBold(True); sales_sub_item.setFont(font)
                        sales_sub_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        table.setItem(curr_row, 3, sales_sub_item)
                        
                        # Populate rows for this category
                        for i, (_, row) in enumerate(cat_rows.iterrows()):
                            # Populate Set Column (Col 0) for sub-rows (matches first row for Excel clarity)
                            if curr_row + i > start_row:
                                table.setItem(curr_row + i, 0, QTableWidgetItem(str(set_key)))
                            
                            # Area -> Column 4
                            table.setItem(curr_row + i, 4, QTableWidgetItem(str(row["Store_Island_Status"])))
                            
                            # Min DA -> Column 5
                            min_val = row["Min_DA"]
                            min_str = f"{min_val:,.0f}" if pd.notnull(min_val) else "-"
                            min_item = QTableWidgetItem(min_str)
                            min_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            table.setItem(curr_row + i, 5, min_item)
                            
                            # Max DA -> Column 6
                            max_val = row["Max_DA"]
                            max_str = f"{max_val:,.0f}" if pd.notnull(max_val) else "-"
                            max_item = QTableWidgetItem(max_str)
                            max_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            table.setItem(curr_row + i, 6, max_item)
                            
                            # Store Count (%) -> Column 7
                            val_count = int(row['Store_Count'])
                            cnt_pct = (val_count / total_count * 100) if total_count > 0 else 0
                            cnt_item = QTableWidgetItem(f"{val_count:,d} ({cnt_pct:.1f}%)")
                            cnt_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            table.setItem(curr_row + i, 7, cnt_item)
                            
                            # Sales Contribution (%) -> Column 8
                            val_sales = float(row['Sales_Sum'])
                            sales_pct = (val_sales / total_sales * 100) if total_sales > 0 else 0
                            sales_item = QTableWidgetItem(f"{val_sales:,.0f} ({sales_pct:.1f}%)")
                            sales_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            table.setItem(curr_row + i, 8, sales_item)
                            
                        # Apply category-level spans
                        if cat_len > 1:
                            table.setSpan(curr_row, 1, cat_len, 1) # Split category column
                            table.setSpan(curr_row, 2, cat_len, 1) # Store Subtotal (%) column
                            table.setSpan(curr_row, 3, cat_len, 1) # Sales Subtotal (%) column
                            
                        curr_row += cat_len
                        
                    # Span the whole Set column (Col 0) for all data rows!
                    if len(grouped) > 1:
                        table.setSpan(start_row, 0, len(grouped), 1)

                    # Grand Total Row
                    item_gt = QTableWidgetItem("Grand Total")
                    font_gt = item_gt.font(); font_gt.setBold(True); item_gt.setFont(font_gt)
                    table.setItem(curr_row, 0, item_gt)
                    table.setItem(curr_row, 1, QTableWidgetItem(""))
                    
                    # Store Subtotal gt -> Column 2
                    gt_pct_item = QTableWidgetItem(f"{total_count:,d} (100.0%)")
                    font_gt_pct = gt_pct_item.font(); font_gt_pct.setBold(True); gt_pct_item.setFont(font_gt_pct)
                    gt_pct_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(curr_row, 2, gt_pct_item)
                    
                    # Sales Subtotal gt -> Column 3
                    sales_gt_sub = QTableWidgetItem(f"{total_sales:,.0f} (100.0%)")
                    font_sales_sub = sales_gt_sub.font(); font_sales_sub.setBold(True); sales_gt_sub.setFont(font_sales_sub)
                    sales_gt_sub.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(curr_row, 3, sales_gt_sub)
                    
                    # Store Count gt -> Column 7
                    cnt_gt = QTableWidgetItem(f"{total_count:,d} (100.0%)")
                    font_cnt = cnt_gt.font(); font_cnt.setBold(True); cnt_gt.setFont(font_cnt)
                    cnt_gt.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(curr_row, 7, cnt_gt)
                    
                    # Sales Contribution gt -> Column 8
                    sales_gt = QTableWidgetItem(f"{total_sales:,.0f} (100.0%)")
                    font_sales = sales_gt.font(); font_sales.setBold(True); sales_gt.setFont(font_sales)
                    sales_gt.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(curr_row, 8, sales_gt)
                    
                    # Span "Grand Total" label across column 0 and column 1
                    table.setSpan(curr_row, 0, 1, 2)
                    
                    for c in range(9):
                        item = table.item(curr_row, c)
                        if not item:
                            item = QTableWidgetItem("")
                            table.setItem(curr_row, c, item)
                        item.setBackground(__import__('PyQt6.QtGui', fromlist=['QColor']).QColor("#e2e8f0"))
                    
                    curr_row += 1
                    
                    # Add 2 empty row gaps unless it's the last set
                    if set_idx < len(sets) - 1:
                        table.setRowCount(table.rowCount() + 2)
                        for _ in range(2):
                            for c in range(9):
                                empty_item = QTableWidgetItem("")
                                empty_item.setFlags(Qt.ItemFlag.NoItemFlags) # Make them non-selectable
                                table.setItem(curr_row, c, empty_item)
                            table.setSpan(curr_row, 0, 1, 9)
                            curr_row += 1
                        
                table.setColumnWidth(0, 100)
                table.setColumnWidth(1, 140)
                table.setColumnWidth(2, 135)
                table.setColumnWidth(3, 180)
                table.setColumnWidth(4, 100)
                table.setColumnWidth(5, 90)
                table.setColumnWidth(6, 90)
                table.setColumnWidth(7, 130)
                table.setColumnWidth(8, 180)

                # --- UPDATE STORE LIST WITH SALES ---
                try:
                    sidf["Sales Amount"] = 0
                    if not sales_df.empty:
                        s_col = "STORE" if "STORE" in sales_df.columns else "M_STORE"
                        v_col = "AVG_TOP2_TOTAL_AMT_SALES"
                        if s_col in sales_df.columns and v_col in sales_df.columns:
                            clean_s_col = sales_df[s_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
                            clean_v_col = pd.to_numeric(sales_df[v_col].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
                            
                            temp_df = pd.DataFrame({s_col: clean_s_col, v_col: clean_v_col})
                            store_sales = temp_df.groupby(s_col)[v_col].sum().reset_index()
                            
                            sidf["Sales Amount"] = sidf["Store_Code"].map(store_sales.set_index(s_col)[v_col]).fillna(0)
                    
                    self.full_store_list_df = sidf.copy() # Update full data
                    self.apply_store_list_filters() # Re-apply filters on new data
                except Exception as sle:
                    print(f"Store List Sales Update Error: {sle}")
                    
            except Exception as e:
                print(f"Error updating All Summary pivot: {e}")
            if not sales_df.empty:
                self._updating_perf_filters = True
                try:
                    if "Dept Type" in sales_df.columns:
                        cur_type = self.perf_filter_type.currentText()
                        self.perf_filter_type.clear(); self.perf_filter_type.addItems(["ALL"] + sorted(sales_df["Dept Type"].dropna().unique().tolist()))
                        idx_t = self.perf_filter_type.findText(cur_type); self.perf_filter_type.setCurrentIndex(idx_t if idx_t >= 0 else 0)
                    
                    if "Dept Details" in sales_df.columns:
                        cur_checked = self.perf_filter_details.checked_items()
                        self.perf_filter_details.clear()
                        details_list = sorted(sales_df["Dept Details"].dropna().unique().tolist())
                        for d in details_list:
                            was_checked = d in cur_checked or (not cur_checked)
                            self.perf_filter_details.add_checkable_item(d, checked=was_checked)
                finally:
                    self._updating_perf_filters = False
                print(f"[TIMING] perf filters: {time.time()-t0:.2f}s"); t0=time.time()

                self.generate_performance_pivot(sales_df)
                print(f"[TIMING] generate_performance_pivot: {time.time()-t0:.2f}s"); t0=time.time()
                
                hidden_cols = ["M_STORE", "M_STORE_DEPARTMENT", "Turnover_Val", "M_STORE_JOIN", 
                               "Normalized Dept", "Dept Details", "COMBINE", "Display Area", "Store Type", "M_POSSYS_DEPT"]
                display_df = sales_df.drop(columns=[c for c in hidden_cols if c in sales_df.columns])
                display_df = display_df.rename(columns={
                    "AVG_TOP2_TOTAL_AMT_SALES": "Sales Amt",
                    "AVG_TOP2_TOTAL_AMT_BALANCE": "Bal Amt",
                    "STORE": "Store",
                    "DEPARTMENT": "Department",
                    "COMBINE": "Combine",
                    "Set1_DA": "Set 1",
                    "Set2_StoreType": "Set 2",
                    "Set3_Type_DA": "Set 3",
                    "Set4_CurrentSetting": "Set 4",
                    "Set5_Mall_SA_Split": "Set 5"
                })
                
                # PREVENT FREEZE: Render max 500 rows to the UI preview grid (Exports & Pivots still use 100%)
                MAX_ROWS = 500
                if len(display_df) > MAX_ROWS:
                    preview_data = display_df.head(MAX_ROWS)
                    self.populate_summary_table(self.table_sales, preview_data)
                else:
                    self.populate_summary_table(self.table_sales, display_df)
                    
                print(f"[TIMING] table_sales ({len(display_df)} total rows, previewed max {MAX_ROWS}): {time.time()-t0:.2f}s"); t0=time.time()
            else:
                self.table_sales.setRowCount(0); self.table_sales.setColumnCount(0)
                self.table_perf_left.setRowCount(0); self.table_perf_left.setColumnCount(0)
                self.table_perf_right.setRowCount(0); self.table_perf_right.setColumnCount(0)
                
            # Get total from master list if available
            if hasattr(self, 'excluded_df') and self.excluded_df is not None:
                if "Store_Code" in self.excluded_df.columns:
                    master_total = total_unique + self.excluded_df["Store_Code"].dropna().nunique()
                else:
                    master_total = total_unique + len(self.excluded_df)
            else:
                master_total = total_unique
            self.status_store_count.setText(f"Analyzed: {total_unique:,} / Total: {master_total:,}")
            self.populate_grouping_analysis_table()
            self.populate_grouping_summary_table()
            self.apply_filter_to_overall()
            print(f"[TIMING] apply_filter_to_overall: {time.time()-t0:.2f}s")
            
        except Exception as e:
            print(f"UI Update Error: {e}")
        finally:
            self.btn_run.setEnabled(True)
            self.progress.setRange(0, 100); self.progress.setValue(100)
            self.progress.setFormat("Completed")
            self.status_msg.setText("Ready")

    def on_all_summary_cell_clicked(self, row, col):
        if not hasattr(self, "full_store_list_df") or self.full_store_list_df is None or self.full_store_list_df.empty:
            return
            
        table = self.table_all_summary
            
        # Determine the Set key by traversing up column 0
        set_key = ""
        r_set = row
        while r_set >= 0:
            item = table.item(r_set, 0)
            if item and item.text().startswith("Set "):
                set_key = item.text()
                break
            r_set -= 1
            
        if not set_key:
            return
            
        df = self.full_store_list_df.copy()
        
        # Traverse upwards to find category name in Column 1 (Split)
        cat_text = ""
        r = row
        # Check first if this is the Grand Total row
        first_item = table.item(row, 0)
        if first_item and first_item.text() == "Grand Total":
            cat_text = "Grand Total"
        else:
            while r >= 0:
                item = table.item(r, 1)
                if item and item.text():
                    cat_text = item.text()
                    break
                r -= 1
            
        if not cat_text:
            return
            
        title = f"{set_key} "
        
        # Clear all active store list filters first for a clean slate
        self.sl_active_filters.clear()
        
        if cat_text == "Grand Total":
            # Entire Fleet
            title += " (Grand Total)"
        else:
            # Check if clicked Column is spanned subtotal (Col 1, 2, 3) or Set col (Col 0)
            if col in [0, 1, 2, 3]:
                # Category Subtotal only
                title += f"» {cat_text} (All Areas)"
                self.sl_active_filters[set_key.upper()] = cat_text
            else:
                # Detail row (Area, Count, or Contribution)
                area_item = table.item(row, 4)
                area_text = area_item.text() if area_item else ""
                if area_text:
                    title += f"» {cat_text} » {area_text}"
                    self.sl_active_filters[set_key.upper()] = cat_text
                    self.sl_active_filters["STORE_ISLAND_STATUS"] = area_text
                else:
                    title += f"» {cat_text}"
                    self.sl_active_filters[set_key.upper()] = cat_text
                    
        # Update and show the dynamic blue warning banner in Store List tab
        self.sl_filter_banner_lbl.setText(f"<b>Active Filter from All Summary:</b> {title}")
        self.sl_filter_banner.setVisible(True)
        
        # Apply the filters to the Store List view
        self.apply_store_list_filters()
        
        # Switch tab index to the "Store List" tab widget
        self.tabs.setCurrentWidget(self.store_list_container)

    def on_perf_filter_changed(self):
        if getattr(self, '_updating_perf_filters', False): return
        if hasattr(self, 'last_sales_df'):
            self.generate_performance_pivot(self.last_sales_df)

    def on_grouping_filter_changed(self):
        if not self.grouping_search:
            return
        search_text = self.grouping_search.text().lower()
        
        table = self.table_grouping_analysis
        for r in range(table.rowCount()):
            match = True
            
            if search_text:
                dept_code_item = table.item(r, 0)
                dept_name_item = table.item(r, 1)
                code_text = dept_code_item.text().lower() if dept_code_item else ""
                name_text = dept_name_item.text().lower() if dept_name_item else ""
                if search_text not in code_text and search_text not in name_text:
                    match = False
                    
            table.setRowHidden(r, not match)

    def ensure_full_df_loaded(self):
        if hasattr(self, 'full_df') and self.full_df is not None and not self.full_df.empty:
            return self.full_df
        if self.possys_path and self.store_path:
            try:
                import pandas as pd
                df, excluded_df = engine.load_and_merge(os.path.dirname(self.possys_path), self.possys_path, self.store_path)
                self.full_df = df
                self.excluded_df = excluded_df
                return df
            except Exception as e:
                print(f"[Engine] Error loading full_df synchronously: {e}")
        import pandas as pd
        return pd.DataFrame()

    def ensure_sales_df_loaded(self):
        if hasattr(self, 'sales_df') and self.sales_df is not None and not self.sales_df.empty:
            return self.sales_df
            
        import pandas as pd
        import numpy as np
        import os
        from pathlib import Path
        
        sales_df = pd.DataFrame()
        s_paths = [p for p in [self.sales_path, self.sales_new_path] if p]
        s_dfs = []
        for sp in s_paths:
            if os.path.exists(sp):
                try:
                    excel_file = pd.ExcelFile(sp, engine='calamine')
                    s_name = "Sales & Balance (No N Dept)"
                    if s_name not in excel_file.sheet_names:
                        if len(excel_file.sheet_names) > 1:
                            s_name = excel_file.sheet_names[1]
                        else:
                            s_name = excel_file.sheet_names[0]
                    
                    temp_df = pd.read_excel(sp, sheet_name=s_name, engine='calamine', dtype=str)
                    temp_df.columns = [str(c).strip() for c in temp_df.columns]
                    
                    rename_map = {}
                    for c in temp_df.columns:
                        c_up = c.upper()
                        if c_up in ["SDTL_STORE", "M_STORE", "STORE_CODE"]:
                            rename_map[c] = "STORE"
                        elif c_up in ["SALES_AMT_PER_30DAYS", "SALES_AMT"]:
                            rename_map[c] = "AVG_TOP2_TOTAL_AMT_SALES"
                        elif c_up in ["BALANCE_AMT"]:
                            rename_map[c] = "AVG_TOP2_TOTAL_AMT_BALANCE"
                    if rename_map:
                        temp_df = temp_df.rename(columns=rename_map)
                        
                    cols_to_drop = [c for c in temp_df.columns if c.upper() == "DATECOUNT"]
                    if cols_to_drop:
                        temp_df = temp_df.drop(columns=cols_to_drop)
                        
                    s_dfs.append(temp_df)
                except Exception as se:
                    print(f"[Engine] Error loading sales file {sp} synchronously: {se}")
        
        if s_dfs:
            sales_df = pd.concat(s_dfs, ignore_index=True)
            sales_df.columns = [str(c).strip() for c in sales_df.columns]
            
            # Enrich sales data
            self.ensure_full_df_loaded()
            if hasattr(self, 'full_df') and self.full_df is not None and not self.full_df.empty:
                dept_info = self.full_df[['M_STORE', 'M_STORE_DEPARTMENT']].drop_duplicates('M_STORE')
            else:
                dept_info = pd.DataFrame()
            
            sales_df = engine.enrich_sales_df(sales_df, dept_info, self.details_path, self.country)
            
            try:
                sales_df = engine.lookup_store_sets(sales_df, self.store_path, {
                    'da_low': 7500, 'da_high': 9500, 'sa_split': 8000, 'da_slicer': 8000,
                    'set5_mall': 10000, 'set5_sa': 8000
                })
            except Exception as e:
                print(f"[Engine] Error in store sets lookup during sync load: {e}")
                for c in ["Display Area", "Store Type", "Set1_DA", "Set2_StoreType", "Set3_Type_DA", "Set4_CurrentSetting", "Set5_Mall_SA_Split"]:
                    sales_df[c] = "Error"
            
            self.sales_df = sales_df
            self.last_sales_df = sales_df
            return sales_df
        return pd.DataFrame()

    def populate_grouping_analysis_table(self):
        """Read department details, sales, and simulation data to populate
        self.table_grouping_analysis dynamically and flexibly for all countries."""
        try:
            import python_calamine
            import pandas as pd
            import numpy as np

            # Read current threshold values unconditionally to avoid UnboundLocalError
            try:
                da_low = int(float(str(self.v_da_low.input.text()).replace(',', '')))
            except:
                da_low = 7500
            try:
                da_high = int(float(str(self.v_da_high.input.text()).replace(',', '')))
            except:
                da_high = 9500
            try:
                sa_split = int(float(str(self.v_sa_split.input.text()).replace(',', '')))
            except:
                sa_split = 8000
            try:
                da_slicer = int(float(str(self.v_da_slicer.input.text()).replace(',', '')))
            except:
                da_slicer = 8000

            def clean_dept_code(val):
                if val is None or pd.isna(val):
                    return ""
                s = str(val).strip()
                if s.endswith(".0"):
                    prefix = s[:-2]
                    if prefix.isdigit():
                        return prefix
                return s

            def dept_sort_key(row):
                code = str(clean_dept_code(row[0])).strip() if row[0] else ""
                num_str = "".join(filter(str.isdigit, code))
                base_num = int(num_str) if num_str else float('inf')
                # L-prefix codes (L881, L894...) go after everything else
                has_l_prefix = 1 if code.upper().startswith("L") else 0
                return (has_l_prefix, base_num, code.lower())

            # Ensure baseline network data is loaded synchronously
            self.ensure_full_df_loaded()
            self.ensure_sales_df_loaded()
            n_cols = 30
            row7 = [""] * n_cols
            row8 = [""] * n_cols
            row9 = [""] * n_cols
            
            # Setup base headers expected by UI
            row7[0] = "Department Desc"
            row7[2] = "Sales Contribution"
            row7[4] = "Set 1"
            row7[9] = "Set 2"
            row7[13] = "Set 3"
            row7[18] = "Set 4"
            row7[23] = "Set 5"

            row8[7] = "Evaluation"
            row8[11] = "Evaluation"
            row8[16] = "Evaluation"
            row8[21] = "Evaluation"
            row8[27] = "Evaluation"

            # Pre-calculate self.overall_df if input paths exist but it is not yet populated
            if (not hasattr(self, 'overall_df') or self.overall_df is None or self.overall_df.empty) and self.possys_path and self.store_path:
                print("[Grouping Analysis] self.overall_df not found. Pre-calculating synchronously...")
                try:
                    set5_mall = 10000
                    try:
                        set5_mall = int(self.v_set5_mall.input.text())
                    except Exception:
                        pass
                    set5_sa = 8000
                    try:
                        set5_sa = int(self.v_set5_sa.input.text())
                    except Exception:
                        pass
                    da_low = 7500
                    try:
                        da_low = int(self.v_da_low.input.text())
                    except Exception:
                        pass
                    da_high = 9500
                    try:
                        da_high = int(self.v_da_high.input.text())
                    except Exception:
                        pass
                    sa_split = 8000
                    try:
                        sa_split = int(self.v_sa_split.input.text())
                    except Exception:
                        pass
                    da_slicer = 8000
                    try:
                        da_slicer = int(self.v_da_slicer.input.text())
                    except Exception:
                        pass
                    
                    df, excluded_df = engine.load_and_merge(os.path.dirname(self.possys_path), self.possys_path, self.store_path)
                    self.excluded_df = excluded_df
                    self.full_df = df
                    overall_df = engine.generate_consolidated_data(df, da_low, da_high, sa_split, da_slicer, set5_mall_split=set5_mall, set5_sa_split=set5_sa)
                    
                    set_names = {
                        "Set1_DA": "Set 1",
                        "Set2_StoreType": "Set 2",
                        "Set3_Type_DA": "Set 3",
                        "Set4_CurrentSetting": "Set 4",
                        "Set5_Mall_SA_Split": "Set 5"
                    }
                    if "Report Set" in overall_df.columns:
                        overall_df["Report Set"] = overall_df["Report Set"].map(set_names).fillna(overall_df["Report Set"])
                    
                    self.overall_df = overall_df
                    print("[Grouping Analysis] Synchronous pre-calculation of self.overall_df completed.")
                except Exception as ex:
                    print(f"[Grouping Analysis] Error in synchronous pre-calculation: {ex}")

            is_dynamic = hasattr(self, 'overall_df') and self.overall_df is not None and not self.overall_df.empty
            cols_12 = ["F1", "F2", "F3", "F4", "F5", "F6", "G1", "G2", "G3", "G4", "G5", "G6"]

            # Load Department Details and remarks from details_path
            dept_remarks = {}
            dept_names = {
                '881': 'Drink', '896': 'FMCG', '920C': 'Direct Disney Corners',
                '920T': 'Direct Disney Theme', '989': 'Bulky Items',
                'L881S': 'Drinks - Short Shelf Life', 'L882S': 'Chocolate & Sweets - Short Shelf Life',
                'L894D': 'Local Dog Products', 'L922': 'Local Premium Dept',
                'L881': 'DRINK', 'L882': 'CHOCOLATE & SWEET', 'L883': 'SNACK & BISCUIT',
                'L883S': 'SNACKS & BISCUITS - SHORT SHELF LIFE', 'L885': 'GROCERY',
                'L885S': 'GROCERY - SHORT SHELF LIFE', 'L894': 'LOCAL PET PRODUCT',
                'L896': 'LOCAL FMCG', 'L896A': 'LOCAL FMCG ALLOWED',
                'L899': 'LOCAL HP AND COMPUTER ACCESSORY', 'L900': 'LOCAL CAR ACCESORY',
                'L901': 'LOCAL TOYS', 'L902': 'LOCAL ELECTRICAL',
                'L903': 'LOCAL CAP/STOCKING/BELT/BAG/SLIPPER', 'L904': 'LOCAL GIFTS',
                'L905': 'LOCAL JEWELLERY & COSMETIC', 'L906': 'LOCAL HOUSEHOLD',
                'L907': 'LOCAL HARDWARE', 'L908': 'LOCAL STATIONARY',
                'L909': 'LOCAL TABLE CLOTH/CURTAIN/SHIRT/SWIM WEAR',
                'L910': 'LOCAL CLOCK / WATCH', 'L920': 'LOCAL DISNEY',
                'L989': 'LOCAL BULKY ITEMS',

                # Base department codes fallback
                '882': 'CHOCOLATE & SWEET', '883': 'SNACK & BISCUIT',
                '883S': 'SNACKS & BISCUITS - SHORT SHELF LIFE', '885': 'GROCERY',
                '885S': 'GROCERY - SHORT SHELF LIFE', '894': 'PET PRODUCT',
                '896A': 'FMCG ALLOWED', '899': 'HP AND COMPUTER ACCESSORY',
                '900': 'CAR ACCESORY', '901': 'TOYS', '902': 'ELECTRICAL',
                '903': 'CAP/STOCKING/BELT/BAG/SLIPPER', '904': 'GIFTS',
                '905': 'JEWELLERY & COSMETIC', '906': 'HOUSEHOLD',
                '907': 'HARDWARE', '908': 'STATIONARY',
                '909': 'TABLE CLOTH/CURTAIN/SHIRT/SWIM WEAR',
                '910': 'CLOCK / WATCH', '920': 'DISNEY'
            }
            
            # ── Load remark directly from Department Details.xlsx ───────────────
            # Bypass engine.load_dept_details entirely — read the file directly
            # so there's no chance of silent failures from URL fallbacks.
            _DEPT_DETAILS_PATH = r"Y:\R&D\Category-based Start Up - Split DB Analysis\Tool Data\Department Details.xlsx"
            _details_candidates = [_DEPT_DETAILS_PATH]
            # Also try any user-specified local paths from self.details_path
            if self.details_path:
                for _p in str(self.details_path).split("|"):
                    _p = _p.strip()
                    if _p and not _p.startswith("http") and _p not in _details_candidates:
                        _details_candidates.append(_p)

            import os as _os
            import openpyxl as _openpyxl

            _loaded_remarks = False
            for _fp in _details_candidates:
                if not _os.path.exists(_fp):
                    print(f"[Grouping Analysis] Remark file not found: {_fp}")
                    continue
                try:
                    _wb = _openpyxl.load_workbook(_fp, data_only=True)
                    c_upper = self.country.upper() if self.country else "ID"

                    # Find the right sheet for this country
                    _sheet_name = None
                    _candidates_sheets = [
                        f"DEPARTMENT {c_upper}",
                        f"DEPARTMENT_{c_upper}",
                        f"{c_upper} REMARK",
                        f"{c_upper}_REMARK",
                        f"DEPARTMENT BR SG",  # fallback for BR/SG
                    ]
                    for _cs in _candidates_sheets:
                        for _s in _wb.sheetnames:
                            if _s.strip().upper() == _cs.upper():
                                _sheet_name = _s
                                break
                        if _sheet_name:
                            break
                    # Secondary: any sheet containing the country code
                    if not _sheet_name:
                        for _s in _wb.sheetnames:
                            if c_upper in _s.upper():
                                _sheet_name = _s
                                break

                    if not _sheet_name:
                        print(f"[Grouping Analysis] No sheet found for country '{c_upper}' in {_fp}. Sheets: {_wb.sheetnames}")
                        continue

                    _ws = _wb[_sheet_name]
                    _rows_data = list(_ws.iter_rows(values_only=True))
                    if not _rows_data:
                        continue

                    _headers = [str(c).strip() if c is not None else "" for c in _rows_data[0]]

                    # Find DEPARTMENT column
                    _dept_col_idx = next(
                        (i for i, h in enumerate(_headers) if "DEPARTMENT" in h.upper()),
                        None
                    )

                    # Find country Remark column: try country-specific first, then generic
                    _remark_col_idx = None
                    for i, h in enumerate(_headers):
                        if c_upper in h.upper() and any(k in h.upper() for k in ["REMARK", "DETAIL", "REM"]):
                            _remark_col_idx = i
                            break
                    if _remark_col_idx is None:
                        # MY special: C2/D2
                        if c_upper == "MY":
                            for i, h in enumerate(_headers):
                                if any(k in h.upper() for k in ["C2", "D2", "REMARK"]):
                                    _remark_col_idx = i
                                    break
                    if _remark_col_idx is None:
                        # Generic fallback
                        for i, h in enumerate(_headers):
                            if any(k in h.upper() for k in ["REMARK", "DETAIL", "REM", "COUNTRY"]):
                                _remark_col_idx = i
                                break

                    # Find DESCRIPTION column
                    _desc_col_idx = next(
                        (i for i, h in enumerate(_headers) if "DESCRIPTION" in h.upper() or "DESC" in h.upper()),
                        None
                    )

                    print(f"[Grouping Analysis] File={_fp}, Sheet={_sheet_name}")
                    print(f"[Grouping Analysis] dept_col_idx={_dept_col_idx} ({_headers[_dept_col_idx] if _dept_col_idx is not None else 'N/A'}), "
                          f"remark_col_idx={_remark_col_idx} ({_headers[_remark_col_idx] if _remark_col_idx is not None else 'N/A'})")

                    if _dept_col_idx is None or _remark_col_idx is None:
                        print("[Grouping Analysis] Could not detect dept or remark column, skipping.")
                        continue

                    for _row in _rows_data[1:]:
                        if len(_row) <= max(_dept_col_idx, _remark_col_idx):
                            continue
                        # Clean dept code: strip .0 suffix, strip leading/trailing spaces
                        _raw_dept = _row[_dept_col_idx]
                        if _raw_dept is None:
                            continue
                        _d_str = str(_raw_dept).strip()
                        if _d_str.endswith(".0"):
                            _d_str = _d_str[:-2]
                        _d_code = _d_str.strip()
                        if not _d_code or _d_code.lower() in ["none", "nan", "department", ""]:
                            continue

                        _raw_rem = _row[_remark_col_idx]
                        _d_rem = str(_raw_rem).strip() if _raw_rem is not None else ""

                        if _d_rem and _d_rem.lower() not in ["none", "nan", "unknown", ""]:
                            dept_remarks[_d_code] = _d_rem
                            # Also map the L-prefix variant
                            _local = f"L{_d_code}" if not _d_code.upper().startswith("L") else _d_code
                            if _local not in dept_remarks:
                                dept_remarks[_local] = _d_rem

                        # Description
                        if _desc_col_idx is not None and len(_row) > _desc_col_idx:
                            _raw_desc = _row[_desc_col_idx]
                            _d_desc = str(_raw_desc).strip() if _raw_desc is not None else ""
                            if _d_desc and _d_desc.lower() not in ["none", "nan", ""]:
                                dept_names[_d_code] = _d_desc

                    print(f"[Grouping Analysis] dept_remarks loaded: {len(dept_remarks)} entries from {_sheet_name}")
                    _loaded_remarks = True
                    break  # Stop after first successful file
                except Exception as _ex:
                    print(f"[Grouping Analysis] Error reading {_fp}: {_ex}")

            if not _loaded_remarks:
                print("[Grouping Analysis] WARNING: Could not load any remark data — Remark column will be empty.")

            # Get unique departments
            unique_depts_set = set()
            if is_dynamic:
                sim_depts = list(self.overall_df["Department"].astype(str).str.strip().unique())
                for d in sim_depts:
                    clean_d = clean_dept_code(d)
                    if clean_d:
                        unique_depts_set.add(clean_d)
            elif self.possys_path and os.path.exists(self.possys_path):
                try:
                    sim_wb = python_calamine.CalamineWorkbook.from_path(self.possys_path)
                    if sim_wb.sheet_names:
                        sim_sheet = sim_wb.get_sheet_by_name(sim_wb.sheet_names[0])
                        sim_rows = sim_sheet.to_python()
                        if len(sim_rows) > 1:
                            s_headers = [str(c).strip().upper() for c in sim_rows[1]]
                            dept_col_idx = s_headers.index("DEPT") if "DEPT" in s_headers else s_headers.index("DEPARTMENT") if "DEPARTMENT" in s_headers else 1
                            for s_row in sim_rows[2:]:
                                if len(s_row) > dept_col_idx and s_row[dept_col_idx] is not None:
                                    clean_d = clean_dept_code(s_row[dept_col_idx])
                                    if clean_d:
                                        unique_depts_set.add(clean_d)
                except Exception as ex:
                    print(f"[Grouping Analysis] Error scanning unique depts from possys: {ex}")

            unique_depts = sorted(list(unique_depts_set), key=lambda d: dept_sort_key([d]))

            # ── Filter: only show departments with remark Dept / Main_Category / Sub_Category ──
            _SHOW_REMARKS = {"dept", "main_category", "sub_category"}
            if dept_remarks:
                unique_depts = [
                    d for d in unique_depts
                    if dept_remarks.get(d, "").strip().lower().replace(" ", "_") in _SHOW_REMARKS
                ]
                print(f"[Grouping Analysis] After remark filter: {len(unique_depts)} depts shown")

            # Retrieve sales amounts
            sales_sums = {}
            if hasattr(self, 'sales_df') and self.sales_df is not None and not self.sales_df.empty:
                s_df = self.sales_df
                dept_col = next((c for c in s_df.columns if "DEPARTMENT" in c.upper()), "Department")
                amt_col = "AVG_TOP2_TOTAL_AMT_SALES" if "AVG_TOP2_TOTAL_AMT_SALES" in s_df.columns else "Sales Amt"
                if dept_col in s_df.columns and amt_col in s_df.columns:
                    s_df_clean = s_df.copy()
                    s_df_clean[dept_col] = s_df_clean[dept_col].apply(clean_dept_code)
                    s_df_clean[amt_col] = pd.to_numeric(s_df_clean[amt_col].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0.0)
                    sales_sums = s_df_clean.groupby(dept_col)[amt_col].sum().to_dict()

            total_sales = sum(v for k, v in sales_sums.items() if k != "Grand Total" and not k.startswith("Total"))
            if total_sales <= 0:
                total_sales = 1.0

            # ── Filter: exclude departments with 0 sales ──────────────────────────
            if sales_sums:
                unique_depts = [
                    d for d in unique_depts
                    if sales_sums.get(d, 0) > 0
                ]
                print(f"[Grouping Analysis] After zero-sales filter: {len(unique_depts)} depts shown")

            # Normalize helper for category string comparisons
            def normalize_category(s):
                if s is None:
                    return ""
                s = str(s).strip().lower()
                s = s.replace(" ", "").replace(",", "").replace(".", "")
                return s

            # Dynamic matching helper based on column index c_idx
            def match_category_row(c_idx, set_df, sheet_cat_val):
                if set_df.empty:
                    return None
                norm_sheet_val = normalize_category(sheet_cat_val)
                for _, r in set_df.iterrows():
                    if normalize_category(r["Category"]) == norm_sheet_val:
                        return r

                for _, r in set_df.iterrows():
                    cat_name = str(r["Category"]).strip().lower()
                    if c_idx == 4 and "<" in cat_name and ">=" not in cat_name and "<=" not in cat_name: return r
                    elif c_idx == 5 and ("<=" in cat_name or (">=" in cat_name and "<" in cat_name)): return r
                    elif c_idx == 6 and ">=" in cat_name and "<" not in cat_name: return r
                    elif c_idx == 9 and "mall" in cat_name: return r
                    elif c_idx == 10 and ("sa" in cat_name or "standalone" in cat_name): return r
                    elif c_idx == 13 and "sa" in cat_name and "<" in cat_name: return r
                    elif c_idx == 14 and "sa" in cat_name and ">=" in cat_name: return r
                    elif c_idx == 15 and "mall" in cat_name: return r
                    elif c_idx == 18 and ("+" in cat_name or ("mall" in cat_name and "sa" in cat_name) or ("sa" in cat_name and "<" in cat_name and "mall" not in cat_name)): return r
                    elif c_idx == 19 and "sa" in cat_name and ">=" in cat_name and "mall" not in cat_name: return r
                    elif c_idx == 20 and "mall" in cat_name and ">=" in cat_name: return r
                    elif c_idx == 23 and "sa" in cat_name and "<" in cat_name: return r
                    elif c_idx == 24 and "sa" in cat_name and ">=" in cat_name: return r
                    elif c_idx == 25 and "mall" in cat_name and "<" in cat_name: return r
                    elif c_idx == 26 and "mall" in cat_name and ">=" in cat_name: return r

                return set_df.iloc[0]

            # Construct dynamic rows for all unique departments
            data_rows = []
            for d_code in unique_depts:
                # Resolve department description with prefix pairings safety fallbacks
                dept_desc = dept_names.get(d_code)
                if not dept_desc:
                    if f"L{d_code}" in dept_names:
                        l_desc = dept_names[f"L{d_code}"]
                        if l_desc.upper().startswith("LOCAL "):
                            dept_desc = l_desc[6:].strip()
                        else:
                            dept_desc = l_desc
                    elif d_code.upper().startswith("L") and d_code[1:] in dept_names:
                        base_desc = dept_names[d_code[1:]]
                        dept_desc = f"LOCAL {base_desc}"
                
                if not dept_desc:
                    dept_desc = f"Dept {d_code}"
                
                dept_desc = dept_desc.upper()
                sales_amount = sales_sums.get(d_code, 0.0)
                sales_pct = sales_amount / total_sales

                row_vals = [None] * n_cols
                row_vals[0] = d_code
                row_vals[1] = dept_desc
                row_vals[2] = sales_amount
                row_vals[3] = sales_pct

                if is_dynamic:
                    dept_df = self.overall_df[self.overall_df["Department"].apply(clean_dept_code) == d_code]

                    # Helper: compute final score from F1-G6 values using grouping_score.py logic
                    def _score(matched_row):
                        y = [float(matched_row[col]) if matched_row[col] is not None else 0.0
                             for col in cols_12]
                        b   = calculate_balance_score(y)
                        sym = calculate_symmetry_score(y)
                        ctr = calculate_center_score(y)
                        return calculate_final_score(b, sym, ctr)

                    def _get_score(c, set_name):
                        if dept_df.empty:
                            return 0.0
                        set_df = dept_df[dept_df["Report Set"] == set_name]
                        matched = match_category_row(c, set_df, row9[c]) if not set_df.empty else None
                        return _score(matched) if matched is not None else 0.0

                    # ── Set 1: cols 4, 5, 6 │ Eval=7 │ Score=8 ────────────────────────
                    s1 = [_get_score(c, "Set 1") for c in [4, 5, 6]]
                    for c, sc in zip([4, 5, 6], s1):
                        row_vals[c] = sc
                    row_vals[8] = sum(s1)                    # Set 1 raw score total
                    row_vals[7] = row_vals[8] * sales_pct   # Set 1 weighted evaluation

                    # ── Set 2: cols 9, 10  │ Eval=11 │ Score=12 ───────────────────────
                    s2 = [_get_score(c, "Set 2") for c in [9, 10]]
                    for c, sc in zip([9, 10], s2):
                        row_vals[c] = sc
                    row_vals[12] = sum(s2) * 1.5
                    row_vals[11] = row_vals[12] * sales_pct

                    # ── Set 3: cols 13, 14, 15 │ Eval=16 │ Score=17 ──────────────────
                    s3 = [_get_score(c, "Set 3") for c in [13, 14, 15]]
                    for c, sc in zip([13, 14, 15], s3):
                        row_vals[c] = sc
                    row_vals[17] = sum(s3)
                    row_vals[16] = row_vals[17] * sales_pct

                    # ── Set 4: cols 18, 19, 20 │ Eval=21 │ Score=22 ──────────────────
                    s4 = [_get_score(c, "Set 4") for c in [18, 19, 20]]
                    for c, sc in zip([18, 19, 20], s4):
                        row_vals[c] = sc
                    row_vals[22] = sum(s4)
                    row_vals[21] = row_vals[22] * sales_pct

                    # ── Set 5: cols 23, 24, 25, 26 │ Eval=27 │ Score=28 ──────────────
                    s5 = [_get_score(c, "Set 5") for c in [23, 24, 25, 26]]
                    for c, sc in zip([23, 24, 25, 26], s5):
                        row_vals[c] = sc
                    row_vals[28] = sum(s5) * 0.75
                    row_vals[27] = row_vals[28] * sales_pct
                else:
                    # Fill blank values for Set columns when simulation not run yet
                    for i in range(4, 29):
                        row_vals[i] = ""

                # Populate remark column at index 29 (Allowed: Dept / Main_Category / Sub_Category only)
                d_rem = dept_remarks.get(d_code)
                if not d_rem:
                    # Strip trailing N (new store suffix) and retry
                    clean_d = d_code.rstrip("N") if d_code.endswith("N") and len(d_code) > 1 else d_code
                    d_rem = dept_remarks.get(clean_d)
                if not d_rem and not d_code.upper().startswith("L"):
                    # Try local (L-prefix) variant
                    d_rem = dept_remarks.get(f"L{d_code}")
                if not d_rem and d_code.upper().startswith("L"):
                    # Try base (strip L-prefix) variant
                    d_rem = dept_remarks.get(d_code[1:])
                row_vals[29] = d_rem if d_rem else ""
                data_rows.append(row_vals)

            # Dynamically calculate subgroup store count headers (row8 / Row 1) using unique stores
            try:
                num_da_low = float(str(self.v_da_low.input.text()).replace(',', ''))
            except: num_da_low = 7500.0
            try:
                num_da_high = float(str(self.v_da_high.input.text()).replace(',', ''))
            except: num_da_high = 9500.0
            try:
                num_sa_split = float(str(self.v_sa_split.input.text()).replace(',', ''))
            except: num_sa_split = 8000.0
            try:
                num_da_slicer = float(str(self.v_da_slicer.input.text()).replace(',', ''))
            except: num_da_slicer = 8000.0
            try:
                num_s5_sa = float(str(self.v_set5_sa.input.text()).replace(',', ''))
            except: num_s5_sa = 8000.0
            try:
                num_s5_mall = float(str(self.v_set5_mall.input.text()).replace(',', ''))
            except: num_s5_mall = 10000.0

            subgroup_store_counts = {}
            if hasattr(self, 'full_df') and self.full_df is not None and not self.full_df.empty:
                unique_stores = self.full_df.drop_duplicates("M_STORE").copy()
                unique_stores["Store_Display_Area_num"] = pd.to_numeric(unique_stores["Store_Display_Area_num"], errors='coerce').fillna(0)
                
                for c in [4, 5, 6, 9, 10, 13, 14, 15, 18, 19, 20, 23, 24, 25, 26]:
                    if c == 4:
                        mask = unique_stores["Store_Display_Area_num"] < num_da_low
                    elif c == 5:
                        mask = (unique_stores["Store_Display_Area_num"] >= num_da_low) & (unique_stores["Store_Display_Area_num"] < num_da_high)
                    elif c == 6:
                        mask = unique_stores["Store_Display_Area_num"] >= num_da_high
                    elif c == 9:
                        mask = unique_stores["Store_Type"].str.lower() == "mall"
                    elif c == 10:
                        mask = unique_stores["Store_Type"].str.lower() == "standalone"
                    elif c == 13:
                        mask = (unique_stores["Store_Type"].str.lower() == "standalone") & (unique_stores["Store_Display_Area_num"] < num_sa_split)
                    elif c == 14:
                        mask = (unique_stores["Store_Type"].str.lower() == "standalone") & (unique_stores["Store_Display_Area_num"] >= num_sa_split)
                    elif c == 15:
                        mask = unique_stores["Store_Type"].str.lower() == "mall"
                    elif c == 18:
                        mask = unique_stores["Store_Display_Area_num"] < num_da_slicer
                    elif c == 19:
                        mask = (unique_stores["Store_Type"].str.lower() == "standalone") & (unique_stores["Store_Display_Area_num"] >= num_da_slicer)
                    elif c == 20:
                        mask = (unique_stores["Store_Type"].str.lower() == "mall") & (unique_stores["Store_Display_Area_num"] >= num_da_slicer)
                    elif c == 23:
                        mask = (unique_stores["Store_Type"].str.lower() == "standalone") & (unique_stores["Store_Display_Area_num"] < num_s5_sa)
                    elif c == 24:
                        mask = (unique_stores["Store_Type"].str.lower() == "standalone") & (unique_stores["Store_Display_Area_num"] >= num_s5_sa)
                    elif c == 25:
                        mask = (unique_stores["Store_Type"].str.lower() == "mall") & (unique_stores["Store_Display_Area_num"] < num_s5_mall)
                    elif c == 26:
                        mask = (unique_stores["Store_Type"].str.lower() == "mall") & (unique_stores["Store_Display_Area_num"] >= num_s5_mall)
                    else:
                        mask = pd.Series(False, index=unique_stores.index)
                    
                    subgroup_store_counts[c] = int(mask.sum())
            else:
                for c in [4, 5, 6, 9, 10, 13, 14, 15, 18, 19, 20, 23, 24, 25, 26]:
                    subgroup_store_counts[c] = 0

            # Update row8 with dynamic store counts
            for c in subgroup_store_counts:
                row8[c] = subgroup_store_counts[c]

            # Find the first department with actual simulation records to get dynamic category names
            first_dept_with_data = None
            if is_dynamic:
                for d_code in unique_depts:
                    dept_df = self.overall_df[self.overall_df["Department"].apply(clean_dept_code) == d_code]
                    if not dept_df.empty:
                        first_dept_with_data = d_code
                        break

            # Update Row 2 (row9) dynamic category names if user changed thresholds
            subgroup_label_mappings = {
                4: f"DA < {da_low}", 5: f"{da_low}<=DA<{da_high}", 6: f"DA >= {da_high}",
                9: "Mall", 10: "Standalone (SA)",
                13: f"SA < {sa_split}", 14: f"SA >= {sa_split}", 15: "Mall",
                18: f"Mall + SA < {da_slicer}", 19: f"SA >= {da_slicer}", 20: f"Mall >= {da_slicer}",
                23: f"SA < {num_s5_sa}", 24: f"SA >= {num_s5_sa}", 25: f"Mall < {num_s5_mall}", 26: f"Mall >= {num_s5_mall}"
            }
            for c in [4, 5, 6, 9, 10, 13, 14, 15, 18, 19, 20, 23, 24, 25, 26]:
                if is_dynamic and first_dept_with_data:
                    dept_df = self.overall_df[self.overall_df["Department"].apply(clean_dept_code) == first_dept_with_data]
                    r_set = "Set 1" if c in [4, 5, 6] else "Set 2" if c in [9, 10] else "Set 3" if c in [13, 14, 15] else "Set 4" if c in [18, 19, 20] else "Set 5"
                    set_df = dept_df[dept_df["Report Set"] == r_set]
                    matched = match_category_row(c, set_df, subgroup_label_mappings[c])
                    if matched is not None:
                        row9[c] = str(matched["Category"])
                    else:
                        row9[c] = subgroup_label_mappings[c]
                else:
                    row9[c] = subgroup_label_mappings[c]

            # Update Row 2 (row9) evaluation and set score column sums
            col_sums = {}
            for c in [7, 8, 11, 12, 16, 17, 21, 22, 27, 28]:
                if is_dynamic:
                    col_sums[c] = sum(r[c] for r in data_rows if r[c] is not None and isinstance(r[c], (int, float)))
                else:
                    col_sums[c] = ""
                row9[c] = col_sums[c]

            # Sort all rows by department code numeric structure
            data_rows.sort(key=dept_sort_key)

            table = self.table_grouping_analysis
            header_table = self.table_grouping_header
            total_ui_cols = 30

            table.clearContents()
            table.clearSpans()
            table.setRowCount(len(data_rows))
            table.setColumnCount(total_ui_cols)

            header_table.clearContents()
            header_table.clearSpans()
            header_table.setRowCount(3)
            header_table.setColumnCount(total_ui_cols)

            # Configure Header heights
            header_table.setRowHeight(0, 30)
            header_table.setRowHeight(1, 30)
            header_table.setRowHeight(2, 32)

            # Helper to map UI column index to the data row index
            def ui_to_data_idx(ui_col):
                if ui_col < 2:
                    return ui_col
                elif ui_col == 2:
                    return -1 # Remark column
                else:
                    return ui_col - 1

            # Populate Header Cells and apply styling
            blue_cols  = [8, 9, 12, 13, 17, 18, 22, 23, 28, 29]
            yellow_cols = [8, 9, 12, 13, 17, 18, 22, 23, 28, 29]

            for c_idx in range(total_ui_cols):
                data_c_idx = ui_to_data_idx(c_idx)

                if c_idx == 2:
                    val0 = "Remark"
                    val1 = ""
                    val2 = ""
                else:
                    # Row 0
                    val0 = str(row7[data_c_idx]).strip() if data_c_idx < len(row7) and data_c_idx >= 0 and row7[data_c_idx] != "" else ""
                    if not val0 and c_idx == 0:
                        val0 = "Department Desc"

                    # Row 1
                    val1_raw = row8[data_c_idx] if data_c_idx < len(row8) and data_c_idx >= 0 else ""
                    if isinstance(val1_raw, (int, float)) and val1_raw == int(val1_raw):
                        val1 = f"{int(val1_raw):,}"
                    else:
                        val1 = str(val1_raw).strip() if val1_raw != "" else ""

                    # Row 2
                    val2_raw = row9[data_c_idx] if data_c_idx < len(row9) and data_c_idx >= 0 else ""
                    if isinstance(val2_raw, (int, float)):
                        val2 = f"{val2_raw:.2f}"
                    else:
                        val2 = str(val2_raw).strip() if val2_raw != "" else ""
                    if not val2 and c_idx == 3:
                        val2 = "Amount"
                    elif not val2 and c_idx == 4:
                        val2 = "%"

                # Create Header items
                item0 = QTableWidgetItem(val0)
                item0.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item0.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                item0.setForeground(QBrush(QColor("#1e293b")))
                item0.setBackground(QBrush(QColor("#ffffff")))
                header_table.setItem(0, c_idx, item0)

                item1 = QTableWidgetItem(val1)
                item1.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item1.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                item1.setForeground(QBrush(QColor("#1e293b")))
                if c_idx in blue_cols:
                    item1.setBackground(QBrush(QColor("#e0f2fe")))
                else:
                    item1.setBackground(QBrush(QColor("#ffffff")))
                header_table.setItem(1, c_idx, item1)

                item2 = QTableWidgetItem(val2)
                item2.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item2.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                item2.setForeground(QBrush(QColor("#1e293b")))
                if c_idx in yellow_cols:
                    item2.setBackground(QBrush(QColor("#ffff00")))
                else:
                    item2.setBackground(QBrush(QColor("#ffffff")))
                header_table.setItem(2, c_idx, item2)

            # Apply Merging Spans on the Header table
            header_table.setSpan(0, 0, 3, 2)  # Dept Code + Name
            header_table.setSpan(0, 2, 3, 1)  # 'Remark'
            header_table.setSpan(0, 3, 2, 2)  # 'Sales Contribution'
            header_table.setSpan(0, 5, 1, 5)  # Set 1
            header_table.setSpan(0, 10, 1, 4) # Set 2
            header_table.setSpan(0, 14, 1, 5) # Set 3
            header_table.setSpan(0, 19, 1, 5) # Set 4
            header_table.setSpan(0, 24, 1, 6) # Set 5

            header_table.setSpan(1, 8, 1, 2)  # Set 1 Evaluation
            header_table.setSpan(1, 12, 1, 2) # Set 2 Evaluation
            header_table.setSpan(1, 17, 1, 2) # Set 3 Evaluation
            header_table.setSpan(1, 22, 1, 2) # Set 4 Evaluation
            header_table.setSpan(1, 28, 1, 2) # Set 5 Evaluation

            # Attach the delegate to Column 4 (Sales Contribution %)
            if not hasattr(self, "_grouping_analysis_delegate"):
                self._grouping_analysis_delegate = DataBarDelegate("#f59e0b")
                
            for col in range(total_ui_cols):
                if col == 4:
                    table.setItemDelegateForColumn(col, self._grouping_analysis_delegate)
                else:
                    table.setItemDelegateForColumn(col, None)

            # Populate data rows
            for r_idx, row_data in enumerate(data_rows):
                for c_idx in range(total_ui_cols):
                    data_c_idx = ui_to_data_idx(c_idx)
                    raw_val = row_data[data_c_idx] if data_c_idx < len(row_data) and data_c_idx >= 0 else ""

                    if c_idx == 2:
                        val = row_data[29] if len(row_data) > 29 and row_data[29] is not None else ""
                        display = str(val).strip()
                    elif c_idx == 0:
                        display = clean_dept_code(raw_val)
                    elif c_idx == 1:
                        display = str(raw_val).strip().upper()
                    elif c_idx == 3:
                        curr_prefix = ""
                        if hasattr(self, 'country') and self.country:
                            currency_map = {
                                "ID": "IDR", "TH": "THB", "MY": "MYR",
                                "BR": "BND", "SG": "SGD", "IN": "INR"
                            }
                            curr_prefix = f"{currency_map.get(self.country, self.country)} "
                        if isinstance(raw_val, (float, int)):
                            display = f"{curr_prefix}{round(raw_val):,}"
                        else:
                            display = f"{curr_prefix}{str(raw_val).strip()}"
                    elif c_idx == 4:
                        if isinstance(raw_val, (float, int)):
                            display = f"{raw_val * 100:.2f}%"
                        else:
                            display = str(raw_val).strip()
                    elif raw_val is None or raw_val == "":
                        display = ""
                    elif isinstance(raw_val, float):
                        display = f"{raw_val:.2f}"
                    else:
                        display = str(raw_val).strip()

                    item = QTableWidgetItem(display)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(r_idx, c_idx, item)

            # Define column widths
            col_widths = {0: 80, 1: 280, 2: 200, 3: 145, 4: 70}
            for i in range(5, total_ui_cols): col_widths[i] = 100

            for i in range(total_ui_cols):
                w = col_widths.get(i, 100)
                table.setColumnWidth(i, w)
                header_table.setColumnWidth(i, w)

            table.horizontalHeader().setStretchLastSection(True)
            header_table.horizontalHeader().setStretchLastSection(True)
            print(f"[Grouping Analysis] Dynamically generated and populated {len(data_rows)} rows.")

        except Exception as e:
            print(f"[Grouping Analysis] Error loading table: {e}")

    def populate_grouping_summary_table(self):
        """Populate the Grouping Analysis Summary table dynamically from network-drive data.
        No local 'Splitting Discussion' file is required. Store counts and sales contributions
        are computed from self.full_df and self.sales_df, which are loaded at startup from the
        network drive. Grouping scores are added when a simulation has been run (self.overall_df).
        Sales Contribution columns are rendered in plain black text."""
        try:
            import pandas as pd
            import numpy as np

            # ── Ensure baseline network data is loaded ──────────────────────────────
            self.ensure_full_df_loaded()
            self.ensure_sales_df_loaded()

            # ── Read current threshold values ────────────────────────────────────────
            def _get_val(widget_attr, default):
                try:
                    return float(str(getattr(self, widget_attr).input.text()).replace(',', ''))
                except Exception:
                    return default

            num_da_low   = _get_val('v_da_low',   7500.0)
            num_da_high  = _get_val('v_da_high',  9500.0)
            num_sa_split = _get_val('v_sa_split',  8000.0)
            num_da_slicer= _get_val('v_da_slicer', 8000.0)
            num_s5_sa    = _get_val('v_set5_sa',   8000.0)
            num_s5_mall  = _get_val('v_set5_mall', 10000.0)

            def fmt(v, decimals=0):
                """Format a threshold number as a comma-separated integer string."""
                try:
                    return f"{int(v):,}"
                except Exception:
                    return str(v)

            d_low   = fmt(num_da_low)
            d_high  = fmt(num_da_high)
            s_split = fmt(num_sa_split)
            d_slicer= fmt(num_da_slicer)
            s5_sa   = fmt(num_s5_sa)
            s5_mall = fmt(num_s5_mall)

            # ── Candidate subgroup definitions ───────────────────────────────────────
            # (Set name, Candidate label, Grouping Analysis column index, Pros/Cons key)
            subgroups_def = [
                ("Set 1", f"DA < {d_low}",             4,  "Set 1"),
                ("Set 1", f"{d_low}<=DA<{d_high}",     5,  "Set 1"),
                ("Set 1", f"DA >= {d_high}",            6,  "Set 1"),

                ("Set 2", "Mall",                       9,  "Set 2"),
                ("Set 2", "Standalone (SA)",            10, "Set 2"),

                ("Set 3", f"SA < {s_split}",            13, "Set 3"),
                ("Set 3", f"SA >= {s_split}",           14, "Set 3"),
                ("Set 3", "Mall",                       15, "Set 3"),

                ("Set 4", f"Mall + SA < {d_slicer}",   18, "Set 4"),
                ("Set 4", f"SA >= {d_slicer}",          19, "Set 4"),
                ("Set 4", f"Mall >= {d_slicer}",        20, "Set 4"),

                ("Set 5", f"SA < {s5_sa}",              23, "Set 5"),
                ("Set 5", f"SA >= {s5_sa}",             24, "Set 5"),
                ("Set 5", f"Mall < {s5_mall}",          25, "Set 5"),
                ("Set 5", f"Mall >= {s5_mall}",         26, "Set 5"),
            ]

            # ── Built-in Pro's / Con's (no local file required) ──────────────────────
            c_upper = str(getattr(self, 'country', '')).strip().upper()
            if c_upper == "ID":
                pros_cons = {
                    "Set 1": {
                        "Pros": "Great grouping distribution.\nSplitting with identical store count.",
                        "Cons": "Ignores store type (Mall vs SA)."
                    },
                    "Set 2": {
                        "Pros": "Easy to separate Mall and Standalone stores.\nGrouping Distribution are ideal when splitting by Mall and SA.",
                        "Cons": "Large size range in one group.\nStore count is heavily biased."
                    },
                    "Set 3": {
                        "Pros": "Good grouping distribution.\nSplitting almost have identical store count.",
                        "Cons": "Large display area for Mall to cover."
                    },
                    "Set 4": {
                        "Pros": "Good grouping distribution.\nDisplay area range is good to cover for each split.",
                        "Cons": "Large store count for Standalone stores."
                    },
                    "Set 5": {
                        "Pros": "Display area range is good to cover for each split.",
                        "Cons": "Large store count for Standalone stores."
                    },
                }
            else:
                # Placeholder for other countries (e.g. TH) so the user can easily put different ones later
                pros_cons = {
                    "Set 1": {"Pros": "", "Cons": ""},
                    "Set 2": {"Pros": "", "Cons": ""},
                    "Set 3": {"Pros": "", "Cons": ""},
                    "Set 4": {"Pros": "", "Cons": ""},
                    "Set 5": {"Pros": "", "Cons": ""},
                }

            # ── Subgroup → store count (always from full_df) ─────────────────────────
            subgroup_store_counts = {c: 0 for _, _, c, _ in subgroups_def}
            if hasattr(self, 'full_df') and self.full_df is not None and not self.full_df.empty:
                us = self.full_df.drop_duplicates("M_STORE").copy()
                us["Store_Display_Area_num"] = pd.to_numeric(
                    us["Store_Display_Area_num"], errors='coerce').fillna(0)
                st = us["Store_Type"].str.lower()
                da = us["Store_Display_Area_num"]

                masks = {
                    4:  da < num_da_low,
                    5:  (da >= num_da_low) & (da < num_da_high),
                    6:  da >= num_da_high,
                    9:  st == "mall",
                    10: st == "standalone",
                    13: (st == "standalone") & (da < num_sa_split),
                    14: (st == "standalone") & (da >= num_sa_split),
                    15: st == "mall",
                    18: da < num_da_slicer,
                    19: (st == "standalone") & (da >= num_da_slicer),
                    20: (st == "mall")       & (da >= num_da_slicer),
                    23: (st == "standalone") & (da < num_s5_sa),
                    24: (st == "standalone") & (da >= num_s5_sa),
                    25: (st == "mall")       & (da < num_s5_mall),
                    26: (st == "mall")       & (da >= num_s5_mall),
                }
                for c, mask in masks.items():
                    subgroup_store_counts[c] = int(mask.sum())

            # ── Distribute-evenly score from live counts ─────────────────────────────
            def _dist_evenly(c_list):
                counts = [subgroup_store_counts.get(c, 0) for c in c_list]
                total_c = sum(counts)
                if total_c == 0:
                    return 0.0
                mean = total_c / len(counts)
                sd = (sum((x - mean) ** 2 for x in counts) / len(counts)) ** 0.5
                cv = sd / mean if mean else 0.0
                return round(100.0 * (1.0 - cv), 1)

            set_c_lists = {
                "Set 1": [4, 5, 6],
                "Set 2": [9, 10],
                "Set 3": [13, 14, 15],
                "Set 4": [18, 19, 20],
                "Set 5": [23, 24, 25, 26],
            }
            dist_evenly = {s: _dist_evenly(cl) for s, cl in set_c_lists.items()}

            # ── Grouping scores — only when simulation has been run ───────────────────
            is_dynamic = (
                hasattr(self, 'overall_df')
                and self.overall_df is not None
                and not self.overall_df.empty
            )

            eval_sums  = {s: "" for s in set_c_lists}
            score_sums = {s: "" for s in set_c_lists}

            if is_dynamic:
                print("[Grouping Summary] Calculating dynamic grouping scores from simulation data.")

                def _norm(s):
                    if s is None:
                        return ""
                    return str(s).strip().lower().replace(" ", "").replace(",", "").replace(".", "")

                subgroup_labels = {
                    4:  f"DA < {d_low}",        5: f"{d_low}<=DA<{d_high}",
                    6:  f"DA >= {d_high}",       9: "Mall",
                    10: "Standalone (SA)",       13: f"SA < {s_split}",
                    14: f"SA >= {s_split}",      15: "Mall",
                    18: f"Mall + SA < {d_slicer}",
                    19: f"SA >= {d_slicer}",     20: f"Mall >= {d_slicer}",
                    23: f"SA < {s5_sa}",         24: f"SA >= {s5_sa}",
                    25: f"Mall < {s5_mall}",     26: f"Mall >= {s5_mall}",
                }

                def _match_cat(c_idx, set_df, label):
                    if set_df.empty:
                        return None
                    norm_label = _norm(label)
                    for _, r in set_df.iterrows():
                        if _norm(r["Category"]) == norm_label:
                            return r
                    for _, r in set_df.iterrows():
                        cn = str(r["Category"]).strip().lower()
                        if c_idx == 4  and "<"  in cn and ">=" not in cn and "<=" not in cn: return r
                        if c_idx == 5  and ("<=" in cn or (">=" in cn and "<" in cn)):       return r
                        if c_idx == 6  and ">=" in cn and "<" not in cn:                     return r
                        if c_idx == 9  and "mall" in cn:                                     return r
                        if c_idx == 10 and ("sa" in cn or "standalone" in cn):               return r
                        if c_idx == 13 and "sa" in cn and "<" in cn:                         return r
                        if c_idx == 14 and "sa" in cn and ">=" in cn:                        return r
                        if c_idx == 15 and "mall" in cn:                                     return r
                        if c_idx == 18 and ("+" in cn or ("mall" in cn and "sa" in cn) or ("sa" in cn and "<" in cn and "mall" not in cn)): return r
                        if c_idx == 19 and "sa" in cn and ">=" in cn and "mall" not in cn:   return r
                        if c_idx == 20 and "mall" in cn and ">=" in cn:                      return r
                        if c_idx == 23 and "sa" in cn and "<" in cn:                         return r
                        if c_idx == 24 and "sa" in cn and ">=" in cn:                        return r
                        if c_idx == 25 and "mall" in cn and "<" in cn:                       return r
                        if c_idx == 26 and "mall" in cn and ">=" in cn:                      return r
                    return set_df.iloc[0]

                def clean_d_code(val):
                    if val is None or pd.isna(val):
                        return ""
                    s = str(val).strip()
                    if s.endswith(".0"):
                        prefix = s[:-2]
                        if prefix.isdigit():
                            return prefix
                    return s

                # Sales sums for dept-weighted evaluation
                sales_sums = {}
                if hasattr(self, 'sales_df') and self.sales_df is not None and not self.sales_df.empty:
                    s_df = self.sales_df
                    dept_col = next((c for c in s_df.columns if "DEPARTMENT" in c.upper()), None)
                    amt_col  = "AVG_TOP2_TOTAL_AMT_SALES" if "AVG_TOP2_TOTAL_AMT_SALES" in s_df.columns else next(
                        (c for c in s_df.columns if "SALES" in c.upper() and "AMT" in c.upper()), None)
                    if dept_col and amt_col:
                        sc = s_df.copy()
                        sc[dept_col] = sc[dept_col].apply(clean_d_code)
                        sc[amt_col]  = pd.to_numeric(sc[amt_col].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
                        sales_sums   = sc.groupby(dept_col)[amt_col].sum().to_dict()

                total_sales = sum(v for k, v in sales_sums.items() if k != "Grand Total" and not k.startswith("Total"))
                if total_sales <= 0:
                    total_sales = 1.0

                cols_12 = ["F1", "F2", "F3", "F4", "F5", "F6", "G1", "G2", "G3", "G4", "G5", "G6"]
                
                # Load Department Details and remarks from details_path (identical to Grouping Analysis)
                dept_remarks = {}
                _DEPT_DETAILS_PATH = r"Y:\R&D\Category-based Start Up - Split DB Analysis\Tool Data\Department Details.xlsx"
                _details_candidates = [_DEPT_DETAILS_PATH]
                if self.details_path:
                    for _p in str(self.details_path).split("|"):
                        _p = _p.strip()
                        if _p and not _p.startswith("http") and _p not in _details_candidates:
                            _details_candidates.append(_p)

                import os as _os
                import openpyxl as _openpyxl
                for _fp in _details_candidates:
                    if _os.path.exists(_fp):
                        try:
                            _wb = _openpyxl.load_workbook(_fp, data_only=True)
                            c_upper = self.country.upper() if self.country else "ID"
                            _sheet_name = None
                            _candidates_sheets = [
                                f"DEPARTMENT {c_upper}",
                                f"DEPARTMENT_{c_upper}",
                                f"{c_upper} REMARK",
                                f"{c_upper}_REMARK",
                                f"DEPARTMENT BR SG",
                            ]
                            for _cs in _candidates_sheets:
                                for _s in _wb.sheetnames:
                                    if _s.strip().upper() == _cs.upper():
                                        _sheet_name = _s
                                        break
                                if _sheet_name:
                                    break
                            if not _sheet_name:
                                for _s in _wb.sheetnames:
                                    if c_upper in _s.upper():
                                        _sheet_name = _s
                                        break
                            if _sheet_name:
                                _ws = _wb[_sheet_name]
                                _rows_data = list(_ws.iter_rows(values_only=True))
                                if _rows_data:
                                    _headers = [str(c).strip() if c is not None else "" for c in _rows_data[0]]
                                    _dept_col_idx = next((i for i, h in enumerate(_headers) if "DEPT" in h.upper() or "DEPARTMENT" in h.upper()), None)
                                    _rem_col_idx = next((i for i, h in enumerate(_headers) if "REMARK" in h.upper()), None)
                                    if _dept_col_idx is not None and _rem_col_idx is not None:
                                        for _row in _rows_data[1:]:
                                            if len(_row) > _dept_col_idx and _row[_dept_col_idx] is not None:
                                                _d_code = clean_d_code(_row[_dept_col_idx])
                                                _d_rem = str(_row[_rem_col_idx]).strip() if len(_row) > _rem_col_idx and _row[_rem_col_idx] is not None else ""
                                                dept_remarks[_d_code] = _d_rem
                                                _local = f"L{_d_code}" if not _d_code.upper().startswith("L") else _d_code
                                                if _local not in dept_remarks:
                                                    dept_remarks[_local] = _d_rem
                                break
                        except Exception as _ex:
                            print(f"[Grouping Summary] Error loading remarks: {_ex}")

                # ── Apply exact same filters as Grouping Analysis ──
                raw_unique_depts = [clean_d_code(d) for d in self.overall_df["Department"].astype(str).str.strip().unique() if clean_d_code(d)]
                
                _SHOW_REMARKS = {"dept", "main_category", "sub_category"}
                if dept_remarks:
                    unique_depts = [
                        d for d in raw_unique_depts
                        if dept_remarks.get(d, "").strip().lower().replace(" ", "_") in _SHOW_REMARKS
                    ]
                else:
                    unique_depts = raw_unique_depts

                if sales_sums:
                    unique_depts = [
                        d for d in unique_depts
                        if sales_sums.get(d, 0) > 0
                    ]

                for r_set, c_list in set_c_lists.items():
                    set_eval_sum  = 0.0
                    set_score_sum = 0.0
                    for d_code in unique_depts:
                        dept_df   = self.overall_df[self.overall_df["Department"].astype(str).str.strip().apply(clean_d_code) == d_code]
                        sales_pct = sales_sums.get(d_code, 0.0) / total_sales
                        s_scores  = []
                        for c in c_list:
                            set_df  = dept_df[dept_df["Report Set"] == r_set]
                            matched = _match_cat(c, set_df, subgroup_labels[c]) if not set_df.empty else None
                            if matched is not None:
                                y      = [float(matched[col]) if matched[col] is not None else 0.0 for col in cols_12]
                                b      = calculate_balance_score(y)
                                sym    = calculate_symmetry_score(y)
                                c_sc   = calculate_center_score(y)
                                score  = calculate_final_score(b, sym, c_sc)
                            else:
                                score = 0.0
                            s_scores.append(score)

                        score_sum = sum(s_scores)
                        if r_set == "Set 2":
                            score_sum *= 1.5
                        elif r_set == "Set 5":
                            score_sum *= 0.75
                        set_score_sum += score_sum
                        set_eval_sum  += score_sum * sales_pct

                    eval_sums[r_set]  = set_eval_sum
                    score_sums[r_set] = set_score_sum

            # ── Assemble rows_data ────────────────────────────────────────────────────
            rows_data = []
            for set_name, cand, c, pros_name in subgroups_def:
                cnt          = subgroup_store_counts.get(c, 0)
                total_in_set = sum(subgroup_store_counts.get(cx, 0)
                                   for sn, _, cx, _ in subgroups_def if sn == set_name)
                pct          = cnt / total_in_set if total_in_set > 0 else 0.0
                dist_score   = dist_evenly[set_name]
                cons_sales   = eval_sums.get(set_name, "")
                w_sales      = score_sums.get(set_name, "")
                row_pros     = pros_cons.get(set_name, {}).get("Pros", "")
                row_cons     = pros_cons.get(set_name, {}).get("Cons", "")
                rows_data.append([set_name, cand, float(cnt), pct, dist_score, cons_sales, w_sales, row_pros, row_cons])

            # ── Subgroup → sales contribution (always from sales_df + full_df) ────────
            subgroup_sales_amt = {}
            subgroup_sales_pct = {}
            grand_total_sales  = 0.0
            if (hasattr(self, 'sales_df') and self.sales_df is not None and not self.sales_df.empty
                    and hasattr(self, 'full_df') and self.full_df is not None and not self.full_df.empty):
                try:
                    s_df = self.sales_df.copy()
                    store_col_s = next((c for c in s_df.columns if "STORE" in c.upper() and "DEPT" not in c.upper()), None)
                    amt_col_s   = "AVG_TOP2_TOTAL_AMT_SALES" if "AVG_TOP2_TOTAL_AMT_SALES" in s_df.columns else next(
                        (c for c in s_df.columns if "SALES" in c.upper() and "AMT" in c.upper()), None)

                    us_sg = self.full_df.drop_duplicates("M_STORE").copy()
                    us_sg["Store_Display_Area_num"] = pd.to_numeric(
                        us_sg["Store_Display_Area_num"], errors='coerce').fillna(0)

                    if store_col_s and amt_col_s:
                        s_df[amt_col_s] = pd.to_numeric(
                            s_df[amt_col_s].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
                        store_sales_map   = s_df.groupby(store_col_s)[amt_col_s].sum().to_dict()
                        grand_total_sales = sum(store_sales_map.values()) or 1.0

                        st_sg = us_sg["Store_Type"].str.lower()
                        da_sg = us_sg["Store_Display_Area_num"]
                        sg_masks = {
                            4:  da_sg < num_da_low,
                            5:  (da_sg >= num_da_low) & (da_sg < num_da_high),
                            6:  da_sg >= num_da_high,
                            9:  st_sg == "mall",
                            10: st_sg == "standalone",
                            13: (st_sg == "standalone") & (da_sg < num_sa_split),
                            14: (st_sg == "standalone") & (da_sg >= num_sa_split),
                            15: st_sg == "mall",
                            18: da_sg < num_da_slicer,
                            19: (st_sg == "standalone")  & (da_sg >= num_da_slicer),
                            20: (st_sg == "mall")        & (da_sg >= num_da_slicer),
                            23: (st_sg == "standalone")  & (da_sg < num_s5_sa),
                            24: (st_sg == "standalone")  & (da_sg >= num_s5_sa),
                            25: (st_sg == "mall")        & (da_sg < num_s5_mall),
                            26: (st_sg == "mall")        & (da_sg >= num_s5_mall),
                        }
                        for _c, sg_mask in sg_masks.items():
                            sg_stores = set(us_sg.loc[sg_mask, "M_STORE"].astype(str).tolist())
                            sg_total  = sum(v for k, v in store_sales_map.items() if str(k) in sg_stores)
                            subgroup_sales_amt[_c] = sg_total
                            subgroup_sales_pct[_c] = sg_total / grand_total_sales
                except Exception as ex:
                    print(f"[Grouping Summary] Error computing subgroup sales: {ex}")

            # ── Build UI tables ──────────────────────────────────────────────────────
            SUMMARY_N_COLS = 11
            table        = self.table_summary_analysis
            header_table = self.table_summary_header

            table.clearContents()
            table.clearSpans()
            table.setRowCount(len(rows_data))
            table.setColumnCount(SUMMARY_N_COLS)

            header_table.clearContents()
            header_table.clearSpans()
            header_table.setRowCount(2)
            header_table.setColumnCount(SUMMARY_N_COLS)

            header_table.setRowHeight(0, 32)
            header_table.setRowHeight(1, 60) # Increased height for word wrap
            header_table.setWordWrap(True)
            for r in range(len(rows_data)):
                table.setRowHeight(r, 34)

            row0 = ['Splitting', '', 'Store Count', '', 'Sales Contribution', '',
                    'Grouping Distribution Score', '', '', 'Analysis', '']
            row1 = ['', '', 'Count', '%', 'Amount', '%',
                    'Distribute Evenly?',
                    'Consider Department Sales Contribution',
                    'Without Consider Department Sales Contribution',
                    "Pro's", "Con's"]

            for c_idx in range(SUMMARY_N_COLS):
                h0 = QTableWidgetItem(row0[c_idx])
                h0.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                h0.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                h0.setForeground(QBrush(QColor("#1e293b")))
                h0.setBackground(QBrush(QColor("#f1f5f9")))
                header_table.setItem(0, c_idx, h0)

                h1 = QTableWidgetItem(row1[c_idx])
                h1.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                h1.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                h1.setForeground(QBrush(QColor("#1e293b")))
                h1.setBackground(QBrush(QColor("#f8fafc")))
                header_table.setItem(1, c_idx, h1)

            header_table.setSpan(0, 0, 2, 2)   # Set + Candidate
            header_table.setSpan(0, 2, 1, 2)   # Store Count
            header_table.setSpan(0, 4, 1, 2)   # Sales Contribution
            header_table.setSpan(0, 6, 1, 3)   # Grouping Distribution Score
            header_table.setSpan(0, 9, 1, 2)   # Analysis

            currency_map = {"ID": "IDR", "TH": "THB", "MY": "MYR",
                            "BR": "BND", "SG": "SGD", "IN": "INR"}
            curr_prefix = f"{currency_map.get(getattr(self, 'country', ''), '')} "

            table.setAlternatingRowColors(False)

            def _to_f(v):
                try:
                    return float(v)
                except:
                    return None

            # Find top 3 scores for columns 6, 7, 8
            top_scores = {6: [], 7: [], 8: []}
            for c_idx in [6, 7, 8]:
                data_col = c_idx - 2
                scores = []
                for _r in rows_data:
                    _v = _r[data_col] if data_col < len(_r) else ""
                    _f = _to_f(_v)
                    if _f is not None and _f != "":
                        scores.append(_f)
                top_scores[c_idx] = sorted(list(set(scores)), reverse=True)[:3]

            for r_idx, row in enumerate(rows_data):
                _set_label  = str(row[0]).strip()
                _cand_label = str(row[1]).strip()
                _sg_col_idx = next(
                    (_ci for _sn, _cnd, _ci, _ in subgroups_def
                     if _sn == _set_label and _cnd == _cand_label), None)

                for c_idx in range(SUMMARY_N_COLS):
                    if c_idx <= 3:
                        data_col = c_idx
                    elif c_idx in [4, 5]:
                        data_col = None
                    else:
                        data_col = c_idx - 2   # rows_data cols 4-8 → UI cols 6-10

                    if c_idx == 4:
                        _amt = subgroup_sales_amt.get(_sg_col_idx, 0.0) if _sg_col_idx is not None else 0.0
                        display = f"{curr_prefix}{round(_amt):,}" if _amt else ""
                        val     = _amt
                    elif c_idx == 5:
                        _pct = subgroup_sales_pct.get(_sg_col_idx, 0.0) if _sg_col_idx is not None else 0.0
                        display = f"{_pct * 100:.2f}%" if _pct else ""
                        val     = _pct
                    else:
                        val = row[data_col] if data_col is not None and data_col < len(row) else ""
                        if c_idx == 2:
                            display = f"{int(val):,}" if isinstance(val, (int, float)) and val != "" else str(val)
                        elif c_idx == 3:
                            display = f"{val * 100:.0f}%" if isinstance(val, (int, float)) and val != "" else str(val)
                        elif c_idx == 6:
                            display = f"{val:.1f}" if isinstance(val, (int, float)) and val != "" else str(val)
                        elif c_idx in [7, 8]:
                            display = f"{val:.2f}" if isinstance(val, (int, float)) and val != "" else str(val)
                        else:
                            display = str(val).strip() if val is not None else ""

                    item = QTableWidgetItem(display)

                    if c_idx in [1, 9, 10]:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    else:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                    if c_idx in [6, 7, 8] and display != "":
                        item.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                        val_f = _to_f(val)
                        if val_f is not None:
                            ts = top_scores[c_idx]
                            if len(ts) > 0 and abs(val_f - ts[0]) < 1e-5:
                                item.setBackground(QBrush(QColor("#778873"))) # Sage Green (Rank 1)
                                item.setForeground(QBrush(QColor("#ffffff"))) # White text
                            elif len(ts) > 1 and abs(val_f - ts[1]) < 1e-5:
                                item.setBackground(QBrush(QColor("#A1BC98"))) # Muted Sage Green (Rank 2)
                                item.setForeground(QBrush(QColor("#000000"))) # Black text
                            elif len(ts) > 2 and abs(val_f - ts[2]) < 1e-5:
                                item.setBackground(QBrush(QColor("#D2DCB6"))) # Light Sage Green (Rank 3)
                                item.setForeground(QBrush(QColor("#000000"))) # Black text
                            else:
                                item.setBackground(QBrush(QColor("#ffffff")))
                                item.setForeground(QBrush(QColor("#000000")))
                        else:
                            item.setBackground(QBrush(QColor("#ffffff")))
                            item.setForeground(QBrush(QColor("#1e293b")))
                    elif c_idx == 0 and display != "":
                        item.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                        item.setBackground(QBrush(QColor("#ffffff")))
                        item.setForeground(QBrush(QColor("#1e293b")))
                    else:
                        item.setBackground(QBrush(QColor("#ffffff")))
                        item.setForeground(QBrush(QColor("#1e293b")))

                    table.setItem(r_idx, c_idx, item)

            # Merge set-level cells
            spans = [(0, 3), (3, 2), (5, 3), (8, 3), (11, 4)]
            for start_row, row_span in spans:
                for col in [0, 6, 7, 8, 9, 10]:
                    table.setSpan(start_row, col, row_span, 1)

            col_widths = {
                0: 90, 1: 170, 2: 70, 3: 60,
                4: 130, 5: 70, 6: 120,
                7: 160, 8: 160, 9: 260, 10: 260,
            }
            for i in range(SUMMARY_N_COLS):
                w = col_widths.get(i, 100)
                table.setColumnWidth(i, w)
                header_table.setColumnWidth(i, w)

            table.horizontalHeader().setStretchLastSection(True)
            header_table.horizontalHeader().setStretchLastSection(True)

            print(f"[Grouping Summary] Table populated (dynamic={is_dynamic}).")
        except Exception as e:
            print(f"[Grouping Summary] Error populating table: {e}")
            import traceback
            traceback.print_exc()

    def populate_group_tab(self):
        """No-op fallback for the deprecated Group tab."""
        pass

    def sync_grouping_horizontal_analysis(self, value):
        if getattr(self, '_syncing_grouping_h', False): return
        self._syncing_grouping_h = True
        if self.table_grouping_header.horizontalScrollBar().value() != value:
            self.table_grouping_header.horizontalScrollBar().setValue(value)
        self._syncing_grouping_h = False

    def sync_grouping_horizontal_header(self, value):
        if getattr(self, '_syncing_grouping_h', False): return
        self._syncing_grouping_h = True
        if self.table_grouping_analysis.horizontalScrollBar().value() != value:
            self.table_grouping_analysis.horizontalScrollBar().setValue(value)
        self._syncing_grouping_h = False

    def sync_perf_vertical_scroll(self, value):
        if getattr(self, '_syncing_perf_v', False): return
        self._syncing_perf_v = True
        for panel in getattr(self, '_perf_panels', []):
            table = panel[-1]
            if table.verticalScrollBar().value() != value:
                table.verticalScrollBar().setValue(value)
        self._syncing_perf_v = False

    def sync_perf_horizontal_scroll(self, value):
        if getattr(self, '_syncing_perf_h', False): return
        self._syncing_perf_h = True
        for panel in getattr(self, '_perf_panels', []):
            h_table = panel[2]
            table = panel[-1]
            if table.horizontalScrollBar().value() != value:
                table.horizontalScrollBar().setValue(value)
            if h_table.horizontalScrollBar().value() != value:
                h_table.horizontalScrollBar().setValue(value)
        self._syncing_perf_h = False


    def generate_performance_pivot(self, df):
        try:
            self.last_sales_df = df # Keep for filter updates
            if df.empty: return
            
            # Apply Filters
            f_df = df.copy()
            f_type = self.perf_filter_type.currentText()
            checked_details = self.perf_filter_details.checked_items()
            
            if f_type != "ALL": f_df = f_df[f_df["Dept Type"] == f_type]
            if checked_details:
                f_df = f_df[f_df["Dept Details"].isin(checked_details)]

            # Force numeric types to avoid pivot aggregation errors
            f_df['AVG_TOP2_TOTAL_AMT_SALES'] = pd.to_numeric(f_df['AVG_TOP2_TOTAL_AMT_SALES'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
            f_df['Turnover_Val'] = pd.to_numeric(f_df['Turnover_Val'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)

            # Get selected Reporting Sets (Global settings)
            try:
                checked = self.combo_set.checked_items() if hasattr(self, 'combo_set') and hasattr(self.combo_set, 'checked_items') else []
            except RuntimeError:
                checked = []
            target_set1 = checked[0] if checked else "Set1_DA"
            checked_compare = checked[1:] if len(checked) > 1 else []
            
            sets_to_process = [target_set1]
            for cs in checked_compare:
                if cs != target_set1 and cs not in sets_to_process:
                    sets_to_process.append(cs)
            # Allow up to 4 sets (primary + 3 compares)

            # Clear all existing dynamic panels
            for panel in self._perf_panels:
                if len(panel) == 4:
                    container, lbl, header_table, table = panel
                else:
                    container, lbl, table = panel
                container.setParent(None)
            self._perf_panels.clear()

            self.column_delegates = []

            set_names_map = {
                "Set1_DA": "Set 1: Display Area",
                "Set2_StoreType": "Set 2: Store Type",
                "Set3_Type_DA": "Set 3: SA Split",
                "Set4_CurrentSetting": "Set 4: DA Split",
                "Set5_Mall_SA_Split": "Set 5: Mall/SA Split"
            }

            for idx, t_set in enumerate(sets_to_process):
                if t_set not in f_df.columns: continue

                # Pivot for this specific set
                p = pd.pivot_table(
                    f_df,
                    values=['AVG_TOP2_TOTAL_AMT_SALES', 'Turnover_Val'],
                    index=['DEPARTMENT'],
                    columns=[t_set],
                    aggfunc='mean'
                ).fillna(0)

                # Create a fresh panel for this set
                container = QFrame()
                vbox = QVBoxLayout(container)
                vbox.setContentsMargins(0, 0, 0, 0); vbox.setSpacing(0)

                lbl = QLabel("")
                lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
                lbl.setStyleSheet("background-color: #1e3a5f; color: #ffffff; font-weight: bold; font-size: 13px; padding: 6px; border-radius: 4px;")

                # Create upper spanned header table
                header_table = QTableWidget(1, 1)
                header_table.verticalHeader().setVisible(False)
                header_table.horizontalHeader().setVisible(False)
                header_table.setFixedHeight(32)
                header_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
                header_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
                header_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
                header_table.setStyleSheet("""
                    QTableWidget { 
                        border: none; 
                        background-color: #f8fafc; 
                        gridline-color: #cbd5e1; 
                    }
                    QScrollBar:vertical {
                        width: 4px;
                        background: transparent;
                        border: none;
                    }
                    QScrollBar::handle:vertical {
                        background: transparent;
                        border: none;
                    }
                    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                        border: none;
                        background: none;
                        height: 0px;
                    }
                    QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                        background: none;
                    }
                    QScrollBar:horizontal {
                        height: 0px;
                        background: transparent;
                        border: none;
                    }
                    QScrollBar::handle:horizontal {
                        background: transparent;
                        border: none;
                    }
                    QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                        border: none;
                        background: none;
                        width: 0px;
                    }
                    QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
                        background: none;
                    }
                """)

                table = QTableWidget()
                table.setAlternatingRowColors(True)
                table.verticalHeader().setVisible(False)
                table.verticalHeader().setDefaultSectionSize(30)
                table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
                table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
                table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
                table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
                table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

                vbox.addWidget(lbl)
                vbox.addWidget(header_table)
                vbox.addWidget(table)
                self.table_perf_splitter.addWidget(container)
                self._perf_panels.append((container, lbl, header_table, table))

                # Sync scrollbars across ALL panels
                table.verticalScrollBar().valueChanged.connect(self.sync_perf_vertical_scroll)
                table.horizontalScrollBar().valueChanged.connect(self.sync_perf_horizontal_scroll)
                header_table.horizontalScrollBar().valueChanged.connect(self.sync_perf_horizontal_scroll)

                table.setSortingEnabled(False)
                
                # Order columns for this set
                available_cats = f_df[t_set].unique()
                da_order = sorted([s for s in available_cats if s not in ["Store Closed", "Not Found", "Searching..."]])

                final_cols = []
                headers = ["DEPARTMENT"]

                for da in da_order:
                    final_cols.append(('AVG_TOP2_TOTAL_AMT_SALES', da))
                    headers.append(f"({t_set})\n{da} Sales")
                    final_cols.append(('Turnover_Val', da))
                    headers.append(f"({t_set})\n{da} Turnover")

                banner_text = set_names_map.get(t_set, t_set)
                lbl.setText(banner_text)

                table.setRowCount(len(p.index) + 1)
                header_table.setColumnCount(len(headers))
                
                # Empty for department
                h_item_dept = QTableWidgetItem("")
                h_item_dept.setBackground(QColor("#f8fafc"))
                header_table.setItem(0, 0, h_item_dept)
                
                col_offset = 1
                for da in da_order:
                    h_item = QTableWidgetItem(str(da))
                    h_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    h_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
                    h_item.setForeground(QColor("#1e293b"))
                    h_item.setBackground(QColor("#e2e8f0"))
                    header_table.setItem(0, col_offset, h_item)
                    header_table.setSpan(0, col_offset, 1, 2)
                    
                    # Set placeholder for spanned col
                    h_item_sp = QTableWidgetItem("")
                    h_item_sp.setBackground(QColor("#e2e8f0"))
                    header_table.setItem(0, col_offset + 1, h_item_sp)
                    
                    col_offset += 2
                    
                table.setColumnCount(len(headers))
                clean_headers = ["Department"] + ["Sales" if "Sales" in h else "Turnover" for h in headers[1:]]
                table.setHorizontalHeaderLabels(clean_headers)
                
                # Hook up sectionResized for subsequent layout updates safely
                table.horizontalHeader().sectionResized.connect(
                    lambda idx, old, new, ht=header_table: ht.setColumnWidth(idx, new)
                )
                
                for c_idx in range(len(final_cols)):
                    col_key = final_cols[c_idx]
                    if col_key not in p.columns: continue
                    col_data = p[col_key].astype(float)
                    if isinstance(col_data, pd.DataFrame): col_data = col_data.iloc[:, 0]

                    c_max = col_data.max()
                    if c_max <= 0: c_max = 1
                    
                    is_to = "Turnover" in headers[c_idx+1]
                    color = "#10b981" if is_to else "#f59e0b"
                    
                    delegate = DataBarDelegate(color, c_max)
                    self.column_delegates.append(delegate)
                    table.setItemDelegateForColumn(c_idx + 1, delegate)

                # Pre-calculate column sums for Sales columns to compute percentage contributions
                col_sums = {}
                for c_idx, col_key in enumerate(final_cols):
                    if col_key not in p.columns: continue
                    is_to = "Turnover" in headers[c_idx+1]
                    if not is_to:
                        col_sums[col_key] = float(p[col_key].sum())

                for r_idx, dept in enumerate(p.index):
                    table.setItem(r_idx, 0, NumericTableItem(str(dept)))
                    for c_idx, col_key in enumerate(final_cols):
                        if col_key not in p.columns: continue
                        row_data = p.loc[dept, col_key]
                        if isinstance(row_data, pd.Series): row_data = row_data.iloc[0]
                        
                        val = float(row_data)
                        is_to = "Turnover" in headers[c_idx+1]
                        
                        if is_to:
                            item = NumericTableItem(f"{val:.1f}%")
                        else:
                            col_sum = col_sums.get(col_key, 0.0)
                            pct = (val / col_sum * 100) if col_sum > 0 else 0
                            item = NumericTableItem(f"{val:,.0f} ({pct:.1f}%)")
                        
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        table.setItem(r_idx, c_idx + 1, item)

                # --- Populate Grand Total Row ---
                gt_row = len(p.index)
                
                # Department column item
                gt_item = NumericTableItem("Grand Total")
                font_gt = gt_item.font(); font_gt.setBold(True); gt_item.setFont(font_gt)
                gt_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(gt_row, 0, gt_item)
                
                for c_idx, col_key in enumerate(final_cols):
                    if col_key not in p.columns: continue
                    is_to = "Turnover" in headers[c_idx+1]
                    
                    if is_to:
                        item = NumericTableItem("-")
                    else:
                        val = col_sums.get(col_key, 0.0)
                        item = NumericTableItem(f"{val:,.0f} (100.0%)")
                        
                    font = item.font(); font.setBold(True); item.setFont(font)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(gt_row, c_idx + 1, item)
                    
                # Set background color to light gray #e2e8f0 for all cells in Grand Total row
                for col in range(table.columnCount()):
                    item = table.item(gt_row, col)
                    if item:
                        item.setBackground(QColor("#e2e8f0"))

                table.resizeColumnsToContents()
                for c in range(table.columnCount()):
                    header_table.setColumnWidth(c, table.columnWidth(c))
                table.setSortingEnabled(True)
            
        except Exception as e:
            print(f"Pivot Error: {e}")

    def on_sales_header_clicked(self, index):
        header_text = self.table_sales.horizontalHeaderItem(index).text()
        
        try:
            from PyQt6.QtWidgets import QInputDialog
            
            if any(x in header_text for x in ["Dept Type", "Dept Details", "Store Type", "Set1_DA", "Set2_StoreType", "Set3_Type_DA"]):
                # Get unique values from the column for dropdown
                unique_vals = set()
                for r in range(self.table_sales.rowCount()):
                    item = self.table_sales.item(r, index)
                    if item and item.text().strip():
                        unique_vals.add(item.text().strip())
                
                label = ""
                for key in ["Dept Type", "Dept Details", "Store Type", "Set1_DA", "Set2_StoreType", "Set3_Type_DA"]:
                    if key in header_text:
                        label = key
                        break
                
                reset_text = f"All {label}s" if not label.endswith("s") else f"All {label}"
                if label == "Dept Type": reset_text = "All Types"
                
                options = [reset_text] + sorted(list(unique_vals))
                text, ok = QInputDialog.getItem(self, f"Filter {label}", f"Select {label}", options, 0, False,
                                               flags=Qt.WindowType.WindowTitleHint | Qt.WindowType.WindowSystemMenuHint | Qt.WindowType.WindowCloseButtonHint)
                if ok:
                    target = "" if text == reset_text else text
                    for r in range(self.table_sales.rowCount()):
                        if not target:
                            self.table_sales.setRowHidden(r, False)
                        else:
                            item = self.table_sales.item(r, index)
                            if item:
                                self.table_sales.setRowHidden(r, item.text() != target)

            elif "Normalized Dept" in header_text or "Department" in header_text or "Dept" in header_text:
                text, ok = QInputDialog.getText(self, "Filter Department", "Enter Department",
                                                flags=Qt.WindowType.WindowTitleHint | Qt.WindowType.WindowSystemMenuHint | Qt.WindowType.WindowCloseButtonHint)
                if ok:
                    target = text.strip().upper()
                    for r in range(self.table_sales.rowCount()):
                        if not target:
                            self.table_sales.setRowHidden(r, False)
                        else:
                            item = self.table_sales.item(r, index)
                            if item:
                                self.table_sales.setRowHidden(r, target not in item.text().upper())
        except Exception as e:
            print(f"Sales Filter Error: {e}")

    def on_cell_clicked(self, table, row, col):
        if not hasattr(self, "full_df") or self.full_df is None: return
        
        header_text = table.horizontalHeaderItem(col).text()
        valid_cols = ["Store Count", "F1", "F2", "F3", "F4", "F5", "F6", "G1", "G2", "G3", "G4", "G5", "G6"]
        if header_text not in valid_cols: return
        
        try:
            # 1. Identify Row Context
            dept = table.item(row, 0).text()
            t_set = table.item(row, 1).text()
            cat = table.item(row, 2).text()
            
            # 2. Filter full_df
            f_df = self.full_df[self.full_df["M_STORE_DEPARTMENT"] == dept].copy()
            
            # Use the config that was active when this table was generated
            import engine
            conf = self.last_run_config
            da_low = int(float(conf.get('da_low', 7500)))
            da_high = int(float(conf.get('da_high', 9500)))
            sa_split = int(float(conf.get('sa_split', 600)))
            da_slicer = int(float(conf.get('da_slicer', 8000)))
            set5_mall = int(float(conf.get('set5_mall', 10000)))
            set5_sa = int(float(conf.get('set5_sa', 8000)))
            
            defs = engine.get_defs(f_df, da_low, da_high, sa_split, da_slicer, set5_mall_split=set5_mall, set5_sa_split=set5_sa)
            inv_set_names = {
                "Set 1": "Set1_DA",
                "Set 2": "Set2_StoreType",
                "Set 3": "Set3_Type_DA",
                "Set 4": "Set4_CurrentSetting",
                "Set 5": "Set5_Mall_SA_Split"
            }
            mapped_set = inv_set_names.get(t_set, t_set)
            
            if mapped_set in defs:
                def_fn = defs[mapped_set]
                stores = []
                for sub_label, mask in def_fn(f_df):
                    # Robust matching: strip commas and decimals
                    clean_sub = sub_label.replace(",", "").replace(".0", "")
                    clean_cat = cat.replace(",", "").replace(".0", "")
                    if clean_sub == clean_cat:
                        cat_df = f_df[mask]
                        if header_text == "Store Count":
                            stores = cat_df["M_STORE"].unique().tolist()
                        else:
                            # Filter specifically for the selected group (e.g., 'F1')
                            stores = cat_df[cat_df["M_STORE_GROUP"] == header_text]["M_STORE"].unique().tolist()
                        break
                
                if stores:
                    self.show_stores_popup(dept, t_set, cat, header_text, sorted(stores))
                else:
                    QMessageBox.information(self, "No Data", "No stores found for this selection.")
            
        except Exception as e:
            print(f"Drill-Down Error: {e}")

    def show_stores_popup(self, dept, t_set, cat, col_name, stores):
        from PyQt6.QtWidgets import QDialog, QTableWidget, QTableWidgetItem, QVBoxLayout, QPushButton, QHBoxLayout, QAbstractItemView
        from PyQt6.QtCore import Qt
        import html
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Stores Detail: {dept} ({cat}) - {col_name}")
        dialog.resize(850, 500)
        
        vbox = QVBoxLayout(dialog)
        
        # Meta info
        dept_esc = html.escape(str(dept))
        t_set_esc = html.escape(str(t_set))
        cat_esc = html.escape(str(cat))
        col_name_esc = html.escape(str(col_name))
        
        meta_label = QLabel(f"<b>Department:</b> {dept_esc} | <b>Set:</b> {t_set_esc} | <b>Category:</b> {cat_esc} | <b>Group:</b> {col_name_esc} | <b>Count:</b> {len(stores)}")
        meta_label.setStyleSheet("font-size: 13px; color: #1e3a5f;")
        vbox.addWidget(meta_label)
        
        # Create Table Widget
        table = QTableWidget()
        table.setAlternatingRowColors(True)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        
        headers = ["Store Code", "Store Name", "Concept", "Area", "Open Date"]
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        
        sl_df = getattr(self, 'full_store_list_df', pd.DataFrame())
        
        code_col = name_col = concept_col = area_col = date_col = None
        if not sl_df.empty:
            code_col = next((c for c in sl_df.columns if any(x in str(c).upper() for x in ["STORE_CODE", "STORE CODE", "STORE", "CODE"])), None)
            name_col = next((c for c in sl_df.columns if any(x in str(c).upper() for x in ["STORE_NAME", "STORE NAME", "NAME"])), None)
            concept_col = next((c for c in sl_df.columns if any(x in str(c).upper() for x in ["STORE_CONCEPT", "CONCEPT"])), None)
            area_col = next((c for c in sl_df.columns if any(x in str(c).upper() for x in ["STORE_AREA", "AREA"])), None)
            date_col = next((c for c in sl_df.columns if any(x in str(c).upper() for x in ["START_BUSINESS_DATE", "BUSINESS_DATE", "OPEN_DATE", "DATE"])), None)
            
        rows_data = []
        for s_code in stores:
            match_row = None
            if not sl_df.empty and code_col:
                match_df = sl_df[sl_df[code_col].astype(str).str.strip() == str(s_code).strip()]
                if not match_df.empty:
                    match_row = match_df.iloc[0]
            
            if match_row is not None:
                name_val = str(match_row.get(name_col, "-")) if name_col else "-"
                concept_val = str(match_row.get(concept_col, "-")) if concept_col else "-"
                area_val = str(match_row.get(area_col, "-")) if area_col else "-"
                date_val = str(match_row.get(date_col, "-")) if date_col else "-"
                
                # Format area value
                try:
                    area_val = f"{float(area_val):,.0f}"
                except:
                    pass
                
                # Format date value
                if " 00:00:00" in date_val:
                    date_val = date_val.replace(" 00:00:00", "")
            else:
                name_val = "Unknown Store"
                concept_val = "-"
                area_val = "-"
                date_val = "-"
                
            rows_data.append((s_code, name_val, concept_val, area_val, date_val))
            
        table.setRowCount(len(rows_data))
        for r_idx, row_item in enumerate(rows_data):
            for c_idx, val in enumerate(row_item):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(r_idx, c_idx, item)
                
        # Ensure sufficient column spacing to prevent header/data clipping
        header_fm = table.horizontalHeader().fontMetrics()
        for i in range(table.columnCount()):
            header_text = table.horizontalHeaderItem(i).text()
            header_w = header_fm.horizontalAdvance(header_text) + 60
            
            # Find the max width of the contents in this column
            max_col_w = header_w
            for r in range(table.rowCount()):
                cell_item = table.item(r, i)
                if cell_item:
                    cell_text = cell_item.text()
                    cell_w = header_fm.horizontalAdvance(cell_text) + 40
                    if cell_w > max_col_w:
                        max_col_w = cell_w
                        
            table.setColumnWidth(i, max(max_col_w, 140))
            
        table.horizontalHeader().setStretchLastSection(True)
        vbox.addWidget(table)
        
        # Action Buttons
        btn_layout = QHBoxLayout()
        btn_copy = QPushButton("Copy Store Codes")
        btn_copy.clicked.connect(lambda: QApplication.clipboard().setText(", ".join(stores)))
        
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dialog.accept)
        
        btn_layout.addWidget(btn_copy)
        btn_layout.addWidget(btn_close)
        vbox.addLayout(btn_layout)
        
        dialog.exec()

    def on_store_list_header_clicked(self, logical_index, table):
        """Opens a simplified filter dialog matching the user's reference style."""
        if not hasattr(self, 'full_store_list_df') or self.full_store_list_df is None:
            return
            
        header_text = table.horizontalHeaderItem(logical_index).text().strip().upper()
        from PyQt6.QtWidgets import QInputDialog
        
        # 1. RANGE FILTERS (Priority)
        if any(x in header_text for x in ["AREA", "DATE", "BUSINESS"]):
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Filter: {header_text}")
            dialog.setFixedWidth(280)
            vbox = QVBoxLayout(dialog)
            vbox.setSpacing(15); vbox.setContentsMargins(20, 20, 20, 20)
            
            filter_widget = None
            if "AREA" in header_text:
                hbox_inputs = QHBoxLayout()
                
                vbox_min = QVBoxLayout()
                lbl_min = QLabel("Min DA:")
                lbl_min.setStyleSheet("font-weight: bold; color: #475569; font-size: 11px;")
                min_spin = QSpinBox(); min_spin.setRange(0, 100000)
                min_spin.setButtonSymbols(QSpinBox.ButtonSymbols.NoButtons)
                vbox_min.addWidget(lbl_min)
                vbox_min.addWidget(min_spin)
                
                vbox_max = QVBoxLayout()
                lbl_max = QLabel("Max DA:")
                lbl_max.setStyleSheet("font-weight: bold; color: #475569; font-size: 11px;")
                max_spin = QSpinBox(); max_spin.setRange(0, 100000)
                max_spin.setButtonSymbols(QSpinBox.ButtonSymbols.NoButtons)
                vbox_max.addWidget(lbl_max)
                vbox_max.addWidget(max_spin)
                
                hbox_inputs.addLayout(vbox_min)
                hbox_inputs.addLayout(vbox_max)
                vbox.addLayout(hbox_inputs)
                
                saved = self.sl_active_filters.get("AREA_RANGE", (0, 100000))
                min_spin.setValue(int(saved[0])); max_spin.setValue(int(saved[1]))
                filter_widget = (min_spin, max_spin)
            else:
                vbox.addWidget(QLabel("Select Date Range:"))
                start_de = QDateEdit(); start_de.setCalendarPopup(True)
                end_de = QDateEdit(); end_de.setCalendarPopup(True)
                saved = self.sl_active_filters.get("DATE_RANGE", (QDate(2000, 1, 1), QDate.currentDate()))
                start_de.setDate(saved[0]); end_de.setDate(saved[1])
                vbox.addWidget(start_de); vbox.addWidget(QLabel("to", alignment=Qt.AlignmentFlag.AlignCenter)); vbox.addWidget(end_de)
                filter_widget = (start_de, end_de)
                
            bbox = QHBoxLayout()
            btn_ok = QPushButton("OK"); btn_ok.clicked.connect(dialog.accept)
            btn_cancel = QPushButton("Cancel"); btn_cancel.clicked.connect(dialog.reject)
            btn_clear = QPushButton("Clear"); btn_clear.clicked.connect(lambda: (self.sl_active_filters.pop("AREA_RANGE" if "AREA" in header_text else "DATE_RANGE", None), dialog.accept()))
            btn_ok.setMinimumWidth(70); btn_cancel.setMinimumWidth(70); btn_clear.setMinimumWidth(70)
            bbox.addWidget(btn_ok); bbox.addWidget(btn_cancel); bbox.addWidget(btn_clear)
            vbox.addLayout(bbox)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                if "AREA" in header_text:
                    self.sl_active_filters["AREA_RANGE"] = (filter_widget[0].value(), filter_widget[1].value())
                else:
                    self.sl_active_filters["DATE_RANGE"] = (filter_widget[0].date(), filter_widget[1].date())
                self.apply_store_list_filters()
            return

        # 2. DYNAMIC DROPDOWN (For categorical columns with few values)
        col_name = next((c for c in self.full_store_list_df.columns if str(c).strip().upper() == header_text), None)
        if col_name:
            unique_vals = sorted(self.full_store_list_df[col_name].astype(str).unique().tolist())
            if len(unique_vals) <= 30:
                items = ["ALL"] + unique_vals
                current = self.sl_active_filters.get(header_text, "ALL")
                item, ok = QInputDialog.getItem(self, f"Filter: {header_text}", f"Select {header_text}:", items, items.index(current) if current in items else 0, False)
                if ok:
                    if item == "ALL": self.sl_active_filters.pop(header_text, None)
                    else: self.sl_active_filters[header_text] = item
                    self.apply_store_list_filters()
                return

        # 3. GENERIC TEXT FILTER (For large data columns like Store Name)
        current = self.sl_active_filters.get(header_text, "")
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Filter: {header_text}")
        vbox = QVBoxLayout(dialog)
        
        vbox.addWidget(QLabel(f"Search in {header_text}:"))
        line_edit = QLineEdit(current)
        vbox.addWidget(line_edit)
        
        col_name = next((c for c in self.full_store_list_df.columns if str(c).strip().upper() == header_text), None)
        if col_name:
            # Generate sorted list of unique string values, ignoring NaNs
            unique_vals = sorted([str(x) for x in self.full_store_list_df[col_name].unique() if pd.notna(x)])
            completer = QCompleter(unique_vals, dialog)
            completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            completer.setFilterMode(Qt.MatchFlag.MatchContains)
            
            # Force light theme styling on the completer dropdown
            popup = completer.popup()
            popup.setStyleSheet("""
                QListView {
                    background-color: #ffffff;
                    color: #0f172a;
                    border: 1px solid #cbd5e1;
                    selection-background-color: #eff6ff;
                    selection-color: #2563eb;
                }
                QListView::item {
                    padding: 6px;
                }
            """)
            
            line_edit.setCompleter(completer)
            
        bbox = QHBoxLayout()
        btn_ok = QPushButton("OK")
        btn_ok.clicked.connect(dialog.accept)
        btn_cancel = QPushButton("Cancel")
        btn_cancel.clicked.connect(dialog.reject)
        
        btn_ok.setMinimumWidth(70)
        btn_cancel.setMinimumWidth(70)
        
        bbox.addWidget(btn_ok)
        bbox.addWidget(btn_cancel)
        vbox.addLayout(bbox)
        
        line_edit.selectAll()
        line_edit.setFocus()
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            text = line_edit.text()
            if not text.strip(): self.sl_active_filters.pop(header_text, None)
            else: self.sl_active_filters[header_text] = text.strip()
            self.apply_store_list_filters()


    def apply_store_list_filters(self):
        """Filters the Store List data based on active column filters."""
        if not hasattr(self, 'full_store_list_df') or self.full_store_list_df is None:
            return
            
        df = self.full_store_list_df.copy()
        
        # Rearrange columns dynamically to match user's preferred layout
        target_patterns = [
            ("STORE_CODE", ["Store_Code", "Store Code", "STORE_CODE", "STORE CODE"]),
            ("STORE_NAME", ["Store_Name", "Store Name", "STORE_NAME", "STORE NAME"]),
            ("STORE_DISPLAY_AREA", ["Store_Display_Area", "Store Display Area", "STORE_DISPLAY_AREA", "STORE DISPLAY AREA", "DISPLAY_AREA", "DISPLAY AREA"]),
            ("STORE_TYPE", ["Store_Type", "Store Type", "STORE_TYPE", "STORE TYPE"]),
            ("SALES_AMOUNT", ["Sales Amount", "Sales_Amount", "SALES AMOUNT", "SALES_AMOUNT"]),
            ("SET_1", ["Set 1", "Set1", "SET_1", "SET1"]),
            ("SET_2", ["Set 2", "Set2", "SET_2", "SET2"]),
            ("SET_3", ["Set 3", "Set3", "SET_3", "SET3"]),
            ("SET_4", ["Set 4", "Set4", "SET_4", "SET4"]),
            ("SET_5", ["Set 5", "Set5", "SET_5", "SET5"]),
            ("COUNTRY_CODE", ["Country_Code", "Country Code", "COUNTRY_CODE", "COUNTRY CODE", "STORE_COUNTRY_CODE", "STORE COUNTRY CODE"])
        ]
        
        ordered_cols = []
        remaining_cols = list(df.columns)
        
        for key, alternatives in target_patterns:
            matched_col = None
            # 1. Try exact matches first
            for alt in alternatives:
                if alt in remaining_cols:
                    matched_col = alt
                    break
            # 2. Try case-insensitive strip matches next
            if not matched_col:
                for alt in alternatives:
                    target_upper = alt.strip().upper().replace("_", " ")
                    for col in remaining_cols:
                        col_upper = str(col).strip().upper().replace("_", " ")
                        if col_upper == target_upper:
                            matched_col = col
                            break
                    if matched_col:
                        break
            
            if matched_col:
                ordered_cols.append(matched_col)
                remaining_cols.remove(matched_col)
                
        final_col_order = ordered_cols + remaining_cols
        df = df[final_col_order]
        
        # Rename all columns professionally to properly format every header
        new_names = []
        for col in df.columns:
            col_str = str(col).strip()
            col_upper = col_str.upper().replace("_", " ")
            
            if col_upper in ["STORE CODE", "STORE_CODE"]:
                new_names.append("Store Code")
            elif col_upper in ["STORE NAME", "STORE_NAME"]:
                new_names.append("Store Name")
            elif col_upper in ["STORE DISPLAY AREA", "STORE_DISPLAY_AREA", "DISPLAY AREA", "DISPLAY_AREA"]:
                new_names.append("Display Area")
            elif col_upper in ["STORE TYPE", "STORE_TYPE"]:
                new_names.append("Store Type")
            elif col_upper in ["SALES AMOUNT", "SALES_AMOUNT"]:
                new_names.append("Sales Amount")
            elif col_upper in ["SET 1", "SET1"]:
                new_names.append("Set 1")
            elif col_upper in ["SET 2", "SET2"]:
                new_names.append("Set 2")
            elif col_upper in ["SET 3", "SET3"]:
                new_names.append("Set 3")
            elif col_upper in ["SET 4", "SET4"]:
                new_names.append("Set 4")
            elif col_upper in ["SET 5", "SET5"]:
                new_names.append("Set 5")
            elif col_upper in ["COUNTRY CODE", "COUNTRY_CODE", "STORE COUNTRY CODE", "STORE_COUNTRY_CODE"]:
                new_names.append("Country Code")
            elif col_upper in ["STORE CONCEPT", "STORE_CONCEPT", "CONCEPT"]:
                new_names.append("Store Concept")
            elif col_upper in ["STORE ISLAND STATUS", "STORE_ISLAND_STATUS", "ISLAND STATUS", "ISLAND_STATUS"]:
                new_names.append("Island Status")
            elif col_upper in ["START BUSINESS DATE", "START_BUSINESS_DATE", "BUSINESS DATE", "BUSINESS_DATE", "OPEN DATE", "OPEN_DATE"]:
                new_names.append("Business Start Date")
            elif col_upper in ["QTA/TA", "QTA_TA", "QTA / TA", "QTA-TA"]:
                new_names.append("QTA/TA")
            else:
                # Fallback: clean underscores and capitalize words beautifully
                clean_name = col_str.replace("_", " ")
                words = []
                for w in clean_name.split():
                    w_upper = w.upper()
                    if w_upper in ["QTA/TA", "QTA", "TA", "ID", "CBM", "F&B", "FMCG", "IR", "SQL"]:
                        words.append(w_upper)
                    elif "/" in w:
                        parts = w.split("/")
                        if all(p.upper() in ["QTA", "TA", "ID", "DA", "SA"] for p in parts):
                            words.append("/".join(p.upper() for p in parts))
                        else:
                            words.append(w.capitalize())
                    else:
                        words.append(w.capitalize())
                clean_name = " ".join(words)
                new_names.append(clean_name)
                
        df.columns = new_names

        # Helper to resolve filters matching renamed columns
        def resolve_filter_key(k):
            k_upper = str(k).strip().upper()
            if k_upper == "SET 1": return "SET 1"
            if k_upper == "SET 2": return "SET 2"
            if k_upper == "SET 3": return "SET 3"
            if k_upper == "SET 4": return "SET 4"
            if k_upper == "SET 5": return "SET 5"
            if k_upper in ["STORE_ISLAND_STATUS", "STORE ISLAND STATUS"]: return "ISLAND STATUS"
            return k_upper

        # 1. Apply all active filters
        special_keys = ["AREA_RANGE", "DATE_RANGE"]
        for key, val in self.sl_active_filters.items():
            if key in special_keys: continue
            resolved_key = resolve_filter_key(key)
            # Find column by name match
            target_col = next((c for c in df.columns if str(c).strip().upper() == resolved_key), None)
            if target_col:
                # Use exact match if the original filter was a dropdown (small number of unique values)
                orig_unique_count = len(df[target_col].unique())
                if orig_unique_count <= 30:
                    df = df[df[target_col].astype(str) == str(val)]
                else:
                    df = df[df[target_col].astype(str).str.contains(str(val), case=False, na=False)]
                
        # 3. Area Processing (Convert all AREA columns to integer)
        # Exclude 'Set' columns because 'Set 1: Display Area' contains strings like '<7,500' which would coerce to 0.
        area_cols = [c for c in df.columns if "AREA" in str(c).upper() and not str(c).upper().startswith("SET")]
        for ac in area_cols:
            df[ac] = pd.to_numeric(df[ac], errors='coerce').fillna(0).astype(int)
            
        # Apply range filter if an area column is found and filter is active
        if area_cols and "AREA_RANGE" in self.sl_active_filters:
            area_min, area_max = self.sl_active_filters["AREA_RANGE"]
            # Filter based on the first area column found
            df = df[(df[area_cols[0]] >= area_min) & (df[area_cols[0]] <= area_max)]
                    
        # 4. Date Processing (Convert all DATE columns to clean YYYY-MM-DD)
        date_cols = [c for c in df.columns if "DATE" in str(c).upper()]
        primary_date_col = next((c for c in date_cols if any(x in str(c).upper() for x in ["START_BUSINESS_DATE", "BUSINESS_DATE", "OPEN_DATE", "DATE"])), None)
        
        for dc in date_cols:
            # Temporarily parse to datetime to clean it
            temp_dates = pd.to_datetime(df[dc], errors='coerce')
            # Create mask for valid dates
            valid_mask = temp_dates.notna()
            # Apply formatting to valid dates only
            df.loc[valid_mask, dc] = temp_dates[valid_mask].dt.strftime('%d/%m/%Y')
            # Ensure "NEW STORE" remains exactly as is (handled by notna/coerce already)

        # Apply date range filter if active
        if primary_date_col and "DATE_RANGE" in self.sl_active_filters:
            # We need to re-parse just for the comparison mask
            comp_dates = pd.to_datetime(df[primary_date_col], errors='coerce')
            start_qdate, end_qdate = self.sl_active_filters["DATE_RANGE"]
            start_dt = pd.to_datetime(start_qdate.toPyDate())
            end_dt = pd.to_datetime(end_qdate.toPyDate())
            
            date_mask = (comp_dates >= start_dt) & (comp_dates <= end_dt)
            new_store_mask = df[primary_date_col].astype(str).str.upper().str.contains("NEW STORE", na=False)
            df = df[date_mask | new_store_mask]

        # 5. Generic Text Filters (Apply all others)
        special_keys = ["CONCEPT", "STORE_TYPE", "AREA_RANGE", "DATE_RANGE"]
        for key, val in self.sl_active_filters.items():
            if key in special_keys: continue
            resolved_key = resolve_filter_key(key)
            # Find column by name match
            target_col = next((c for c in df.columns if str(c).strip().upper() == resolved_key), None)
            if target_col:
                df = df[df[target_col].astype(str).str.contains(str(val), case=False, na=False)]

        # Update Table
        self.last_store_list_df = df.copy()
        self.populate_summary_table(self.table_store_list, df, frozen_cols=2)
        
        # Re-apply compact sizing with the user's preferred widths
        self.table_store_list_frozen.resizeColumnsToContents()
        self.table_store_list.resizeColumnsToContents()
        
        # Use fixed widths as per preference
        self.table_store_list_frozen.setColumnWidth(0, 150)
        self.table_store_list_frozen.setColumnWidth(1, 200)
        
        for table in [self.table_store_list_frozen, self.table_store_list]:
            header_fm = table.horizontalHeader().fontMetrics()
            for i in range(table.columnCount()):
                if table == self.table_store_list_frozen and i < 2: continue
                header_text = table.horizontalHeaderItem(i).text()
                header_w = header_fm.horizontalAdvance(header_text) + 60
                new_w = max(table.columnWidth(i) + 15, header_w, 80)
                table.setColumnWidth(i, new_w)
        
        total_w = 150 + 200 + 10
        self.table_store_list_frozen.setFixedWidth(total_w)

    def clear_all_store_list_filters(self):
        self.sl_active_filters.clear()
        self.sl_filter_banner.setVisible(False)
        self.apply_store_list_filters()

    def populate_summary_table(self, table, df, **kwargs):
        frozen_cols = kwargs.get('frozen_cols', 0)
        frozen_table = getattr(table, 'frozen_side', None)
        
        if frozen_table and frozen_cols > 0:
            # Recursive call to populate both parts, ensuring we don't recurse infinitely
            df_frozen = df.iloc[:, :frozen_cols]
            df_main = df.iloc[:, frozen_cols:]
            self.populate_summary_table(frozen_table, df_frozen, frozen_cols=0)
            self.populate_summary_table(table, df_main, frozen_cols=0)
            return

        table.setSortingEnabled(False)
        
        table.setRowCount(df.shape[0])
        table.setColumnCount(df.shape[1])
        table.setHorizontalHeaderLabels(df.columns)
        
        spark_col = -1
        if "Sparkline" in df.columns:
            spark_col = list(df.columns).index("Sparkline")
            table.setItemDelegateForColumn(spark_col, self.sparkline_delegate)

        # Heatmap calculations
        heatmap = kwargs.get('heatmap', False)
        grid_min = 0.0
        grid_max = 1.0
        grid_range = 1.0
        count_cols = ["F1", "F2", "F3", "F4", "F5", "F6", "G1", "G2", "G3", "G4", "G5", "G6"]
        
        if heatmap:
            active_cols = [c for c in df.columns if c in count_cols]
            if active_cols:
                numeric_df = df[active_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
                grid_min = float(numeric_df.values.min())
                grid_max = float(numeric_df.values.max())
                grid_range = grid_max - grid_min if grid_max > grid_min else 1.0

        for r in range(df.shape[0]):
            for c in range(df.shape[1]):
                val = df.iloc[r, c]
                col_name = df.columns[c]
                
                if c == spark_col:
                    item = QTableWidgetItem("")
                    item.setData(Qt.ItemDataRole.UserRole, val)
                elif pd.isna(val) or val == "-":
                    item = NumericTableItem("-")
                else:
                    try:
                        num_val = float(val)
                        if col_name in ["AVG_TOP2_TOTAL_AMT_SALES", "AVG_TOP2_TOTAL_AMT_BALANCE", "Sales Amt", "Bal Amt", "Sales Amount"]:
                            curr_prefix = ""
                            if hasattr(self, 'country') and self.country:
                                currency_map = {
                                    "ID": "IDR",
                                    "TH": "THB",
                                    "MY": "MYR",
                                    "BR": "BND",
                                    "SG": "SGD",
                                    "IN": "INR"
                                }
                                curr_prefix = f"{currency_map.get(self.country, self.country)} "
                            formatted_val = f"{curr_prefix}{num_val:,.0f}"
                        elif not num_val.is_integer():
                            formatted_val = f"{num_val:.4f}"
                        else:
                            formatted_val = str(int(num_val))
                        item = NumericTableItem(formatted_val)
                    except:
                        item = QTableWidgetItem(str(val))
                
                # Apply Set Slicer Colors for Store List Tab
                if "SET " in col_name.upper():
                    try:
                        val_str = str(val).upper()
                        bg_color = None
                        
                        from PyQt6.QtGui import QColor, QBrush
                        
                        # Premium Designer Pastel Color Scheme (Soothing & Cohesive)
                        soft_blue   = QColor(224, 242, 254)  # Tailwind sky-100 (Fresh Ice Blue)
                        soft_green  = QColor(209, 250, 229)  # Tailwind emerald-100 (Elegant Mint Green)
                        soft_amber  = QColor(254, 243, 199)  # Tailwind amber-100 (Warm Muted Gold)
                        soft_purple = QColor(243, 232, 255)  # Tailwind purple-100 (Soft Lavender)
                        
                        if "SET 1" in col_name.upper():
                            if ">=" in val_str and "<" in val_str:
                                bg_color = soft_green
                            elif "<" in val_str:
                                bg_color = soft_blue
                            elif ">=" in val_str:
                                bg_color = soft_amber
                                
                        elif "SET 2" in col_name.upper():
                            if "MALL + EXPRESS" in val_str:
                                bg_color = soft_amber
                            elif "SA + EXPRESS" in val_str:
                                bg_color = soft_purple
                            elif "MALL" in val_str:
                                bg_color = soft_blue
                            elif "SA" in val_str:
                                bg_color = soft_green
                                
                        elif "SET 3" in col_name.upper():
                            if "MALL" in val_str:
                                bg_color = soft_blue
                            elif "<" in val_str:
                                bg_color = soft_green
                            else:
                                bg_color = soft_amber
                                
                        elif "SET 4" in col_name.upper():
                            if "MALL" in val_str:
                                if "<" in val_str:
                                    bg_color = soft_blue
                                else:
                                    bg_color = soft_green
                            else:
                                if "<" in val_str:
                                    bg_color = soft_amber
                                else:
                                    bg_color = soft_purple
                                    
                        elif "SET 5" in col_name.upper():
                            if "MALL" in val_str:
                                if "<" in val_str:
                                    bg_color = soft_blue
                                else:
                                    bg_color = soft_green
                            elif "SA" in val_str:
                                if "<" in val_str:
                                    bg_color = soft_amber
                                else:
                                    bg_color = soft_purple
                                    
                        if bg_color:
                            item.setBackground(QBrush(bg_color))
                            item.setForeground(QBrush(QColor(51, 65, 85))) # Tailwind slate-700 (premium contrast text)
                    except Exception as e:
                        print(f"Set coloring error: {e}")
                
                # Apply Heatmap Color
                if heatmap and col_name in count_cols:
                    try:
                        try:
                            num_val = float(val)
                        except:
                            num_val = 0.0
                            
                        ratio = (num_val - grid_min) / grid_range
                        
                        # Premium scaling: White (255, 255, 255) -> Very dark red (127, 20, 20)
                        r_col = int(255 - (255 - 127) * ratio)
                        g_col = int(255 - (255 - 20) * ratio)
                        b_col = int(255 - (255 - 20) * ratio)
                        
                        from PyQt6.QtGui import QColor, QBrush
                        item.setBackground(QBrush(QColor(r_col, g_col, b_col)))
                        
                        # Text color adjustment for high-contrast readability
                        if ratio > 0.45:
                            item.setForeground(QBrush(QColor(255, 255, 255)))
                        else:
                            item.setForeground(QBrush(QColor(30, 41, 59))) # Tailwind slate-800
                            
                        if ratio > 0.7:
                            font = item.font()
                            font.setBold(True)
                            item.setFont(font)
                    except:
                        pass

                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(r, c, item)
        table.setSortingEnabled(True)
        table.resizeColumnsToContents()
        if spark_col != -1:
            table.setColumnWidth(spark_col, 100)

    def sync_table_selection(self, source, target):
        if getattr(self, '_syncing_selection', False): return
        self._syncing_selection = True
        try:
            target.clearSelection()
            selected_rows = set(index.row() for index in source.selectionModel().selectedRows())
            for row in selected_rows:
                target.selectRow(row)
        finally:
            self._syncing_selection = False

    def export_csv(self, choice):
        if not self.possys_path or not self.store_path: return
        
        if choice == "Overall Summary":
            save_path, _ = QFileDialog.getSaveFileName(self, "Save Consolidated Data", "Store_Splitting_Summary.xlsx", "Excel Files (*.xlsx)")
            if not save_path: return
            
            config = {
                'base_folder': os.path.dirname(self.possys_path), 'possys_path': self.possys_path, 'store_path': self.store_path,
                'da_low': self.v_da_low.input.text(), 'da_high': self.v_da_high.input.text(), 'sa_split': self.v_sa_split.input.text(),
                'da_slicer': self.v_da_slicer.input.text(),
                'set5_mall': self.v_set5_mall.input.text(), 'set5_sa': self.v_set5_sa.input.text()
            }
            self.btn_export.setEnabled(False); self.progress.setVisible(True)
            self.progress.setRange(0, 0); self.progress.setFormat("Exporting Overall Summary... %p%")
            self._discard_thread('exporter')
            self.exporter = ExportWorker(config, save_path)
            self.exporter.export_finished.connect(self.on_export_finished); self.exporter.error.connect(self.on_error); self.exporter.start()
            
        elif choice == "Sales Performance":
            if not hasattr(self, 'last_sales_df') or self.last_sales_df.empty:
                QMessageBox.warning(self, "No Data", "Please click 'UPDATE CHART' first to generate Sales Performance data.")
                return
                
            save_path, _ = QFileDialog.getSaveFileName(self, "Save Sales Performance", "Sales_Performance_Full_Extraction.xlsx", "Excel Files (*.xlsx)")
            if not save_path: return
            
            try:
                df = self.last_sales_df.copy()
                # Identify all set columns dynamically for flexibility (Set1, Set2, Set3, etc.)
                set_cols = sorted([c for c in df.columns if str(c).startswith("Set")])
                
                with pd.ExcelWriter(save_path) as writer:
                    # 1. RAW DATA SHEET
                    hidden_cols = ["Turnover_Val", "M_STORE_JOIN"]
                    raw_export = df.drop(columns=[c for c in hidden_cols if c in df.columns])
                    raw_export = raw_export.rename(columns={
                        "AVG_TOP2_TOTAL_AMT_SALES": "Sales Amt",
                        "AVG_TOP2_TOTAL_AMT_BALANCE": "Bal Amt",
                        "STORE": "Store",
                        "DEPARTMENT": "Department",
                        "COMBINE": "Combine"
                    })
                    raw_export.to_excel(writer, sheet_name="Raw Data", index=False)
                    
                    # 2. PIVOT SHEETS FOR EACH SET
                    df['AVG_TOP2_TOTAL_AMT_SALES'] = pd.to_numeric(df['AVG_TOP2_TOTAL_AMT_SALES'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
                    df['Turnover_Val'] = pd.to_numeric(df['Turnover_Val'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
                    
                    for t_set in set_cols:
                        set_df = engine.build_sales_performance_pivot(df, t_set)
                        if set_df is not None:
                            set_df.to_excel(writer, sheet_name=t_set[:31], index=False)
                            
                QMessageBox.information(self, "Export Success", f"Full extraction saved to: {save_path}")
            except Exception as e:
                err_msg = str(e)
                if "permission denied" in err_msg.lower() or "errno 13" in err_msg.lower() or isinstance(e, PermissionError):
                    QMessageBox.critical(self, "Export Error",
                        f"Permission denied: Could not write to:\n{save_path}\n\n"
                        "The file is currently open in Excel (or another program).\n"
                        "Please close the file in Excel and try again.")
                else:
                    QMessageBox.critical(self, "Error", f"Failed to export Sales Performance: {err_msg}")
                
        elif choice == "Performance Analysis":
            if not hasattr(self, 'last_sales_df') or self.last_sales_df.empty:
                QMessageBox.warning(self, "No Data", "Please click 'UPDATE CHART' first to generate Performance Analysis data.")
                return
                
            save_path, _ = QFileDialog.getSaveFileName(self, "Save Performance Analysis Pivot", "Performance_Analysis_Pivot.xlsx", "Excel Files (*.xlsx)")
            if not save_path: return
            
            try:
                df = self.last_sales_df.copy()
                f_type = self.perf_filter_type.currentText()
                checked_details = self.perf_filter_details.checked_items()
                
                if f_type != "ALL": df = df[df["Dept Type"] == f_type]
                if checked_details: df = df[df["Dept Details"].isin(checked_details)]

                df['AVG_TOP2_TOTAL_AMT_SALES'] = pd.to_numeric(df['AVG_TOP2_TOTAL_AMT_SALES'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
                df['Turnover_Val'] = pd.to_numeric(df['Turnover_Val'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)

                try:
                    checked = self.combo_set.checked_items() if hasattr(self, 'combo_set') and hasattr(self.combo_set, 'checked_items') else []
                except RuntimeError:
                    checked = []
                target_set1 = checked[0] if checked else "Set1_DA"
                checked_compare = checked[1:] if len(checked) > 1 else []
                sets_to_process = [target_set1]
                for cs in checked_compare:
                    if cs != target_set1 and cs not in sets_to_process:
                        sets_to_process.append(cs)
                sets_to_process = sets_to_process[:2]

                with pd.ExcelWriter(save_path) as writer:
                    for t_set in sets_to_process:
                        if t_set not in df.columns: continue
                        set_df = engine.build_sales_performance_pivot(df, t_set)
                        if set_df is not None:
                            set_df.to_excel(writer, sheet_name=t_set[:31], index=False)
                            
                QMessageBox.information(self, "Export Success", f"Saved to: {save_path}")
            except Exception as e:
                err_msg = str(e)
                if "permission denied" in err_msg.lower() or "errno 13" in err_msg.lower() or isinstance(e, PermissionError):
                    QMessageBox.critical(self, "Export Error",
                        f"Permission denied: Could not write to:\n{save_path}\n\n"
                        "The file is currently open in Excel (or another program).\n"
                        "Please close the file in Excel and try again.")
                else:
                    QMessageBox.critical(self, "Error", f"Failed to export Performance Analysis: {err_msg}")
                
        elif choice == "Store List":
            if not hasattr(self, 'last_store_list_df') or self.last_store_list_df.empty:
                QMessageBox.warning(self, "No Data", "Store list is empty or has not been loaded.")
                return
                
            save_path, _ = QFileDialog.getSaveFileName(self, "Save Store List", "Filtered_Store_List.xlsx", "Excel Files (*.xlsx);;CSV Files (*.csv)")
            if not save_path: return
            
            try:
                df = self.last_store_list_df.copy()
                if save_path.endswith(".csv"):
                    df.to_csv(save_path, index=False)
                else:
                    df.to_excel(save_path, index=False)
                QMessageBox.information(self, "Export Success", f"Store List saved to: {save_path}")
            except Exception as e:
                err_msg = str(e)
                if "permission denied" in err_msg.lower() or "errno 13" in err_msg.lower() or isinstance(e, PermissionError):
                    QMessageBox.critical(self, "Export Error",
                        f"Permission denied: Could not write to:\n{save_path}\n\n"
                        "The file is currently open in Excel (or another program).\n"
                        "Please close the file in Excel and try again.")
                else:
                    QMessageBox.critical(self, "Error", f"Failed to export Store List: {err_msg}")

        elif choice == "All Summary":
            table = self.table_all_summary
            
            if table.rowCount() == 0:
                QMessageBox.warning(self, "No Data", "All Summary table is empty. Please click 'UPDATE' first to generate data.")
                return

            save_path, _ = QFileDialog.getSaveFileName(self, "Save All Summary", "All_Summary_Sets.xlsx", "Excel Files (*.xlsx)")
            if not save_path: return
            
            try:
                import xlsxwriter
                workbook = xlsxwriter.Workbook(save_path)
                
                # Define formatting styles
                header_format = workbook.add_format({
                    'font_name': 'Segoe UI',
                    'font_size': 11,
                    'bold': False,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border': 1,
                    'bg_color': '#f8fafc'
                })
                
                normal_format = workbook.add_format({
                    'font_name': 'Segoe UI',
                    'font_size': 11,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border': 1
                })
                
                bold_format = workbook.add_format({
                    'font_name': 'Segoe UI',
                    'font_size': 11,
                    'bold': False,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border': 1
                })
                
                total_format = workbook.add_format({
                    'font_name': 'Segoe UI',
                    'font_size': 11,
                    'bold': False,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border': 1,
                    'bg_color': '#e2e8f0'
                })
                
                # Normal numeric/percentage formats
                normal_num_format = workbook.add_format({
                    'font_name': 'Segoe UI',
                    'font_size': 11,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border': 1,
                    'num_format': '#,##0'
                })
                normal_pct_format = workbook.add_format({
                    'font_name': 'Segoe UI',
                    'font_size': 11,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border': 1,
                    'num_format': '0.0%'
                })
                
                # Bold numeric/percentage formats
                bold_num_format = workbook.add_format({
                    'font_name': 'Segoe UI',
                    'font_size': 11,
                    'bold': False,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border': 1,
                    'num_format': '#,##0'
                })
                bold_pct_format = workbook.add_format({
                    'font_name': 'Segoe UI',
                    'font_size': 11,
                    'bold': False,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border': 1,
                    'num_format': '0.0%'
                })
                
                # Total numeric/percentage formats
                total_num_format = workbook.add_format({
                    'font_name': 'Segoe UI',
                    'font_size': 11,
                    'bold': False,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border': 1,
                    'bg_color': '#e2e8f0',
                    'num_format': '#,##0'
                })
                total_pct_format = workbook.add_format({
                    'font_name': 'Segoe UI',
                    'font_size': 11,
                    'bold': False,
                    'align': 'center',
                    'valign': 'vcenter',
                    'border': 1,
                    'bg_color': '#e2e8f0',
                    'num_format': '0.0%'
                })
                
                def parse_number_and_percentage(val):
                    if not val or val == "-":
                        return None, None
                    match = re.match(r"^([\d,]+(?:\.\d+)?)\s*\(([\d\.]+)%\)$", val.strip())
                    if match:
                        try:
                            num_str = match.group(1).replace(",", "")
                            pct_str = match.group(2)
                            num_val = float(num_str) if "." in num_str else int(num_str)
                            pct_val = float(pct_str) / 100.0
                            return num_val, pct_val
                        except ValueError:
                            pass
                    return None, None

                all_worksheet = workbook.add_worksheet("ALL Summary")
                all_worksheet.hide_gridlines(2)
                all_current_row = 0
                all_col_max_lens = {}

                def get_excel_cols(c):
                    if c == 0: return False, 0, None
                    elif c == 1: return False, 1, None
                    elif c == 2: return True, 2, 3
                    elif c == 3: return True, 4, 5
                    elif c == 4: return False, 6, None
                    elif c == 5: return False, 7, None
                    elif c == 6: return False, 8, None
                    elif c == 7: return True, 9, 10
                    elif c == 8: return True, 11, 12
                    return False, c, None
                
                # Write headers
                headers = []
                for col in range(table.columnCount()):
                    header_item = table.horizontalHeaderItem(col)
                    header_text = header_item.text() if header_item else f"Column {col+1}"
                    is_comp, excel_c_num, excel_c_pct = get_excel_cols(col)
                    if is_comp:
                        base_name = header_text.replace(" (%)", "").strip()
                        headers.append(base_name)
                        headers.append(header_text)
                    else:
                        headers.append(header_text)
                        
                for idx, header in enumerate(headers):
                    all_worksheet.write(0, idx, header, header_format)
                    all_col_max_lens[idx] = max(all_col_max_lens.get(idx, 0), len(header))
                
                all_worksheet.set_row(0, 26)
                all_current_row += 1
                
                all_merged_cells = set()
                
                # Write data rows
                for r in range(table.rowCount()):
                    all_worksheet.set_row(all_current_row, 22)
                    
                    is_total_row = False
                    item_zero = table.item(r, 0)
                    if item_zero and "Grand Total" in item_zero.text():
                        is_total_row = True
                        
                    # Check for empty gap rows
                    if not item_zero or not item_zero.text():
                        all_current_row += 1
                        continue
                        
                    # Determine base formats
                    if is_total_row:
                        cell_format = total_format
                        num_cell_format = total_num_format
                        pct_cell_format = total_pct_format
                    else:
                        if item_zero and item_zero.font().bold():
                            cell_format = bold_format
                            num_cell_format = bold_num_format
                            pct_cell_format = bold_pct_format
                        else:
                            cell_format = normal_format
                            num_cell_format = normal_num_format
                            pct_cell_format = normal_pct_format
                            
                    for c in range(table.columnCount()):
                        item = table.item(r, c)
                        val = item.text() if item else ""
                        
                        row_span = table.rowSpan(r, c)
                        col_span = table.columnSpan(r, c)
                        
                        cell_format_c = cell_format
                        num_cell_format_c = num_cell_format
                        pct_cell_format_c = pct_cell_format
                        if not is_total_row and item and item.font().bold() and cell_format != bold_format:
                            cell_format_c = bold_format
                            num_cell_format_c = bold_num_format
                            pct_cell_format_c = bold_pct_format
                            
                        is_comp, excel_c_num, excel_c_pct = get_excel_cols(c)
                        if is_comp:
                            if (r, c) not in all_merged_cells:
                                parsed = False
                                num_val, pct_val = parse_number_and_percentage(val)
                                
                                if num_val is not None:
                                    all_col_max_lens[excel_c_num] = max(all_col_max_lens.get(excel_c_num, 0), len(f"{num_val:,.0f}"))
                                if pct_val is not None:
                                    all_col_max_lens[excel_c_pct] = max(all_col_max_lens.get(excel_c_pct, 0), len(f"{pct_val*100:.1f}%"))
                                    
                                if num_val is not None and pct_val is not None:
                                    if row_span > 1:
                                        all_worksheet.merge_range(all_current_row, excel_c_num, all_current_row + row_span - 1, excel_c_num, num_val, num_cell_format_c)
                                        all_worksheet.merge_range(all_current_row, excel_c_pct, all_current_row + row_span - 1, excel_c_pct, pct_val, pct_cell_format_c)
                                    else:
                                        all_worksheet.write_number(all_current_row, excel_c_num, num_val, num_cell_format_c)
                                        all_worksheet.write_number(all_current_row, excel_c_pct, pct_val, pct_cell_format_c)
                                    parsed = True
                                    
                                if not parsed:
                                    all_col_max_lens[excel_c_num] = max(all_col_max_lens.get(excel_c_num, 0), len(val))
                                    if row_span > 1:
                                        all_worksheet.merge_range(all_current_row, excel_c_num, all_current_row + row_span - 1, excel_c_num, val, cell_format_c)
                                        all_worksheet.merge_range(all_current_row, excel_c_pct, all_current_row + row_span - 1, excel_c_pct, "", cell_format_c)
                                    else:
                                        all_worksheet.write(all_current_row, excel_c_num, val, cell_format_c)
                                        all_worksheet.write(all_current_row, excel_c_pct, "", cell_format_c)
                                        
                                if row_span > 1:
                                    for dr in range(row_span):
                                        all_merged_cells.add((r + dr, c))
                        else:
                            excel_c = excel_c_num
                            if (r, c) not in all_merged_cells:
                                all_col_max_lens[excel_c] = max(all_col_max_lens.get(excel_c, 0), len(val))
                                is_written_all = False
                                if c in [5, 6] and val != "-" and not is_total_row:  # Min DA, Max DA
                                    try:
                                        clean_val = float(val.replace(",", "").strip())
                                        all_worksheet.write_number(all_current_row, excel_c, clean_val, num_cell_format_c)
                                        is_written_all = True
                                    except ValueError:
                                        pass
                                
                                if not is_written_all:
                                    if row_span > 1 or col_span > 1:
                                        for dr in range(row_span):
                                            for dc in range(col_span):
                                                if dr > 0 or dc > 0:
                                                    all_merged_cells.add((r + dr, c + dc))
                                        all_worksheet.merge_range(all_current_row, excel_c, all_current_row + row_span - 1, excel_c + col_span - 1, val, cell_format_c)
                                    else:
                                        all_worksheet.write(all_current_row, excel_c, val, cell_format_c)
                                        
                    all_current_row += 1
                    
                # Auto-fit columns
                for col, max_len in all_col_max_lens.items():
                    all_worksheet.set_column(col, col, max(max_len + 3, 12))
                        
                workbook.close()
                QMessageBox.information(self, "Export Success", f"All Summary exported successfully to:\n{save_path}")
            except Exception as e:
                err_msg = str(e)
                if "permission denied" in err_msg.lower() or "errno 13" in err_msg.lower() or isinstance(e, PermissionError):
                    QMessageBox.critical(self, "Export Error",
                        f"Permission denied: Could not write to:\n{save_path}\n\n"
                        "The file is currently open in Excel (or another program).\n"
                        "Please close the file in Excel and try again.")
                else:
                    QMessageBox.critical(self, "Export Error", f"Failed to export All Summary:\n{err_msg}")

    def get_paths_for_country(self, country_code):
        # If country_code is the currently active country, use current loaded files if available
        if country_code == self.country and self.possys_path and self.store_path:
            return self.possys_path, self.store_path
            
        # Otherwise, try to auto-locate them on the network drive
        import glob
        drives = [r"Z:\\", r"Y:\\", r"X:\\", r"U:\\", r"T:\\"]
        rd_root = None
        for drive in drives:
            try:
                candidate = os.path.join(drive, "R&D", "Category-based Start Up - Split DB Analysis")
                if os.path.exists(candidate):
                    rd_root = candidate
                    break
            except:
                continue
                
        tool_data_dir = None
        if rd_root and os.path.exists(rd_root):
            tool_data_dir = os.path.join(rd_root, "Tool Data")
            

                
        if tool_data_dir and os.path.exists(tool_data_dir):
            if True:
                def find_latest(subfolder, pattern):
                    try:
                        folder_path = os.path.join(tool_data_dir, subfolder)
                        if os.path.exists(folder_path):
                            matches = glob.glob(os.path.join(folder_path, pattern))
                            if matches:
                                valid_matches = []
                                for m in matches:
                                    try:
                                        os.path.getmtime(m)
                                        valid_matches.append(m)
                                    except:
                                        pass
                                if valid_matches:
                                    return max(valid_matches, key=os.path.getmtime)
                    except:
                        pass
                    return None
                    
                p_store = find_latest("Store List", f"{country_code} - Store List*.xlsx")
                if not p_store:
                    p_store = find_latest("Store List", f"*{country_code}*Store List*.xlsx")
                    
                p_grouping = find_latest("Grouping", f"{country_code} - Grouping*.xlsx")
                if not p_grouping:
                    p_grouping = find_latest("Grouping", f"*{country_code}*Grouping*.xlsx")
                    
                if p_store and p_grouping:
                    return p_grouping, p_store
                    
        return None, None

    def export_charts(self):
        if not self.possys_path or not self.store_path:
            QMessageBox.warning(self, "Missing Files", "Please make sure the POSSYS / Grouping File and Store List are loaded.")
            return
            
        country_names = {
            "ID": "Indonesia",
            "TH": "Thailand",
            "BR": "Brunei",
            "MY": "Malaysia",
            "SG": "Singapore",
            "IN": "India"
        }
        countries = ["ID", "BR", "IN", "MY", "SG", "TH"]
        active_country = self.country if hasattr(self, 'country') else "ID"
        
        dialog = CountryExportDialog(countries, country_names, active_country, self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
            
        chosen_codes = dialog.result_selection
        if not chosen_codes:
            return
            
        output_dir = QFileDialog.getExistingDirectory(self, "Select Output Folder for Charts")
        if not output_dir:
            return
            
        # Resolve grouping and store files for each selected country
        export_tasks = []
        for code in chosen_codes:
            full_name = country_names.get(code, code)
            p_grouping, p_store = self.get_paths_for_country(code)
            
            if not p_grouping or not p_store:
                reply = QMessageBox.warning(
                    self,
                    "Files Not Found",
                    f"Could not locate the Grouping/POSSYS and Store List files for {full_name} ({code}) on the network drive.\n\nDo you want to skip {full_name} and continue with others?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.No:
                    return
                continue
                
            export_tasks.append({
                'code': code,
                'name': full_name,
                'possys': p_grouping,
                'store': p_store
            })
            
        if not export_tasks:
            QMessageBox.critical(self, "No Valid Countries", "No valid files could be located for any of the selected countries. Export aborted.")
            return
            
        try:
            checked = self.combo_set.checked_items() if hasattr(self, 'combo_set') and hasattr(self.combo_set, 'checked_items') else []
        except RuntimeError:
            checked = []
        target_set = checked[0] if checked else "Set1_DA"
        target_set_compare = checked[1:] if len(checked) > 1 else []
        
        config = {
            'da_low': self.v_da_low.input.text(), 
            'da_high': self.v_da_high.input.text(), 
            'sa_split': self.v_sa_split.input.text(),
            'da_slicer': self.v_da_slicer.input.text(),
            'set5_mall': self.v_set5_mall.input.text(),
            'set5_sa': self.v_set5_sa.input.text(),
            'target_group': self.combo_dept.currentText(), 
            'target_set': target_set,
            'target_set_compare': target_set_compare
        }
        
        self.btn_export.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        self.progress.setFormat("Exporting Charts... 0% (%p%)")
        
        self._discard_thread('exporter')
        self.exporter = ChartExportWorker(config, output_dir, export_tasks)
        self.exporter.progress.connect(self.on_chart_export_progress)
        self.exporter.export_finished.connect(self.on_chart_export_finished)
        self.exporter.error.connect(self.on_error)
        self.exporter.start()

    def on_chart_export_progress(self, current, total):
        self.progress.setRange(0, total)
        self.progress.setValue(current)
        self.progress.setFormat(f"Exporting Charts... {current}/{total} (%p%)")

    def on_chart_export_finished(self):
        self.btn_export.setEnabled(True)
        self.progress.setVisible(False)
        QMessageBox.information(self, "Export Success", "Charts successfully exported and sorted by Set & Department!")

    def update_readme_with_excluded(self, excluded_df):
        if not hasattr(self, 'readme_lbl') or self.readme_lbl is None: return
        
        # Original static text
        base_html = self.readme_lbl.text()
        if "<!-- EXCLUDED_SECTION_START -->" in base_html:
            base_html = base_html.split("<!-- EXCLUDED_SECTION_START -->")[0]
        elif "<h3 style='color: #ef4444;'>Excluded Stores</h3>" in base_html:
            base_html = base_html.split("<h3 style='color: #ef4444;'>Excluded Stores</h3>")[0]
            # Strip the preceding dynamic wrapper or automation note if present
            if "<div style='margin-top: 20px; padding-top: 15px;" in base_html:
                base_html = base_html.split("<div style='margin-top: 20px; padding-top: 15px;")[0]
            elif "<div style=\"margin-top: 20px; padding-top: 15px;\"" in base_html:
                base_html = base_html.split("<div style=\"margin-top: 20px; padding-top: 15px;\"")[0]
        elif "<h3>Excluded Stores</h3>" in base_html:
            base_html = base_html.split("<h3>Excluded Stores</h3>")[0]

        if excluded_df is None or excluded_df.empty:
            count_text = "<p><b>Excluded Stores:</b> 0</p>"
            list_html = "<p>No stores were excluded.</p>"
        else:
            count_text = f"<p><b>Excluded Stores:</b> {len(excluded_df)}</p>"
            # Table of excluded stores
            rows = []
            # Find columns for Start Date and Name
            date_col = next((c for c in excluded_df.columns if "date" in c.lower()), "Start_Business_Date")
            name_col = next((c for c in excluded_df.columns if "name" in c.lower()), "Store_Name")
            
            for _, row in excluded_df.sort_values("Store_Code").iterrows():
                val = row.get(date_col)
                import pandas as pd
                if pd.isna(val) or val is None or str(val).strip().lower() in ["nan", "nat", "", "none", "n/a"]:
                    clean_date = "NEW STORE"
                else:
                    raw_date = str(val).strip()
                    clean_date = raw_date.split(' ')[0] if ' ' in raw_date else raw_date
                rows.append(f"<tr><td>{row.get('Store_Code', 'N/A')}</td><td>{row.get(name_col, 'N/A')}</td><td>{clean_date}</td><td>{row.get('Store_Concept', 'N/A')}</td><td>{row.get('Store _Dummy_Type', 'N/A')}</td></tr>")
            
            list_html = f"""
            <table border='1' style='border-collapse: collapse; width: 100%; font-size: 11px;'>
                <tr style='background-color: #f1f5f9;'><th>Code</th><th>Name</th><th>Start Date</th><th>Concept</th><th>Dummy Type</th></tr>
                {''.join(rows)}
            </table>
            """

        # Automation Summary Note (Dynamic)
        express_count = 0
        if hasattr(self, 'full_df') and self.full_df is not None:
            # Look for the exact concept string we used in engine.py
            is_special = self.full_df["Store_Concept"].astype(str).str.upper().str.contains("EXPRESS") | \
                         self.full_df["Store _Dummy_Type"].astype(str).str.upper().str.contains("RANKING_BASED")
            express_df = self.full_df[is_special]
            store_col_key = next((c for c in express_df.columns if "store" in c.lower()), "Store_Code")
            express_count = express_df[store_col_key].nunique()

        automation_note = f"""
        <div style='margin-bottom: 15px; padding: 10px; background-color: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 4px;'>
            <p style='color: #166534; font-size: 12px; margin: 0;'>
                <b>✓ Automation Active:</b> {express_count} stores (Express/Ranking Based) were automatically assigned to G6.
            </p>
        </div>
        """
        
        excluded_section = f"""<!-- EXCLUDED_SECTION_START -->
        <div style='margin-top: 20px; padding-top: 15px; border-top: 1px solid #e2e8f0;'>
            {automation_note}
            <h3 style='color: #ef4444;'>Excluded Stores</h3>
            {count_text}
            <div style='max-height: 300px; overflow-y: auto;'>
                {list_html}
            </div>
            <p style='font-size: 11px; color: #64748b; margin-top: 10px;'>
                * Note: These stores are in the Store List but missing from the Simulation file.
            </p>
        </div>
        """
        self.readme_lbl.setText(base_html + excluded_section)

    def on_export_finished(self, path):
        self.btn_export.setEnabled(True)
        self.progress.setRange(0, 100); self.progress.setValue(100)
        self.progress.setFormat("Export Completed!")
        QMessageBox.information(self, "Export Success", f"Saved to: {path}")

    def on_error(self, msg):
        self.btn_run.setEnabled(True); self.btn_export.setEnabled(True); self.progress.setVisible(False)
        if "permission denied" in msg.lower() or "errno 13" in msg.lower():
            QMessageBox.critical(self, "Export Error",
                "Permission denied: Could not write to the file.\n\n"
                "The file is currently open in Excel (or another program).\n"
                "Please close the file in Excel and try again.")
        else:
            QMessageBox.critical(self, "Error", msg)

    def create_new_custom_tab(self):
        active_numbers = set()
        for idx in range(self.tabs.count()):
            title = self.tabs.tabText(idx).strip()
            if title.startswith("Workspace"):
                try:
                    num = int(title.split()[-1])
                    active_numbers.add(num)
                except (ValueError, IndexError):
                    pass
                    
        new_num = 1
        while new_num in active_numbers:
            new_num += 1
            
        tab_widget = QWidget()
        tab_layout = QVBoxLayout(tab_widget) # Use a container layout
        tab_layout.setContentsMargins(0, 0, 0, 0); tab_layout.setSpacing(0)
        
        from PyQt6.QtWidgets import QSplitter
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(10)
        splitter.setStyleSheet("QSplitter::handle { background: #e2e8f0; margin: 4px; border-radius: 2px; } QSplitter::handle:hover { background: #cbd5e1; }")
        
        # Left Panel: Data Viewer
        left_panel = QWidget()
        left_vbox = QVBoxLayout(left_panel)
        left_vbox.setContentsMargins(0, 0, 0, 0); left_vbox.setSpacing(10)
        
        toolbar_h = QHBoxLayout()
        
        lbl_src = QLabel("Data Source:")
        lbl_src.setStyleSheet("font-weight: bold; color: #1e293b; font-size: 12px;")
        combo_source = QComboBox()
        combo_source.addItems(["-- Select File --", "ID store list.xlsx", "Sales_&_Balance_Summary 01 02 03 - Original.xlsx"])
        combo_source.setStyleSheet("padding: 5px; border: 1px solid #cbd5e1; border-radius: 4px; background-color: white;")
        toolbar_h.addWidget(lbl_src)
        toolbar_h.addWidget(combo_source, 1)
        
        btn_add_pivot = QPushButton("+")
        btn_add_pivot.setStyleSheet("padding: 8px 16px; background-color: #10b981; color: white; font-weight: bold; border-radius: 4px; border: none;")
        btn_add_pivot.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        toolbar_h.addWidget(btn_add_pivot)
        
        btn_save = QPushButton("Save")
        btn_save.setStyleSheet("padding: 8px 16px; background-color: #3b82f6; color: white; font-weight: normal; border-radius: 4px; border: none;")
        btn_save.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        toolbar_h.addWidget(btn_save)
        
        btn_load = QPushButton("Load")
        btn_load.setStyleSheet("padding: 8px 16px; background-color: #64748b; color: white; font-weight: normal; border-radius: 4px; border: none;")
        btn_load.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        toolbar_h.addWidget(btn_load)
        
        left_vbox.addLayout(toolbar_h)
        
        from PyQt6.QtWidgets import QScrollArea
        canvas_scroll = QScrollArea()
        canvas_scroll.setWidgetResizable(True)
        canvas_scroll.setStyleSheet("background-color: #f1f5f9; border: 1px solid #e2e8f0; border-radius: 6px;")
        
        canvas_widget = QWidget()
        canvas_layout = QGridLayout(canvas_widget)
        canvas_layout.setSpacing(15)
        canvas_layout.setContentsMargins(15, 15, 15, 15)
        canvas_scroll.setWidget(canvas_widget)
        
        left_vbox.addWidget(canvas_scroll)
        
        tab_widget.cards = []
        tab_widget.active_card = None
        
        # Right Panel: Pivot Options
        right_panel = QFrame()
        right_panel.setMinimumWidth(280)
        right_panel.setStyleSheet("QFrame { background-color: #f8fafc; border: 1px solid #cbd5e1; border-radius: 6px; } QLabel { border: none; background: transparent; }")
        right_vbox = QVBoxLayout(right_panel)
        right_vbox.setContentsMargins(0, 0, 0, 0); right_vbox.setSpacing(0)
        
        v_splitter = QSplitter(Qt.Orientation.Vertical)
        v_splitter.setHandleWidth(10)
        v_splitter.setStyleSheet("QSplitter::handle { background: #e2e8f0; margin: 4px; border-radius: 2px; } QSplitter::handle:hover { background: #cbd5e1; }")
        
        # Top Container: Search and Field List
        top_container = QWidget()
        top_vbox = QVBoxLayout(top_container)
        top_vbox.setContentsMargins(15, 15, 15, 5); top_vbox.setSpacing(10)
        
        lbl_title = QLabel("PivotTable Fields")
        lbl_title.setStyleSheet("font-size: 15px; font-weight: bold; color: #0f172a;")
        top_vbox.addWidget(lbl_title)
        
        lbl_sub = QLabel("Choose fields to add to report:")
        lbl_sub.setStyleSheet("font-size: 11px; color: #64748b;")
        top_vbox.addWidget(lbl_sub)
        
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("Search")
        search_edit.setStyleSheet("padding: 6px; border: 1px solid #cbd5e1; border-radius: 4px; background-color: #ffffff; margin-bottom: 2px;")
        top_vbox.addWidget(search_edit)
        
        from PyQt6.QtWidgets import QListWidget, QListWidgetItem
        field_list = QListWidget()
        field_list.setDragEnabled(True)
        field_list.setDragDropMode(QAbstractItemView.DragDropMode.DragOnly)
        field_list.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        field_list.setStyleSheet("background-color: white; border: 1px solid #cbd5e1; border-radius: 4px;")
        top_vbox.addWidget(field_list, 1)
        
        v_splitter.addWidget(top_container)
        
        # Bottom Container: Drag Label and Grid
        bottom_container = QWidget()
        bottom_vbox = QVBoxLayout(bottom_container)
        bottom_vbox.setContentsMargins(15, 5, 15, 15); bottom_vbox.setSpacing(10)
        
        lbl_drag = QLabel("Drag fields between areas below:")
        lbl_drag.setStyleSheet("font-size: 11px; font-weight: bold; color: #475569; margin-top: 5px;")
        bottom_vbox.addWidget(lbl_drag)
        
        def filter_fields(text):
            for i in range(field_list.count()):
                item = field_list.item(i)
                item.setHidden(text.lower() not in item.text().lower())
                
        search_edit.textChanged.connect(filter_fields)
        
        grid = QGridLayout()
        grid.setSpacing(8)
        
        lbl_f = QLabel("Filters")
        lbl_f.setStyleSheet("font-size: 11px; font-weight: bold; color: #1e293b;")
        combo_f = DropListWidget()
        combo_f.setStyleSheet("background-color: white; border: 1px solid #cbd5e1; border-radius: 4px;")
        combo_f.setMinimumHeight(60)
        combo_f.itemDoubleClicked.connect(lambda item: combo_f.takeItem(combo_f.row(item)))
        
        lbl_c = QLabel("Columns")
        lbl_c.setStyleSheet("font-size: 11px; font-weight: bold; color: #1e293b;")
        combo_c = DropListWidget()
        combo_c.setStyleSheet("background-color: white; border: 1px solid #cbd5e1; border-radius: 4px;")
        combo_c.setMinimumHeight(60)
        combo_c.itemDoubleClicked.connect(lambda item: combo_c.takeItem(combo_c.row(item)))
        
        lbl_r = QLabel("Rows")
        lbl_r.setStyleSheet("font-size: 11px; font-weight: bold; color: #1e293b;")
        combo_r = DropListWidget()
        combo_r.setStyleSheet("background-color: white; border: 1px solid #cbd5e1; border-radius: 4px;")
        combo_r.setMinimumHeight(60)
        combo_r.itemDoubleClicked.connect(lambda item: combo_r.takeItem(combo_r.row(item)))
        
        lbl_v = QLabel("Values")
        lbl_v.setStyleSheet("font-size: 11px; font-weight: bold; color: #1e293b;")
        combo_v = DropListWidget(is_value_area=True)
        combo_v.setStyleSheet("background-color: white; border: 1px solid #cbd5e1; border-radius: 4px;")
        combo_v.setMinimumHeight(60)
        combo_v.itemDoubleClicked.connect(lambda item: combo_v.takeItem(combo_v.row(item)))
                
        combo_f.siblings = {'f': combo_f, 'c': combo_c, 'r': combo_r, 'v': combo_v}
        combo_c.siblings = combo_f.siblings
        combo_r.siblings = combo_f.siblings
        combo_v.siblings = combo_f.siblings
        combo_f.parent_tab = tab_widget
        combo_c.parent_tab = tab_widget
        combo_r.parent_tab = tab_widget
        combo_v.parent_tab = tab_widget
        
        tab_widget.combo_f = combo_f
        tab_widget.combo_c = combo_c
        tab_widget.combo_r = combo_r
        tab_widget.combo_v = combo_v
        
        grid.addWidget(lbl_f, 0, 0); grid.addWidget(combo_f, 1, 0)
        grid.addWidget(lbl_c, 0, 1); grid.addWidget(combo_c, 1, 1)
        grid.addWidget(lbl_r, 2, 0); grid.addWidget(combo_r, 3, 0)
        grid.addWidget(lbl_v, 2, 1); grid.addWidget(combo_v, 3, 1)
        
        def set_active_pivot(card):
            tab_widget.active_card = card
            for c in tab_widget.cards:
                c.setStyleSheet("QFrame { background-color: #ffffff; border: 1px solid #e2e8f0; border-radius: 6px; }")
            card.setStyleSheet("QFrame { background-color: #ffffff; border: 2px solid #2563eb; border-radius: 6px; }")
            
            # Disable signals while loading to avoid circular updates
            combo_f.blockSignals(True); combo_c.blockSignals(True); combo_r.blockSignals(True); combo_v.blockSignals(True)
            
            # Clear and reload fields safely
            combo_f.clear()
            for text in card.filt_c:
                combo_f.add_card_item(text)
                
            combo_c.clear()
            for text in card.col_c:
                combo_c.add_card_item(text)
                
            combo_r.clear()
            for text in card.row_c:
                combo_r.add_card_item(text)
                
            combo_v.clear()
            for text, agg, disp in card.val_items:
                combo_v.add_card_item(text, display_name=disp)
                
            combo_f.blockSignals(False); combo_c.blockSignals(False); combo_r.blockSignals(False); combo_v.blockSignals(False)
                
        tab_widget.set_active_pivot = set_active_pivot
        
        def rearrange_cards():
            # Simply re-index cards in the grid
            # First, remove all from layout
            for card in tab_widget.cards:
                canvas_layout.removeWidget(card)
                
            # Then add back in order
            for i, card in enumerate(tab_widget.cards):
                row = i // 2
                col = i % 2
                canvas_layout.addWidget(card, row, col)
            
            if not tab_widget.active_card and tab_widget.cards:
                set_active_pivot(tab_widget.cards[0])
            elif not tab_widget.cards:
                combo_f.clear(); combo_c.clear(); combo_r.clear(); combo_v.clear()
                
        tab_widget.rearrange_cards = rearrange_cards
        
        def add_pivot_table():
            row = len(tab_widget.cards) // 2
            col = len(tab_widget.cards) % 2
            new_card = PivotTableCard(f"Pivot Table {len(tab_widget.cards)+1}", tab_widget)
            canvas_layout.addWidget(new_card, row, col)
            tab_widget.cards.append(new_card)
            set_active_pivot(new_card)
            
        btn_add_pivot.clicked.connect(add_pivot_table)
        
        def save_workspace():
            from PyQt6.QtWidgets import QFileDialog
            import json
            
            path, _ = QFileDialog.getSaveFileName(self, "Save Workspace", "", "Pivot Workspace (*.pivot)")
            if not path: return
            
            data = {
                "source": combo_source.currentText(),
                "cards": []
            }
            
            for card in tab_widget.cards:
                # Convert aggregations to strings for JSON
                serializable_vals = []
                for d, agg, disp in card.val_items:
                    if callable(agg):
                        # Detect StdDevp/Varp from display name
                        if disp.startswith("StdDevp"): agg_str = "std_p"
                        elif disp.startswith("Varp"): agg_str = "var_p"
                        else: agg_str = "count"
                    else:
                        agg_str = str(agg)
                    serializable_vals.append((d, agg_str, disp))
                    
                card_data = {
                    "title": card.edit_title.text(),
                    "filt_c": card.filt_c,
                    "row_c": card.row_c,
                    "col_c": card.col_c,
                    "val_items": serializable_vals,
                    "filter_values": card.filter_values
                }
                data["cards"].append(card_data)
                
            try:
                with open(path, 'w') as f:
                    json.dump(data, f)
                QMessageBox.information(self, "Success", "Workspace saved successfully!")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save workspace: {str(e)}")

        btn_save.clicked.connect(save_workspace)
        
        def load_workspace():
            from PyQt6.QtWidgets import QFileDialog
            import json
            path, _ = QFileDialog.getOpenFileName(self, "Load Workspace", "", "Pivot Workspace (*.pivot)")
            if not path: return
            
            try:
                with open(path, 'r') as f:
                    data = json.load(f)
                
                # Check if source matches, if not, update it
                if data["source"] != combo_source.currentText():
                    idx = combo_source.findText(data["source"])
                    if idx >= 0:
                        combo_source.setCurrentIndex(idx)
                        load_pivot_source(idx)
                
                # Restore cards (Append)
                for card_data in data["cards"]:
                    new_card = PivotTableCard(card_data["title"], tab_widget)
                    new_card.filt_c = card_data["filt_c"]
                    new_card.row_c = card_data["row_c"]
                    new_card.col_c = card_data["col_c"]
                    
                    # Reconstruct aggregations
                    restored_vals = []
                    for d, agg_str, disp in card_data["val_items"]:
                        if agg_str == "std_p": agg = lambda x: np.std(x, ddof=0)
                        elif agg_str == "var_p": agg = lambda x: np.var(x, ddof=0)
                        else: agg = agg_str
                        restored_vals.append((d, agg, disp))
                    
                    new_card.val_items = restored_vals
                    new_card.filter_values = card_data["filter_values"]
                    
                    row = len(tab_widget.cards) // 2
                    col = len(tab_widget.cards) % 2
                    canvas_layout.addWidget(new_card, row, col)
                    tab_widget.cards.append(new_card)
                
                if tab_widget.cards:
                    set_active_pivot(tab_widget.cards[-1]) # Focus last loaded
                
                QMessageBox.information(self, "Success", "Workspace tables imported!")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load workspace: {str(e)}")

        btn_load.clicked.connect(load_workspace)
        
        
        bottom_vbox.addLayout(grid)
        v_splitter.addWidget(bottom_container)
        
        v_splitter.setStretchFactor(0, 3)
        v_splitter.setStretchFactor(1, 2)
        
        right_vbox.addWidget(v_splitter)
        
        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        
        # Set initial sizes (e.g. 70% left, 30% right)
        splitter.setStretchFactor(0, 7)
        splitter.setStretchFactor(1, 3)
        
        tab_layout.addWidget(splitter)
        
        # State Storage
        tab_widget.source_df = None
        
        # Connect source selector
        def load_pivot_source(idx):
            if idx == 0:
                field_list.clear()
                combo_f.clear(); combo_c.clear(); combo_r.clear(); combo_v.clear()
                tab_widget.source_df = None
                return
            
            file_name = combo_source.currentText()
            path = ""
            if "store" in file_name.lower():
                path = self.store_path
            elif "sales" in file_name.lower():
                path = self.sales_path
                
            if not path or not os.path.exists(path):
                QMessageBox.warning(self, "Warning", "Selected file path not found.")
                return
                
            try:
                df = pd.read_excel(path)
                tab_widget.source_df = df
                field_list.clear()
                for col in df.columns:
                    item = QListWidgetItem(col)
                    item.setData(Qt.ItemDataRole.UserRole, col)
                    field_list.addItem(item)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load file: {str(e)}")
                
        combo_source.currentIndexChanged.connect(load_pivot_source)
        
        # Connect Pivot logic
        def apply_pivot_logic():
            if tab_widget.source_df is None:
                return
                
            if not tab_widget.active_card:
                return # Silent return for auto-updates
                
            df = tab_widget.source_df
            def get_list_items(lw):
                items = []
                for i in range(lw.count()):
                    d = lw.item(i).data(Qt.ItemDataRole.UserRole)
                    if d:
                        items.append(d)
                    else:
                        items.append(lw.item(i).text())
                return [x for x in items if x != "-- None --"]
                
            def get_value_items(lw):
                items = []
                for i in range(lw.count()):
                    item = lw.item(i)
                    card = lw.itemWidget(item)
                    d = item.data(Qt.ItemDataRole.UserRole) or item.text()
                    if d == "-- None --": continue
                    
                    agg = "count"
                    display_name = f"Count of {d}"
                    if card:
                        display_name = card.lbl.text()
                        for f in ["Sum", "Count", "Average", "Max", "Min", "Product", "Count Numbers", "StdDevp", "StdDev", "Varp", "Var"]:
                            if display_name.startswith(f):
                                if f == "Sum": agg = "sum"
                                elif f == "Count": agg = "count"
                                elif f == "Average": agg = "mean"
                                elif f == "Max": agg = "max"
                                elif f == "Min": agg = "min"
                                elif f == "Product": agg = "prod"
                                elif f == "Count Numbers": agg = "count"
                                elif f == "StdDevp": agg = lambda x: np.std(x, ddof=0)
                                elif f == "StdDev": agg = "std"
                                elif f == "Varp": agg = lambda x: np.var(x, ddof=0)
                                elif f == "Var": agg = "var"
                                break
                    items.append((d, agg, display_name))
                return items
                
            row_c = get_list_items(combo_r)
            col_c = get_list_items(combo_c)
            filt_c = get_list_items(combo_f)
            val_items = get_value_items(combo_v)
            
            # Save configurations safely
            tab_widget.active_card.filt_c = filt_c
            tab_widget.active_card.row_c = row_c
            tab_widget.active_card.col_c = col_c
            tab_widget.active_card.val_items = val_items

            try:
                res = engine.compute_workspace_pivot(df, row_c, col_c, val_items, tab_widget.active_card.filter_values)
                if res is None:
                    return
                            
                target_table = tab_widget.active_card.table
                target_table.clear()
                if isinstance(res, pd.DataFrame):
                    # Flatten multi-index columns
                    if isinstance(res.columns, pd.MultiIndex):
                        res.columns = ['_'.join(map(str, col)).strip('_') for col in res.columns.values]
                    # Flatten multi-index index
                    if hasattr(res.index, 'names'):
                        new_names = []
                        for i, name in enumerate(res.index.names):
                            if name is None or str(name) == 'None' or str(name) == '':
                                new_names.append(f"Row_Level_{i}")
                            else:
                                new_names.append(name)
                        res.index.names = new_names
                        
                    res = res.reset_index()
                    
                    target_table.setRowCount(len(res))
                    target_table.setColumnCount(len(res.columns))
                    target_table.setHorizontalHeaderLabels([str(c) for c in res.columns])
                    
                    from PyQt6.QtGui import QColor
                    for r_i in range(len(res)):
                        is_total_row = any(str(res.iloc[r_i, x]) == "Grand Total" for x in range(len(res.columns)))
                        for c_i in range(len(res.columns)):
                            is_total_col = str(res.columns[c_i]) == "Grand Total"
                            v = res.iloc[r_i, c_i]
                            item = QTableWidgetItem(str(v) if pd.notnull(v) else "")
                            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            
                            if is_total_row or is_total_col:
                                item.setBackground(QColor("#bfdbfe"))
                                font = item.font(); font.setBold(True); item.setFont(font)
                            
                            target_table.setItem(r_i, c_i, item)
                else:
                    res = res.reset_index()
                    target_table.setRowCount(len(res))
                    target_table.setColumnCount(len(res.columns))
                    target_table.setHorizontalHeaderLabels([str(c) for c in res.columns])
                    
                    from PyQt6.QtGui import QColor
                    for r_i in range(len(res)):
                        is_total_row = any(str(res.iloc[r_i, x]) == "Grand Total" for x in range(len(res.columns)))
                        for c_i in range(len(res.columns)):
                            is_total_col = str(res.columns[c_i]) == "Grand Total"
                            v = res.iloc[r_i, c_i]
                            item = QTableWidgetItem(str(v) if pd.notnull(v) else "")
                            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            
                            if is_total_row or is_total_col:
                                item.setBackground(QColor("#bfdbfe"))
                                font = item.font(); font.setBold(True); item.setFont(font)
                                
                            target_table.setItem(r_i, c_i, item)
                            
                target_table.resizeColumnsToContents()
                
            except Exception as e:
                QMessageBox.critical(self, "Pivot Error", f"Failed to update pivot: {str(e)}")
                
        tab_widget.apply_pivot_logic = apply_pivot_logic
        
        # Connect signals for auto-updates
        for box in [combo_f, combo_c, combo_r, combo_v]:
            box.model().rowsInserted.connect(apply_pivot_logic)
            box.model().rowsRemoved.connect(apply_pivot_logic)
        
        # Add first table automatically for fresh workspaces
        add_pivot_table()
        
        tab_idx = self.tabs.addTab(tab_widget, f" Workspace {new_num} ")
        self.tabs.setCurrentIndex(tab_idx)

    def close_current_custom_tab(self):
        idx = self.tabs.currentIndex()
        if idx > 8:
            self.tabs.removeTab(idx)

    def export_grouping_analysis(self):
        self.export_grouping_excel(self.table_grouping_header, self.table_grouping_analysis, "Grouping Analysis")

    def export_grouping_summary(self):
        self.export_grouping_excel(self.table_summary_header, self.table_summary_analysis, "Grouping Summary")

    def export_grouping_excel(self, header_table, data_table, default_name):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import xlsxwriter

        path, _ = QFileDialog.getSaveFileName(self, f"Export {default_name}", f"{default_name}.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return

        try:
            workbook = xlsxwriter.Workbook(path)
            worksheet = workbook.add_worksheet(default_name[:31])
            self.write_qtable_to_sheet(workbook, worksheet, data_table, default_name, header_table=header_table)
            workbook.close()
            QMessageBox.information(self, "Export Success", f"Successfully exported to {path}")
        except Exception as e:
            err_msg = str(e)
            if "permission denied" in err_msg.lower() or "errno 13" in err_msg.lower() or isinstance(e, PermissionError):
                QMessageBox.critical(self, "Export Error",
                    f"Permission denied: Could not write to:\n{path}\n\n"
                    "The file is currently open in Excel (or another program).\n"
                    "Please close the file in Excel and try again.")
            else:
                QMessageBox.critical(self, "Export Error", f"Failed to export: {err_msg}")

    def write_qtable_to_sheet(self, workbook, worksheet, table_widget, sheet_name, header_table=None):
        worksheet.hide_gridlines(2)
        
        # Proportional column width setting
        cols = table_widget.columnCount() if table_widget else (header_table.columnCount() if header_table else 0)
        for c in range(cols):
            w = table_widget.columnWidth(c) if table_widget else (header_table.columnWidth(c) if header_table else 0)
            if w > 0:
                worksheet.set_column(c, c, max(w * 0.13, 10))

        # Format cache: reuse identical format objects to stay under Excel's ~64k unique format limit
        _fmt_cache = {}
        def get_cached_fmt(props):
            key = tuple(sorted(props.items()))
            if key not in _fmt_cache:
                _fmt_cache[key] = workbook.add_format(props)
            return _fmt_cache[key]

        skip_cells = set()
        combined_cols = set()   # cols with "value (pct%)" format
        pct_only_cols = set()   # cols with plain "xx.x%" format
        curr_row = 0

        def write_widget_rows(start_row, widget, is_header_style=False):
            if not widget:
                return start_row
                
            w_rows = widget.rowCount()
            w_cols = widget.columnCount()
            
            for r in range(w_rows):
                h = widget.rowHeight(r)
                if h > 0:
                    worksheet.set_row(start_row + r, h * 0.75)
                    
                for c in range(w_cols):
                    if (r, c) in skip_cells:
                        continue
                        
                    item = widget.item(r, c)
                    text = item.text() if item else ""
                    
                    cell_fmt_dict = {
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter'
                    }
                    
                    if is_header_style or (item and item.font().bold()):
                        cell_fmt_dict['bold'] = True
                        
                    # Background Color
                    if is_header_style:
                        cell_fmt_dict['bg_color'] = '#f8fafc'
                        cell_fmt_dict['text_wrap'] = True
                    elif item and item.background().color().isValid():
                        bg_name = item.background().color().name()
                        if bg_name != "#000000":
                            cell_fmt_dict['bg_color'] = bg_name
                            
                    # Text Color
                    if item and item.foreground().color().isValid():
                        fg_name = item.foreground().color().name()
                        if fg_name != "#000000":
                            cell_fmt_dict['font_color'] = fg_name

                    # Currency stripping logic
                    import re
                    clean_text = text.strip()
                    m = re.match(r"^([A-Z]{3})\s*([\d\.,\s\-]+)$", clean_text)
                    if m:
                        text = m.group(2).replace(" ", "")
                        cell_fmt_dict['num_format'] = '#,##0'
                    else:
                        for prefix in ["IDR", "THB", "MYR", "BND", "SGD", "INR", "USD", "EUR"]:
                            if clean_text.startswith(prefix):
                                text = clean_text.replace(prefix, "").strip()
                                cell_fmt_dict['num_format'] = '#,##0'
                                break

                    # Detect combined "amount (percentage%)" format, e.g. "437 (34.1%)" or "125,781,519,954 (28.9%)"
                    comb_m = re.match(r"^([^\(]+)\s*\(([\d\.]+)%\)$", clean_text)

                    # Percentage parsing and formatting
                    parsed_val = text
                    is_numeric = False
                    
                    if comb_m:
                        # Write the PERCENTAGE as a numeric value with a standard format string.
                        # Using '0.0%' (same format for ALL combined cells) avoids the 64k format
                        # explosion, AND numeric values are required for Excel data bars to render.
                        group_2 = comb_m.group(2).strip()
                        try:
                            parsed_val = float(group_2) / 100.0
                            is_numeric = True
                            cell_fmt_dict['num_format'] = '0.0%'
                            if not is_header_style:
                                combined_cols.add(c)
                        except ValueError:
                            parsed_val = clean_text
                            is_numeric = False
                    elif text and text != "-" and text != "":
                        try:
                            clean_num_text = text.replace(",", "")
                            if "%" in text:
                                parsed_val = float(clean_num_text.replace("%", "").strip()) / 100.0
                                cell_fmt_dict['num_format'] = '0.00%'
                                is_numeric = True
                                if not is_header_style:
                                    pct_only_cols.add(c)
                            else:
                                val = float(clean_num_text)
                                parsed_val = val
                                is_numeric = True
                        except ValueError:
                            pass

                    # Use cached format to avoid creating thousands of duplicate format objects
                    cell_fmt = get_cached_fmt(cell_fmt_dict)
                    
                    row_span = widget.rowSpan(r, c)
                    col_span = widget.columnSpan(r, c)
                    
                    if row_span > 1 or col_span > 1:
                        if is_numeric:
                            worksheet.merge_range(
                                start_row + r, c,
                                start_row + r + row_span - 1, c + col_span - 1,
                                parsed_val, cell_fmt
                            )
                        else:
                            worksheet.merge_range(
                                start_row + r, c,
                                start_row + r + row_span - 1, c + col_span - 1,
                                text, cell_fmt
                            )
                        for dr in range(row_span):
                            for dc in range(col_span):
                                skip_cells.add((r + dr, c + dc))
                    else:
                        if is_numeric:
                            worksheet.write_number(start_row + r, c, parsed_val, cell_fmt)
                        else:
                            worksheet.write_string(start_row + r, c, text, cell_fmt)
                            
            return start_row + w_rows

        # Horizontal headers — always write column headers for every table that has columns
        if table_widget and cols > 0:
            worksheet.set_row(curr_row, 26)
            header_fmt = get_cached_fmt({
                'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter',
                'bg_color': '#f1f5f9', 'text_wrap': True
            })
            for c in range(cols):
                header_item = table_widget.horizontalHeaderItem(c)
                header_text = header_item.text() if header_item else f"Column {c+1}"
                worksheet.write(curr_row, c, header_text, header_fmt)
            curr_row += 1

        if header_table:
            curr_row = write_widget_rows(curr_row, header_table, is_header_style=True)
            
        if table_widget:
            skip_cells.clear()  # Always clear skip cells before writing data table
            curr_row = write_widget_rows(curr_row, table_widget, is_header_style=False)

        # Apply conditional formatting for Sales Contribution % data bars
        if sheet_name == "Grouping Analysis" and table_widget:
            data_rows_count = table_widget.rowCount()
            start_data_row = curr_row - data_rows_count
            if curr_row - 1 >= start_data_row:
                worksheet.conditional_format(start_data_row, 4, curr_row - 1, 4, {
                    'type': 'data_bar',
                    'bar_color': '#f59e0b',
                    'bar_solid': True,
                    'min_type': 'min',
                    'max_type': 'max'
                })

        # Apply conditional formatting for data bars on any table using combined percentage cols or pure % cols
        all_bar_cols = combined_cols | pct_only_cols
        if table_widget and all_bar_cols:
            data_rows_count = table_widget.rowCount()
            start_data_row = curr_row - data_rows_count
            if curr_row - 1 >= start_data_row:
                for c in all_bar_cols:
                    # Determine bar colour from column header text
                    col_header = ''
                    if table_widget:
                        h_item = table_widget.horizontalHeaderItem(c)
                        if h_item:
                            col_header = h_item.text().lower()
                    if 'sales' in col_header or 'turnover' in col_header or 'contribution' in col_header:
                        bar_color = '#10b981'  # Emerald green for sales/revenue columns
                    else:
                        bar_color = '#f59e0b'  # Amber/orange for store count / subtotal columns

                    worksheet.conditional_format(start_data_row, c, curr_row - 1, c, {
                        'type': 'data_bar',
                        'bar_color': bar_color,
                        'bar_solid': True,
                        'min_type': 'min',
                        'max_type': 'max'
                    })

    def export_master(self):
        if not self.possys_path or not self.store_path:
            QMessageBox.warning(self, "Missing Files", "Please make sure the POSSYS / Grouping File and Store List are loaded.")
            return

        if not hasattr(self, 'last_sales_df') or self.last_sales_df.empty or not hasattr(self, 'last_store_list_df') or self.last_store_list_df.empty or self.table_all_summary.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "Please click 'UPDATE' first to generate all analysis tables before exporting Master.xlsx.")
            return

        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        from PyQt6.QtGui import QColor
        import xlsxwriter
        import re

        save_path, _ = QFileDialog.getSaveFileName(self, "Save Master Workbook", "Master.xlsx", "Excel Files (*.xlsx)")
        if not save_path:
            return

        try:
            self.btn_export.setEnabled(False)
            self.progress.setVisible(True)
            self.progress.setRange(0, 0)
            self.progress.setFormat("Exporting Master Workbook... %p%")
            
            workbook = xlsxwriter.Workbook(save_path)

            # 1. Dept Summary
            ws_dept = workbook.add_worksheet("Dept Summary")
            self.write_qtable_to_sheet(workbook, ws_dept, self.table_summary, "Dept Summary")

            # 2. Overall Summary
            ws_overall = workbook.add_worksheet("Overall Summary")
            self.write_qtable_to_sheet(workbook, ws_overall, self.table_overall, "Overall Summary")

            # 3. Store List
            ws_store = workbook.add_worksheet("Store List")
            self.write_qtable_to_sheet(workbook, ws_store, self.table_store_list, "Store List")

            # 4. Sales Performance
            ws_sales = workbook.add_worksheet("Sales Performance")
            self.write_qtable_to_sheet(workbook, ws_sales, self.table_sales, "Sales Performance")

            # 5. Performance Analysis
            for panel in getattr(self, '_perf_panels', []):
                try:
                    container, lbl, header_table, data_table = panel
                    set_label = lbl.text().strip()
                    clean_label = re.sub(r'[\[\]\*\?\:\\\/]', '', set_label)
                    sheet_name = f"Perf - {clean_label}"[:31]
                    ws_perf = workbook.add_worksheet(sheet_name)
                    self.write_qtable_to_sheet(workbook, ws_perf, data_table, sheet_name, header_table=header_table)
                except Exception as pe:
                    print(f"Failed to export performance panel: {pe}")

            # 6. All Summary
            ws_all = workbook.add_worksheet("All Summary")
            self.write_qtable_to_sheet(workbook, ws_all, self.table_all_summary, "All Summary")

            # 7. Grouping Analysis
            ws_group_an = workbook.add_worksheet("Grouping Analysis")
            self.write_qtable_to_sheet(workbook, ws_group_an, self.table_grouping_analysis, "Grouping Analysis", header_table=self.table_grouping_header)

            # 8. Grouping Summary
            ws_group_sum = workbook.add_worksheet("Grouping Summary")
            self.write_qtable_to_sheet(workbook, ws_group_sum, self.table_summary_analysis, "Grouping Summary", header_table=self.table_summary_header)

            workbook.close()
            
            self.btn_export.setEnabled(True)
            self.progress.setVisible(False)
            QMessageBox.information(self, "Export Success", f"Successfully exported all tabs with exact formatting and styles to:\n{save_path}")
        except Exception as e:
            self.btn_export.setEnabled(True)
            self.progress.setVisible(False)
            err_msg = str(e)
            if "permission denied" in err_msg.lower() or "errno 13" in err_msg.lower() or isinstance(e, PermissionError):
                QMessageBox.critical(self, "Export Error",
                    f"Permission denied: Could not write to:\n{save_path}\n\n"
                    "The file is currently open in Excel (or another program).\n"
                    "Please close the file in Excel and try again.")
            else:
                QMessageBox.critical(self, "Export Error", f"Failed to export Master workbook:\n{err_msg}")

if __name__ == "__main__":
    import signal
    # Prevent KeyboardInterrupt (Ctrl+C from terminal) from propagating into
    # matplotlib's Qt event filter and crashing the process with a traceback.
    signal.signal(signal.SIGINT, signal.SIG_IGN)
    app = QApplication(sys.argv)
    window = StoreChartApp()
    window.show()
    sys.exit(app.exec())
