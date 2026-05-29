import re
import os
from pathlib import Path
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scipy.stats
from collections import defaultdict
from matplotlib.figure import Figure

# ============================================================
# FILE CACHE (mtime-based — serves from RAM if file unchanged)
# ============================================================
_cache = {}
def get_cached_excel(path, sheet_name=0, **kwargs):
    try: mtime = os.path.getmtime(path)
    except OSError: mtime = 0
    cache_key = (str(path), str(sheet_name))
    if cache_key in _cache:
        cached_mtime, cached_df = _cache[cache_key]
        if cached_mtime == mtime:
            return cached_df.copy()
    
    path_str = str(path)
    if path_str.lower().endswith('.csv'):
        # Remove Excel-specific kwargs
        csv_kwargs = {k: v for k, v in kwargs.items() if k not in ['engine', 'sheet_name']}
        df = pd.read_csv(path, **csv_kwargs)
    else:
        df = pd.read_excel(path, sheet_name=sheet_name, **kwargs)
        
    _cache[cache_key] = (mtime, df.copy())
    return df

def get_grouping_sheet_name(path):
    """
    Looks for a sheet named 'FINAL' (case-insensitive) in the Excel file.
    If found, returns it; otherwise, returns 0 (the first sheet).
    """
    try:
        import pandas as pd
        xl = pd.ExcelFile(path, engine='calamine')
        for s in xl.sheet_names:
            if str(s).strip().upper() == "FINAL":
                return s
        return xl.sheet_names[0]
    except Exception:
        return 0

# ============================================================
# DEFAULTS
# ============================================================
STORE_INFO_SHEET = None  # Discovered dynamically per file
_FG_LABELS = ["F1", "F2", "F3", "F4", "F5", "F6", "G1", "G2", "G3", "G4", "G5", "G6"]
H_TO_FG = {}
for i, label in enumerate(_FG_LABELS, start=1):
    H_TO_FG[f"H{i:02d}"] = label
    H_TO_FG[f"H{i + 20:02d}"] = label

PREFERRED_GROUP_ORDER = _FG_LABELS

PREMIUM_COLORS = [
    "#4f46e5",  # Indigo (Modern, deep, professional blue-violet)
    "#0ea5e9",  # Sky Blue (Vibrant, high-contrast cyan-blue)
    "#10b981",  # Emerald Green (Clean, professional green)
    "#f59e0b",  # Amber/Orange (Warm accent)
    "#ec4899",  # Pink/Rose (Vibrant second accent)
    "#8b5cf6",  # Violet (Premium secondary color)
    "#f43f5e",  # Rose Red (Vibrant contrast)
    "#14b8a6"   # Teal (Calm, modern cyan-green)
]

# ============================================================
# HELPERS
# ============================================================
def clean_store_code(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.replace(r"\.0$", "", regex=True)

def get_base_id(dept_name: str) -> str:
    name = str(dept_name).strip()
    if name.endswith('N'): return name[:-1]
    return name

def classify_concept(row):
    concept = str(row.get("Store_Concept", "")).strip().upper()
    if "EXPRESS" in concept or "RANKING_BASED" in concept: return "Express"
    if "DIY" in concept: return "DIY"
    if "EXPRESS" in str(row["M_STORE"]).upper() or "RANKING_BASED" in str(row["M_STORE"]).upper(): return "Express"
    return "DIY"

def enrich_summary_df(df):
    """Adds statistical and visual metrics to a summary dataframe."""
    if df.empty: return df
    counts_cols = PREFERRED_GROUP_ORDER
    # Ensure numeric for math
    for c in counts_cols: df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    
    df["Store Count"] = df[counts_cols].sum(axis=1)
    df["Skewness"] = scipy.stats.skew(df[counts_cols].values.astype(float), axis=1, bias=False)
    df["Sparkline"] = df[counts_cols].values.tolist()
    df["Ranking"] = df.groupby(["Department", "Report Set"])["Skewness"].rank(method='min', ascending=True).fillna(0).astype(int)
    
    cols = ["Department", "Report Set", "Category"] + counts_cols + ["Store Count", "Dept Type", "Skewness", "Ranking", "Sparkline"]
    return df.reindex(columns=cols)

def _render_set_on_axes(axes, set_name, sub_defs_fn, dept_group, all_group_df, colors):
    sub_defs_list = sub_defs_fn(all_group_df)
    summary_recs = []

    for col_idx, (ax, (sub_label, _)) in enumerate(zip(axes, sub_defs_list)):
        bar_width = 0.8 / len(dept_group)
        x_vals = np.arange(len(PREFERRED_GROUP_ORDER))
        table_data, row_labels, row_colors = [], [], []

        for d_idx, dept in enumerate(dept_group):
            dept_df = all_group_df[all_group_df["M_STORE_DEPARTMENT"] == dept].copy()
            _, current_mask = sub_defs_fn(dept_df)[col_idx]
            sub_df = dept_df[current_mask].copy()
            
            counts = sub_df.groupby("M_STORE_GROUP")["M_STORE"].count().reindex(PREFERRED_GROUP_ORDER, fill_value=0)
            
            offset = (d_idx - (len(dept_group) - 1) / 2) * bar_width
            
            # Use user-specified colors. If single department, match exactly. If multiple, use emerald shades.
            if len(dept_group) == 1:
                bar_color = "#065F46"  # Dark green
                line_color = "#10B981" # Exact #10B981
            else:
                emerald_bars = ["#065F46", "#047857", "#064E3B", "#14532D", "#166534"]
                emerald_lines = ["#10B981", "#34D399", "#6EE7B7", "#A7F3D0", "#D1FAE5"]
                bar_color = emerald_bars[d_idx % len(emerald_bars)]
                line_color = emerald_lines[d_idx % len(emerald_lines)]
                
            table_data.append([str(int(v)) if v > 0 else "0" for v in counts.values])
            row_labels.append(dept)
            row_colors.append(line_color)

            bars = ax.bar(x_vals + offset, counts.values, width=bar_width, label=dept, color=bar_color, alpha=0.8)
            y = counts.values.astype(float)
            if y.sum() > 0:
                try:
                    z = np.polyfit(x_vals, y, 2); p = np.poly1d(z)
                    x_smooth = np.linspace(0, len(PREFERRED_GROUP_ORDER)-1, 50)
                    y_smooth = np.maximum(0, p(x_smooth))
                    ax.plot(x_smooth, y_smooth, "--", color=line_color, linewidth=1.5)
                except: pass
            
            for bar, val in zip(bars, counts.values):
                if val > 0:
                    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height()+0.2, str(int(val)), ha="center", fontsize=8, color=line_color, fontweight="bold")

            # Add record for summary table
            row = {
                "Department": dept, "Report Set": set_name, "Category": sub_label, "Dept Type": ("New" if str(dept).upper().endswith("N") else "Normal"),
            }
            for k, v in counts.items(): row[k] = v
            summary_recs.append(row)

        ax.set_title(f"[{set_name.split('_')[0]}] {sub_label}", fontsize=9, pad=4, fontweight='bold')
        ax.set_xticks(x_vals); ax.set_xticklabels(PREFERRED_GROUP_ORDER, rotation=45, fontsize=7)
        
        # Table data
        table_rows = []
        dept_counts_list = []
        for d_idx, dept in enumerate(dept_group):
            row_vals = [int(v) for v in table_data[d_idx]]
            dept_counts_list.append(row_vals)
            table_rows.append([str(v) for v in row_vals] + [str(sum(row_vals))])

        if table_rows:
            final_row_labels = row_labels
            final_col_labels = PREFERRED_GROUP_ORDER + ["SUM"]
            t = ax.table(cellText=table_rows, rowLabels=final_row_labels, colLabels=final_col_labels, loc='bottom', bbox=[0, -0.28, 1, 0.22])
            t.auto_set_font_size(False); t.set_fontsize(8)
            for i, r_color in enumerate(row_colors):
                if i < len(final_row_labels):
                    t[(i+1, -1)].get_text().set_color(r_color)
    
    return summary_recs

def plot_single_targeted_set(set_name, sub_defs_fn, groups, df):
    if not groups or len(groups) == 0:
        fig = Figure(figsize=(10, 6), layout='constrained')
        return fig, pd.DataFrame()
    temp_sub_defs = sub_defs_fn(df.iloc[:1])
    n_cols = len(temp_sub_defs)
    colors = PREMIUM_COLORS
    dept_group = groups[0]
    
    width_per_col = 6 if n_cols == 3 else 7
    fig = Figure(figsize=(width_per_col * n_cols, 6.0), layout='constrained')
    axes = [fig.add_subplot(1, n_cols, i+1) for i in range(n_cols)]
    
    all_group_df = df[df["M_STORE_DEPARTMENT"].isin(dept_group)].copy()
    if all_group_df.empty: return fig, pd.DataFrame()
 
    recs = _render_set_on_axes(axes, set_name, sub_defs_fn, dept_group, all_group_df, colors)
    return fig, enrich_summary_df(pd.DataFrame(recs))
 
def plot_dual_targeted_sets(set1_name, sub_defs1_fn, set2_name, sub_defs2_fn, groups, df):
    colors = PREMIUM_COLORS
    if not groups or len(groups) == 0:
        fig = Figure(figsize=(10, 6), layout='constrained')
        return fig, pd.DataFrame()
    dept_group = groups[0]
    all_group_df = df[df["M_STORE_DEPARTMENT"].isin(dept_group)].copy()
    
    # Determine max columns needed
    n_cols1 = len(sub_defs1_fn(df.iloc[:1]))
    n_cols2 = len(sub_defs2_fn(df.iloc[:1]))
    max_cols = max(n_cols1, n_cols2)
    
    width_per_col = 6 if max_cols == 3 else 7
    fig = Figure(figsize=(width_per_col * max_cols, 12.0), layout='constrained')
    
    axes1 = [fig.add_subplot(2, max_cols, i+1) for i in range(n_cols1)]
    axes2 = [fig.add_subplot(2, max_cols, max_cols + i+1) for i in range(n_cols2)]
    
    recs1 = _render_set_on_axes(axes1, set1_name, sub_defs1_fn, dept_group, all_group_df, colors)
    recs2 = _render_set_on_axes(axes2, set2_name, sub_defs2_fn, dept_group, all_group_df, colors)
    
    combined_df = enrich_summary_df(pd.DataFrame(recs1 + recs2))
    return fig, combined_df

def plot_multi_targeted_sets(sets_info, groups, df):
    """Handles 1–4 reporting sets stacked in rows. sets_info = [(set_name, sub_defs_fn), ...]"""
    colors = PREMIUM_COLORS
    if not groups or len(groups) == 0:
        fig = Figure(figsize=(10, 6), layout='constrained')
        return fig, pd.DataFrame()
    dept_group = groups[0]
    all_group_df = df[df["M_STORE_DEPARTMENT"].isin(dept_group)].copy()
    if all_group_df.empty:
        fig = Figure(figsize=(10, 6), layout='constrained')
        return fig, pd.DataFrame()

    n_rows = len(sets_info)
    # Find the max columns needed across all sets
    max_cols = max(len(sub_defs_fn(df.iloc[:1])) for _, sub_defs_fn in sets_info)

    width_per_col = 6 if max_cols == 3 else 7
    row_height = 6.0
    fig = Figure(figsize=(width_per_col * max_cols, row_height * n_rows), layout='constrained')

    all_recs = []
    for row_idx, (set_name, sub_defs_fn) in enumerate(sets_info):
        n_cols = len(sub_defs_fn(df.iloc[:1]))
        axes = [fig.add_subplot(n_rows, max_cols, row_idx * max_cols + col + 1) for col in range(n_cols)]
        recs = _render_set_on_axes(axes, set_name, sub_defs_fn, dept_group, all_group_df, colors)
        all_recs.extend(recs)

    combined_df = enrich_summary_df(pd.DataFrame(all_recs))
    return fig, combined_df

# ============================================================
# PROCESSING & EXPORT LOGIC
# ============================================================
def load_and_merge(base_folder, possys_file, store_info_file):
    """
    Loads and merges POSSYS + Store List data.
    Accepts either full absolute paths or filenames relative to base_folder.
    """
    bp = Path(base_folder)

    # Accept full absolute path OR a filename
    possys_path     = Path(possys_file)    if Path(possys_file).is_absolute()    else bp / possys_file
    store_info_path = Path(store_info_file) if Path(store_info_file).is_absolute() else bp / store_info_file

    # 1. Load POSSYS / Grouping File
    try:
        s_name = get_grouping_sheet_name(possys_path)

        # Auto-detect header row: require STORE and DEPT columns to be named (not Unnamed)
        def _detect_grouping_header(path, sheet):
            """Returns 0 or 1 — the row where STORE and DEPT columns are properly named."""
            path_str = str(path)
            # Clear any cached reads so probes are fresh
            for h in [0, 1]:
                _cache.pop((path_str, str(h)), None)
                try:
                    tmp = pd.read_excel(path, sheet_name=sheet, header=h, engine='calamine', dtype=str, nrows=3)
                    cols = [str(c).strip().lower() for c in tmp.columns]
                    has_store = any(c in ["store", "m_store", "store_code"] for c in cols)
                    has_dept  = any(c in ["dept", "department", "m_store_department"] for c in cols)
                    if has_store and has_dept:
                        return h
                except:
                    pass
            return 1  # default to row 2 as safe fallback for new Grouping format

        h_idx = _detect_grouping_header(possys_path, s_name)
        pdf = get_cached_excel(possys_path, sheet_name=s_name, header=h_idx, engine='calamine', dtype=str)
        pdf.columns = [str(c).strip() for c in pdf.columns]

        def find_col(exact_names, kw, index, default):
            """Priority: 1) Exact name match, 2) Keyword match, 3) Column index fallback"""
            # 1. Exact name match (case-insensitive)
            for name in exact_names:
                for c in pdf.columns:
                    if c.strip().upper() == name.upper(): return c
            # 2. Keyword match
            for c in pdf.columns:
                if all(k.lower() in str(c).lower() for k in kw): return c
            # 3. Column index fallback (A=0, B=1, F=5)
            if index < len(pdf.columns):
                return pdf.columns[index]
            return default

        # For simulation file: STORE=ColA, DEPT=ColB, GRAD_GROUPING=ColF
        m_store_col = find_col(["STORE", "M_STORE", "Store_Code"],   ["store"],          0, "M_STORE")
        m_dept_col  = find_col(["DEPT", "DEPARTMENT", "M_STORE_DEPARTMENT"], ["dept"],   1, "M_STORE_DEPARTMENT")
        m_group_col = find_col(["GRAD_GROUPING", "M_STORE_GROUP"],   ["grad"],           5, "M_STORE_GROUP")

        print(f"[Engine] Mapped: Store='{m_store_col}' | Dept='{m_dept_col}' | Group='{m_group_col}'")

        rename_dict = {}
        if "M_STORE" not in pdf.columns:
            rename_dict[m_store_col] = "M_STORE"
        if "M_STORE_DEPARTMENT" not in pdf.columns:
            rename_dict[m_dept_col] = "M_STORE_DEPARTMENT"
        if "M_STORE_GROUP" not in pdf.columns:
            rename_dict[m_group_col] = "M_STORE_GROUP"

        pdf = pdf.rename(columns=rename_dict)

        pdf["M_STORE"]            = pdf["M_STORE"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        pdf["M_STORE_DEPARTMENT"] = pdf["M_STORE_DEPARTMENT"].astype(str).str.strip()
        pdf["M_STORE_GROUP"]      = pdf["M_STORE_GROUP"].map(H_TO_FG).fillna(pdf["M_STORE_GROUP"])
    except Exception as e:
        print(f"POSSYS Load Error: {e}")
        pdf = pd.DataFrame(columns=["M_STORE", "M_STORE_DEPARTMENT", "M_STORE_GROUP"])

    # 2. Load Store List (Smart Sheet Discovery)
    try:
        xl = pd.ExcelFile(store_info_path, engine='calamine')
        sidf = None
        
        # Look for the sheet that actually contains store size data
        for s_name in xl.sheet_names:
            # Try the first 2 rows as potential headers
            for h_test in [0, 1]:
                temp_df = xl.parse(s_name, header=h_test, nrows=5, dtype=str)
                temp_df.columns = [str(c).lower() for c in temp_df.columns]
                if any(k in " ".join(temp_df.columns) for k in ["store_code", "display area", "sqm", "da"]):
                    sidf = xl.parse(s_name, header=h_test, dtype=str)
                    print(f"[Engine] Found Store Data in sheet: '{s_name}' at Row {h_test+1}")
                    break
            if sidf is not None: break
        
        # Fallback to first sheet if nothing found
        if sidf is None:
            sidf = xl.parse(xl.sheet_names[0], dtype=str)
            
        sidf.columns = [str(c).strip() for c in sidf.columns]

        def find_sid(kw, default):
            for c in sidf.columns:
                if all(k.lower() in c.lower() for k in kw): return c
            return default

        st_code_col    = find_sid(["store", "code"],    "Store_Code")
        st_type_col    = find_sid(["store", "type"],    None) or find_sid(["lot", "type"], None) or "Store_Type"
        st_da_col      = find_sid(["display", "area"], None) or find_sid(["sqm"], None) or find_sid(["da"], "Store_Display_Area")
        st_concept_col = find_sid(["concept"],          "Store_Concept")

        sidf = sidf.rename(columns={
            st_code_col:    "Store_Code",
            st_type_col:    "Store_Type",
            st_da_col:      "Store_Display_Area",
            st_concept_col: "Store_Concept"
        })

        # Ensure standard columns are present
        if "Store_Type" not in sidf.columns:
            sidf["Store_Type"] = "Mall"
        else:
            is_sa = sidf["Store_Type"].astype(str).str.strip().str.upper().isin(["SALONE", "STANDALONE", "SA"])
            sidf["Store_Type"] = np.where(is_sa, "Standalone", "Mall")

        if "Store_Concept" not in sidf.columns:
            sidf["Store_Concept"] = "Mr_DIY"

        # Check if country is Thailand (TH) based on path names
        is_th = "TH" in str(possys_path).upper() or "TH" in str(store_info_path).upper() or "TH" in str(base_folder).upper()

        if "Store _Dummy_Type" not in sidf.columns:
            sidf["Store _Dummy_Type"] = "Department based" if is_th else "Unknown"
        else:
            if is_th:
                sidf["Store _Dummy_Type"] = sidf["Store _Dummy_Type"].fillna("Department based")
                sidf.loc[sidf["Store _Dummy_Type"].astype(str).str.strip().str.lower().isin(["nan", "", "unknown", "none", "n/a"]), "Store _Dummy_Type"] = "Department based"

        # Create a normalized 'match_key' for both sides to ignore prefixes and leading zeros
        def get_match_key(s):
            # Aggressively strip everything to find the numeric ID
            val = "".join(filter(str.isdigit, str(s))).lstrip("0")
            return val if val else "0"

        # Create a prefix-aware key for distinguishing duplicates like BB4003 vs JW4003
        def get_full_key(s):
            return str(s).strip().upper().replace("0", "") # Basic normalization

        pdf["match_key"] = pdf["M_STORE"].apply(get_match_key)
        sidf["match_key"] = sidf["Store_Code"].apply(get_match_key)
        
        # We merge based on numeric ID, but if there are duplicates (like 4003), 
        # we check the prefix from the Store_Code to match the right one.
        # For simplicity and speed, let's use a two-pass approach or just fix the numeric collision
        
        # Prepare store attributes from sidf - keep ALL unique Store_Codes
        sidf["Store_Display_Area_num"] = pd.to_numeric(
            sidf["Store_Display_Area"].astype(str).str.replace(",", ""), errors="coerce"
        ).fillna(0)
        s_attr = sidf.drop_duplicates(subset=["Store_Code"])
        
        # 1. Match by Full Store Code (e.g. 'GW4001' -> 'GW4001')
        full_df = pdf.merge(s_attr[["Store_Code", "Store_Type", "Store_Display_Area_num", "Store_Concept", "Store _Dummy_Type"]], 
                           left_on="M_STORE", right_on="Store_Code", how="left")
        
        # 2. Fallback: If no match by full code, try matching by numeric ID
        unmatched_mask = full_df["Store_Code"].isna()
        if unmatched_mask.any():
            numeric_attr = s_attr.drop_duplicates("match_key", keep="first")
            full_df.loc[unmatched_mask, "Store_Code"] = full_df.loc[unmatched_mask, "match_key"].map(numeric_attr.set_index("match_key")["Store_Code"])
            full_df.loc[unmatched_mask, "Store_Type"] = full_df.loc[unmatched_mask, "match_key"].map(numeric_attr.set_index("match_key")["Store_Type"])
            full_df.loc[unmatched_mask, "Store_Display_Area_num"] = full_df.loc[unmatched_mask, "match_key"].map(numeric_attr.set_index("match_key")["Store_Display_Area_num"])
            full_df.loc[unmatched_mask, "Store_Concept"] = full_df.loc[unmatched_mask, "match_key"].map(numeric_attr.set_index("match_key")["Store_Concept"])
            full_df.loc[unmatched_mask, "Store _Dummy_Type"] = full_df.loc[unmatched_mask, "match_key"].map(numeric_attr.set_index("match_key")["Store _Dummy_Type"])

        # Priority Check: Ensure Concept is cleaned
        full_df["Store_Concept"] = full_df["Store_Concept"].fillna("MR_DIY")
        
        # INJECT MISSING SPECIAL STORES (Express or Ranking_Based)
        # These are included in analysis and assigned G6 even if missing from POSSYS
        is_special = sidf["Store_Concept"].astype(str).str.upper().str.contains("EXPRESS") | \
                     sidf["Store _Dummy_Type"].astype(str).str.upper().str.contains("RANKING_BASED")
        
        special_sidf = sidf[is_special].copy()
        
        # We only inject stores that are NOT already in full_df
        missing_special_codes = set(special_sidf["Store_Code"]) - set(full_df["Store_Code"].dropna())
        
        if missing_special_codes:
            v_depts = pdf["M_STORE_DEPARTMENT"].dropna().unique()
            new_rows = []
            for code in missing_special_codes:
                s_data = special_sidf[special_sidf["Store_Code"] == code].iloc[0]
                for dept in v_depts:
                    new_rows.append({
                        "M_STORE": s_data["Store_Code"],
                        "M_STORE_DEPARTMENT": dept,
                        "M_STORE_GROUP": "G6",
                        "Store_Code": s_data["Store_Code"],
                        "Store_Type": s_data["Store_Type"],
                        "Store_Display_Area_num": s_data["Store_Display_Area_num"],
                        "Store_Concept": s_data["Store_Concept"],
                        "Store _Dummy_Type": s_data["Store _Dummy_Type"]
                    })
            
            if new_rows:
                inj_df = pd.DataFrame(new_rows)
                full_df = pd.concat([full_df, inj_df], ignore_index=True)
                print(f"[Engine] Injected {len(missing_special_codes)} missing Special stores (Express/Ranking) across {len(v_depts)} depts.")

        # Capture excluded stores
        # A store is excluded if it is in the Store List (sidf) but NOT in our analyzed data (full_df)
        # and it's not an Express store (since Express stores are always injected/analyzed)
        analyzed_codes = set(full_df["Store_Code"].dropna().unique())
        is_special_sidf = sidf["Store_Concept"].astype(str).str.upper().str.contains("EXPRESS") | \
                          sidf["Store _Dummy_Type"].astype(str).str.upper().str.contains("RANKING_BASED")
        
        excluded_df = sidf[
            (~sidf["Store_Code"].isin(analyzed_codes)) & (~is_special_sidf)
        ].copy()

        # Cleanup
        full_df = full_df.drop(columns=["match_key"])
        
        if "Store_Display_Area_num" not in full_df.columns:
            full_df["Store_Display_Area_num"] = 0
            
        full_df["Concept"] = full_df.apply(classify_concept, axis=1)

        # AUTO-ASSIGN G6 TO MR_EXPRESS STORES
        # If a store is 'Express', force its group to 'G6' regardless of what's in the file
        is_express = full_df["Concept"] == "Express"
        if is_express.any():
            full_df.loc[is_express, "M_STORE_GROUP"] = "G6"
            print(f"[Engine] Auto-assigned G6 to {is_express.sum()} MR_EXPRESS store records.")

        return full_df, excluded_df
    except Exception as e:
        print(f"Store List Load Error: {e}")
        # Ensure fallback PDF has the expected columns even if merge failed
        for col in ["Store_Type", "Store_Display_Area_num", "Store_Concept", "Concept"]:
            if col not in pdf.columns:
                pdf[col] = "Unknown" if "Concept" in col or "Type" in col else 0
        return pdf, pd.DataFrame()

def get_defs(df, da_low, da_high, sa_split, da_slicer, set5_mall_split=10000, set5_sa_split=8000):
    def s1(d): return [(f"DA < {da_low:,}", d["Store_Display_Area_num"] < da_low), (f"{da_low:,}<=DA<{da_high:,}", (d["Store_Display_Area_num"]>=da_low)&(d["Store_Display_Area_num"]<da_high)), (f"DA >= {da_high:,}", d["Store_Display_Area_num"]>=da_high)]
    def s2(d): return [("Mall", d["Store_Type"].str.lower()=="mall"), ("Standalone (SA)", d["Store_Type"].str.lower()=="standalone")]
    def s3(d):
        m, s = d["Store_Type"].str.lower()=="mall", d["Store_Type"].str.lower()=="standalone"
        return [(f"SA < {sa_split:,}", s&(d["Store_Display_Area_num"]<sa_split)), (f"SA >= {sa_split:,}", s&(d["Store_Display_Area_num"]>=sa_split)), ("Mall", m)]
    def s4(d):
        m, s = d["Store_Type"].str.lower()=="mall", d["Store_Type"].str.lower()=="standalone"
        da = d["Store_Display_Area_num"]
        return [
            (f"Mall+SA < {da_slicer:,}", da < da_slicer),
            (f"SA >= {da_slicer:,}", s & (da >= da_slicer)),
            (f"Mall >= {da_slicer:,}", m & (da >= da_slicer))
        ]
    def s5(d):
        m, s = d["Store_Type"].str.lower()=="mall", d["Store_Type"].str.lower()=="standalone"
        da = d["Store_Display_Area_num"]
        return [
            (f"SA < {set5_sa_split:,}", s & (da < set5_sa_split)),
            (f"SA >= {set5_sa_split:,}", s & (da >= set5_sa_split)),
            (f"Mall < {set5_mall_split:,}", m & (da < set5_mall_split)),
            (f"Mall >= {set5_mall_split:,}", m & (da >= set5_mall_split))
        ]
    return {"Set1_DA": s1, "Set2_StoreType": s2, "Set3_Type_DA": s3, "Set4_CurrentSetting": s4, "Set5_Mall_SA_Split": s5}

def generate_consolidated_data(df, da_low, da_high, sa_split, da_slicer, set5_mall_split=10000, set5_sa_split=8000):
    defs = get_defs(df, da_low, da_high, sa_split, da_slicer, set5_mall_split, set5_sa_split)
    v_depts = sorted(df["M_STORE_DEPARTMENT"].dropna().unique().tolist())
    all_rows = []
    
    # Process each set/category across all departments
    for set_name, def_fn in defs.items():
        for sub_label, mask in def_fn(df):
            sub_df = df[mask]
            
            if not sub_df.empty:
                # Group by Dept and Store Group
                counts = sub_df.groupby(["M_STORE_DEPARTMENT", "M_STORE_GROUP"])["M_STORE"].count()
                counts_unstacked = counts.unstack(fill_value=0).reindex(columns=PREFERRED_GROUP_ORDER, fill_value=0)
            else:
                counts_unstacked = pd.DataFrame(columns=PREFERRED_GROUP_ORDER, index=v_depts).fillna(0)
                
            # Ensure every department has a row
            missing_depts = [d for d in v_depts if d not in counts_unstacked.index]
            if missing_depts:
                extra = pd.DataFrame(0, index=missing_depts, columns=PREFERRED_GROUP_ORDER)
                counts_unstacked = pd.concat([counts_unstacked, extra])
            
            counts_unstacked.index.name = "Department"
            cat_df = counts_unstacked.reset_index()
            cat_df.insert(1, "Report Set", set_name)
            cat_df.insert(2, "Category", sub_label)
            
            # Label Dept Type
            cat_df["Dept Type"] = "Normal"
            cat_df.loc[cat_df["Department"].astype(str).str.upper().str.endswith("N"), "Dept Type"] = "New"
            
            all_rows.append(cat_df)
            
    if not all_rows:
        return enrich_summary_df(pd.DataFrame())

    combined_df = pd.concat(all_rows, ignore_index=True)
    
    # Sort by Department name
    combined_df = combined_df.sort_values(by=["Report Set", "Category", "Department"]).reset_index(drop=True)
    
    return enrich_summary_df(combined_df)

def generate_quarter_data(summary_df):
    """Aggregates F1-G6 into Q1-Q4 for Quarter Analysis."""
    if summary_df.empty: return summary_df
    
    df = summary_df.copy()
    # Define Quarters
    df["Q1"] = df["F1"] + df["F2"] + df["F3"]
    df["Q2"] = df["F4"] + df["F5"] + df["F6"]
    df["Q3"] = df["G1"] + df["G2"] + df["G3"]
    df["Q4"] = df["G4"] + df["G5"] + df["G6"]
    
    # Calculate Skewness for Quarters
    q_cols = ["Q1", "Q2", "Q3", "Q4"]
    df["Skewness"] = df[q_cols].apply(lambda x: pd.Series(x).skew(), axis=1)
    df["Sparkline"] = df[q_cols].values.tolist()
    
    # Keep metadata and Quarters
    meta_cols = ["Department", "Report Set", "Category", "Store Count", "Dept Type"]
    final_cols = meta_cols[:3] + q_cols + meta_cols[3:] + ["Skewness", "Ranking", "Sparkline"]
    return df.reindex(columns=final_cols)

def plot_display_area_distribution(df, set_name, da_low, da_high, sa_split, da_slicer, set_compare="None", set5_mall_split=10000, set5_sa_split=8000):
    """Generates a box-and-whisker plot that adapts to the selected Report Set(s)."""
    
    set_names = {
        "Set1_DA": "Set 1: Display Area",
        "Set2_StoreType": "Set 2: Store Type",
        "Set3_Type_DA": "Set 3: SA Split",
        "Set4_CurrentSetting": "Set 4: DA Split",
        "Set5_Mall_SA_Split": "Set 5: Mall/SA Split"
    }
    
    if df.empty: 
        fig = Figure(figsize=(10, 6), layout='constrained')
        return fig
        
    # CRITICAL: We must look at unique stores only for the distribution
    unique_stores = df.drop_duplicates("M_STORE").copy()
    defs = get_defs(unique_stores, da_low, da_high, sa_split, da_slicer, set5_mall_split=set5_mall_split, set5_sa_split=set5_sa_split)
    
    # Decide how many sets to plot
    sets_to_plot = [set_name]
    if isinstance(set_compare, list):
        for cs in set_compare:
            if cs != set_name and cs in defs and cs not in sets_to_plot:
                sets_to_plot.append(cs)
    elif isinstance(set_compare, str) and set_compare != "None" and set_compare != set_name and set_compare in defs:
        sets_to_plot.append(set_compare)
        
    n_sets = len(sets_to_plot)
    
    if n_sets == 1:
        fig = Figure(figsize=(10, 6), layout='constrained')
        axes = [fig.add_subplot(111)]
    elif n_sets == 2:
        fig = Figure(figsize=(16, 6), layout='constrained')
        axes = [fig.add_subplot(121), fig.add_subplot(122)]
    elif n_sets == 3:
        fig = Figure(figsize=(18, 6), layout='constrained')
        axes = [fig.add_subplot(131), fig.add_subplot(132), fig.add_subplot(133)]
    else:
        fig = Figure(figsize=(16, 10), layout='constrained')
        axes = [fig.add_subplot(221), fig.add_subplot(222), fig.add_subplot(223), fig.add_subplot(224)]
        
    for ax_idx, ax in enumerate(axes):
        s_name = sets_to_plot[ax_idx]
        if s_name not in defs: 
            s_name = "Set2_StoreType" # Default fallback
            
        sub_defs = defs[s_name](unique_stores)
        data = []
        labels = []
        for sub_label, mask in sub_defs:
            d_series = unique_stores[mask]["Store_Display_Area_num"].dropna()
            count = len(d_series)
            data.append(d_series)
            labels.append(f"{sub_label}\n({count} Stores)")
            
        # Create boxplot
        bp = ax.boxplot(data, labels=labels, patch_artist=True, showmeans=True,
                       widths=0.4,
                       medianprops={'color': 'black', 'linewidth': 2},
                       meanprops={'marker': 'x', 'markeredgecolor': 'black', 'markersize': 10},
                       flierprops={'marker': 'o', 'markersize': 4, 'markerfacecolor': 'black', 'markeredgecolor': 'black'})
                       
        # Coloring
        colors = ['#60a5fa', '#34d399', '#f87171', '#fbbf24', '#a78bfa'] 
        for i, patch in enumerate(bp['boxes']):
            color = colors[i % len(colors)]
            patch.set_facecolor(color)
            patch.set_alpha(0.7)
            
        ax.set_ylabel("Display Area", fontsize=12)
        ax.set_title(set_names.get(s_name, s_name), fontsize=13, fontweight='bold', pad=10)
        ax.grid(True, axis='y', linestyle='--', alpha=0.7)
        
        all_stats = []
        category_stats = []
        for i, d in enumerate(data):
            if len(d) == 0:
                category_stats.append({})
                continue
                
            mean_val = d.mean()
            ax.text(i + 1.05, mean_val, f"{mean_val:,.0f}", ha='left', va='bottom', fontsize=7, color='black')
            
            q1, q3 = d.quantile(0.25), d.quantile(0.75)
            iqr = q3 - q1
            s = {
                "Max": d.max(),
                "Upper": q3 + (1.5 * iqr),
                "Q3": q3,
                "Median": d.median(),
                "Q1": q1,
                "Lower": max(0, q1 - (1.5 * iqr)),
                "Min": d.min()
            }
            category_stats.append(s)
            all_stats.extend(s.values())
            all_stats.append(mean_val)
            
        if all_stats:
            s_min, s_max = min(all_stats), max(all_stats)
            ax.set_ylim(s_min, s_max * 1.1)
            
        y_min, y_max = ax.get_ylim()
        y_range = y_max - y_min
        
        for i, stats in enumerate(category_stats):
            if not stats: continue
            sorted_stats = sorted(stats.items(), key=lambda x: x[1])
            last_y = -999999
            for label, val in sorted_stats:
                if label in ["Q1", "Q3"]:
                    continue
                y_pos = val
                if abs(y_pos - last_y) < (y_range * 0.02):
                    y_pos += (y_range * 0.01)
                ax.text(i + 1.25, y_pos, f"{label}: {val:,.0f}", va='center', fontsize=9)
                last_y = y_pos
            
    return fig

def save_consolidated_excel(df, save_path):
    """
    Saves dataframe to Excel with Formulas and Sparklines using XlsxWriter.
    Columns:
    A:Department, B:Report Set, C:Category, D-O:F1-G6, 
    P:Store Count, Q:Dept Type, R:Skewness, S:Ranking, T:Sparkline
    """
    n_rows = len(df)
    writer = pd.ExcelWriter(save_path, engine='xlsxwriter')
    
    # 1. Insert 'Store Count' placeholder in the dataframe at index 15 (Column P)
    # df currently: Dept, Set, Category, F1...G6, Dept Type
    # Insert Store Count at 15
    if "Store Count" not in df.columns:
        df.insert(15, "Store Count", 0)
    
    # Write to Excel (Let pandas handle the headers for A-Q)
    df.to_excel(writer, index=False, sheet_name='Summary')
    
    workbook  = writer.book
    worksheet = writer.sheets['Summary']
    
    # Header format
    header_fmt = workbook.add_format({'bold': True, 'bg_color': '#DDEBF7', 'border': 1})
    
    # Add Column Headers for the dynamic formula columns (R, S, T)
    # (P and Q headers are already written by df.to_excel)
    worksheet.write('R1', 'Skewness', header_fmt)
    worksheet.write('S1', 'Ranking', header_fmt)
    worksheet.write('T1', 'Sparkline', header_fmt)

    # Freeze top row
    worksheet.freeze_panes(1, 0)
    
    for row_num in range(1, n_rows + 1):
        idx = row_num + 1
        
        # P: Store Count =SUM(D2:O2) (Index 15)
        worksheet.write_formula(row_num, 15, f'=SUM(D{idx}:O{idx})')
        
        # R: Skewness =SKEW(D2:O2) (Index 17)
        worksheet.write_formula(row_num, 17, f'=SKEW(D{idx}:O{idx})')
        
        # S: Ranking =COUNTIFS(...) (Index 18)
        # Using R (Index 17) for Skewness comparison
        formula_rank = (
            f'=COUNTIFS($A$2:$A${n_rows+1}, A{idx}, '
            f'$B$2:$B${n_rows+1}, B{idx}, '
            f'$R$2:$R${n_rows+1}, "<"&R{idx}) + 1'
        )
        worksheet.write_formula(row_num, 18, formula_rank)
        
        # T: Sparkline (Trend D to O)
        worksheet.add_sparkline(row_num, 19, {
            'range': f'Summary!D{idx}:O{idx}',
            'type': 'line',
            'style': 18,
        })

    writer.close()

def get_available_groups(possys_path):
    try:
        s_name = get_grouping_sheet_name(possys_path)

        # Auto-detect header: need both STORE and DEPT columns to be named
        h_idx = 1  # default to row 2 for new Grouping format
        path_str = str(possys_path)
        for h in [0, 1]:
            _cache.pop((path_str, str(h)), None)
            try:
                tmp = pd.read_excel(possys_path, sheet_name=s_name, header=h, engine='calamine', dtype=str, nrows=3)
                cols = [str(c).strip().lower() for c in tmp.columns]
                has_store = any(c in ["store", "m_store", "store_code"] for c in cols)
                has_dept  = any(c in ["dept", "department", "m_store_department"] for c in cols)
                if has_store and has_dept:
                    h_idx = h
                    break
            except:
                pass

        df = get_cached_excel(possys_path, sheet_name=s_name, header=h_idx, engine='calamine', dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        # Fuzzy-find OR fallback to Column B (Index 1) for Department
        dept_col = next(
            (c for c in df.columns if "department" in c.lower() or c.upper() == "DEPT"),
            df.columns[1] if len(df.columns) > 1 else (df.columns[0] if len(df.columns) > 0 else None)
        )
        
        if dept_col is None:
            return []

        depts = df[dept_col].dropna().unique()
        return sorted(list(set(get_base_id(d) for d in depts)))
    except:
        return []


# ============================================================
# SALES DATA PROCESSING (moved from main.py)
# ============================================================

def load_dept_details(details_path, country="ID"):
    """
    Loads department details from either a Google Sheet URL or a local Excel/CSV path.
    Supports a pipe-delimited list of paths to try URL first, then fallbacks.
    Uses urllib with browser-like headers to avoid 401 Unauthorized from Google.
    """
    import urllib.request
    import io

    parts = str(details_path).split("|")

    # 1. Try URL(s) first — use urllib with browser headers to bypass Google's 401 filter
    for part in parts:
        part = part.strip()
        if part.startswith("http"):
            try:
                req = urllib.request.Request(
                    part,
                    headers={
                        "User-Agent": (
                            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                            "AppleWebKit/537.36 (KHTML, like Gecko) "
                            "Chrome/124.0.0.0 Safari/537.36"
                        ),
                        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                    }
                )
                with urllib.request.urlopen(req, timeout=10) as resp:
                    raw = resp.read().decode("utf-8")
                df = pd.read_csv(io.StringIO(raw), dtype=str)
                df.columns = [str(c).strip() for c in df.columns]
                print(f"[Engine] Successfully loaded Department Details from Google Sheet.")
                return df
            except Exception as e:
                # Silently fall through to local fallback — 401/network errors are expected
                # when the sheet is restricted or no internet access is available.
                print(f"[Engine] Google Sheet unavailable ({type(e).__name__}), using local fallback.")

    # 2. Try local fallback file paths (e.g. from Downloads/Department Details.xlsx)
    import os
    for part in parts:
        part = part.strip()
        if not part.startswith("http") and os.path.exists(part):
            try:
                if part.endswith(".csv"):
                    df = pd.read_csv(part, dtype=str)
                    df.columns = [str(c).strip() for c in df.columns]
                    print(f"[Engine] Successfully loaded Department Details from local CSV: {part}")
                    return df
                else:
                    # Excel fallback
                    import openpyxl
                    wb_details = openpyxl.load_workbook(part, data_only=True)
                    
                    expected_sheets = []
                    c_upper = country.upper()
                    if c_upper == "ID":
                        expected_sheets = ["DEPARTMENT ID", "DEPARTMENT_ID", "ID REMARK", "ID_REMARK", "REMARK CRITERIA"]
                    elif c_upper == "MY":
                        expected_sheets = ["DEPARTMENT MY", "DEPARTMENT_MY", "MY REMARK", "MY"]
                    elif c_upper == "BR":
                        expected_sheets = ["DEPARTMENT BR SG", "DEPARTMENT BR", "DEPARTMENT_BR", "BR REMARK", "BR"]
                    elif c_upper == "SG":
                        expected_sheets = ["DEPARTMENT BR SG", "DEPARTMENT SG", "DEPARTMENT_SG", "SG REMARK", "SG"]
                    elif c_upper == "TH":
                        expected_sheets = ["DEPARTMENT TH", "DEPARTMENT_TH", "TH REMARK", "TH"]
                    elif c_upper == "IN":
                        expected_sheets = ["DEPARTMENT IN", "DEPARTMENT_IN", "IN REMARK", "IN"]

                    expected_sheets.append(f"DEPARTMENT {c_upper}")
                    expected_sheets.append(f"DEPARTMENT_{c_upper}")
                    expected_sheets.append(f"{c_upper} REMARK")
                    expected_sheets.append(f"{c_upper}_REMARK")
                    expected_sheets = [s.upper() for s in expected_sheets]

                    sheet_name = None
                    for exp in expected_sheets:
                        for s in wb_details.sheetnames:
                            if s.strip().upper() == exp:
                                sheet_name = s
                                break
                        if sheet_name:
                            break

                    if not sheet_name:
                        for s in wb_details.sheetnames:
                            if c_upper in s.upper():
                                sheet_name = s
                                break
                    if not sheet_name:
                        for s in wb_details.sheetnames:
                            if "DEPT" in s.upper() or "REMARK" in s.upper():
                                sheet_name = s
                                break
                    if not sheet_name:
                        sheet_name = wb_details.sheetnames[0]

                    ws = wb_details[sheet_name]
                    rows_data = []
                    for row in ws.iter_rows(values_only=True):
                        rows_data.append(row)
                    
                    if len(rows_data) > 0:
                        headers = [str(c).strip() if c is not None else "" for c in rows_data[0]]
                        data_rows = rows_data[1:]
                        # Convert to DataFrame
                        df = pd.DataFrame(data_rows, columns=headers, dtype=str)
                        df.columns = [str(c).strip() for c in df.columns]
                        print(f"[Engine] Successfully loaded Department Details from local Excel: {part} (Sheet: {sheet_name})")
                        return df
            except Exception as ex:
                print(f"[Engine] Error loading from local fallback {part}: {ex}")

    return None


def enrich_sales_df(df, dept_info, details_path="", country="ID"):
    """
    Combines Sales/Balance data with POSSYS Dept info and applies Dept Type mapping.
    Uses fuzzy column detection to handle different report headers.
    """
    if df.empty: return df
    df = df.copy()
    
    # Clean Column Names
    df.columns = [str(c).strip() for c in df.columns]
    
    def find_col(keywords, mandatory=False, default=None):
        for c in df.columns:
            if all(k.lower() in c.lower() for k in keywords):
                return c
        if mandatory: return default if default else keywords[0].upper()
        return None

    # Detect key columns — supports both Existing (AVG_TOP2_*) and New Store (SALES_AMT_PER_30DAYS) files
    store_col = find_col(["store"], mandatory=True, default="STORE")
    # Sales: prefer AVG_TOP2 first, then per-30-days variant
    sales_col = find_col(["sales", "avg"]) or find_col(["sales_amt"]) or find_col(["sales"])
    if not sales_col: sales_col = "AVG_TOP2_TOTAL_AMT_SALES"
    # Balance: prefer AVG_TOP2 first, then raw BALANCE_AMT
    bal_col = find_col(["balance", "avg"]) or find_col(["balance_amt"]) or find_col(["balance"])
    if not bal_col: bal_col = "AVG_TOP2_TOTAL_AMT_BALANCE"
    dept_col  = find_col(["department"]) or "DEPARTMENT"

    # Ensure numeric for calculations (strip commas first to avoid NaN on formatted numbers)
    df[sales_col] = pd.to_numeric(df[sales_col].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
    df[bal_col]   = pd.to_numeric(df[bal_col].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
    
    # Calculate Turnover
    df["Turnover_Val"] = np.where(df[bal_col] > 0, (df[sales_col] / df[bal_col] * 100), 0.0)
    df["Turnover"]     = df["Turnover_Val"].map(lambda x: f"{x:.2f}%")
    
    # Normalized Dept + Dept Type
    if dept_col in df.columns:
        dept_str = df[dept_col].astype(str)
        df["Normalized Dept"] = dept_str.str.replace("N", "", regex=False).str.strip()
        df["Dept Type"]       = np.where(dept_str.str.endswith("N"), "New", "Normal")
    else:
        df["Normalized Dept"] = "Unknown"
        df["Dept Type"]       = "Unknown"

    # Merge with Dept Info from POSSYS (for filtering/context)
    dept_info.columns = ["M_STORE_JOIN", "M_POSSYS_DEPT"]
    df = df.merge(dept_info, left_on=store_col, right_on="M_STORE_JOIN", how="left")
    
    # Dept Details lookup (Remark)
    d_df = load_dept_details(details_path, country)
    if d_df is not None:
        try:
            # Fuzzy find columns: DEPARTMENT and ID Remark (case-insensitive)
            dept_col_detail = next((c for c in d_df.columns if "DEPARTMENT" in c.upper()), None)
            
            # Fuzzy find remark column dynamically based on country
            remark_col_detail = None
            # 1. Try f"{country} Remark" or f"{country} Details" (e.g. BR Remark / SG Remark)
            for c in d_df.columns:
                if country.upper() in c.upper() and ("REMARK" in c.upper() or "DETAIL" in c.upper() or "REM" in c.upper()):
                    remark_col_detail = c
                    break
            # 2. Try C2/D2 for Malaysia (MY)
            if not remark_col_detail and country.upper() == "MY":
                for c in d_df.columns:
                    if "C2" in c.upper() or "D2" in c.upper() or "REMARK" in c.upper():
                        remark_col_detail = c
                        break
            # 3. Fallback to standard Remark or Detail search
            if not remark_col_detail:
                remark_col_detail = next((c for c in d_df.columns if "REMARK" in c.upper() or "DETAIL" in c.upper() or "REM" in c.upper()), None)
            
            if dept_col_detail and remark_col_detail:
                # Clean and normalize department codes to avoid mismatch on .0, trailing N, and whitespace
                d_df_clean = d_df.copy()
                def clean_code(val):
                    if val is None or pd.isna(val):
                        return ""
                    s = str(val).strip()
                    if s.endswith(".0"):
                        s = s[:-2]
                    s = s.replace("N", "").strip()
                    return s
                d_df_clean[dept_col_detail] = d_df_clean[dept_col_detail].apply(clean_code)
                raw_mapping = d_df_clean.set_index(dept_col_detail)[remark_col_detail].to_dict()
                
                mapping = {}
                for k, v in raw_mapping.items():
                    val = str(v).strip() if pd.notna(v) else ""
                    if val.lower() not in ["none", "nan", "unknown", ""]:
                        mapping[k] = val
                    else:
                        mapping[k] = ""
                
                # Also normalize/clean Normalized Dept for mapping to succeed
                df["Normalized Dept"] = df["Normalized Dept"].apply(clean_code)
                df["Dept Details"] = df["Normalized Dept"].map(mapping).fillna("")
            else:
                df["Dept Details"] = "Unknown"
        except Exception as e:
            print(f"[Engine] Error parsing Department Details: {e}")
            df["Dept Details"] = "Unknown"
    else:
        df["Dept Details"] = "Unknown"

    return df

    return df


def lookup_store_sets(sales_df, store_path, config):
    """
    Classifies each store row into Set1–Set4 based on Display Area and Store Type.
    Also adds 'Display Area' and 'Store Type' columns.

    Parameters:
        sales_df   : Enriched sales DataFrame (must have 'STORE' column)
        store_path : Path to 'ID store list.xlsx'
        config     : Dict with keys: da_low, da_high, sa_split, da_slicer

    Returns:
        sales_df with added Set columns, reordered appropriately
    """
    sales_df = sales_df.copy()
    target_cols = ["Display Area", "Store Type", "Set1_DA", "Set2_StoreType", "Set3_Type_DA", "Set4_CurrentSetting", "Set5_Mall_SA_Split"]

    # Fuzzy detection for Store column
    def find_col(kw, df):
        for c in df.columns:
            if all(k.lower() in c.lower() for k in kw): return c
        return None
    store_col = find_col(["store"], sales_df) or "STORE"

    status = "Searching..."
    if not os.path.exists(store_path): status = "File Missing"
    elif store_col not in sales_df.columns: status = "No STORE Col"

    for c in target_cols:
        sales_df[c] = status

    if os.path.exists(store_path) and store_col in sales_df.columns:
        try:
            # Smart sheet discovery for Store List
            def _find_store_sheet(path):
                xl = pd.ExcelFile(path, engine='calamine')
                for s_name in xl.sheet_names:
                    for h_test in [0, 1]:
                        tmp = xl.parse(s_name, header=h_test, nrows=5, dtype=str)
                        cols_lower = [str(c).strip().lower() for c in tmp.columns]
                        if any(k in " ".join(cols_lower) for k in ["store_code", "store code", "display area", "sqm", "da"]):
                            return xl.parse(s_name, header=h_test, dtype=str)
                # Fallback: known sheet name
                if "ID STORES SHORT NAME" in xl.sheet_names:
                    return xl.parse("ID STORES SHORT NAME", header=1, dtype=str)
                return xl.parse(xl.sheet_names[0], dtype=str)
            st_df = _find_store_sheet(store_path)
            if not st_df.empty:
                st_df.columns = [str(c).strip() for c in st_df.columns]
                
                # Dynamic Column Detection for Store List
                def f_st(kw):
                    for c in st_df.columns:
                        if all(k.lower() in str(c).lower() for k in kw): return c
                    return None
                
                st_code_col = f_st(["store", "code"]) or f_st(["code"]) or st_df.columns[2]
                st_da_col   = f_st(["display", "area"]) or f_st(["sqm"]) or f_st(["da"]) or (st_df.columns[16] if len(st_df.columns)>16 else None)
                st_type_col = f_st(["store", "type"]) or f_st(["lot", "type"]) or f_st(["type"])

                st_df["M_STORE_JOIN"] = st_df[st_code_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
                lookup_keys = sales_df[store_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)

                da_low    = float(config.get('da_low',    7500))
                da_high   = float(config.get('da_high',   9500))
                sa_split  = float(config.get('sa_split',  600))
                da_slicer = float(config.get('da_slicer', 8000))
                set5_mall = float(config.get('set5_mall', 10000))
                set5_sa   = float(config.get('set5_sa',   8000))

                if st_da_col:
                    da_values = lookup_keys.map(st_df.set_index("M_STORE_JOIN")[st_da_col].to_dict())
                    sales_df["Display Area"] = da_values.fillna("Not Found")
                    n_da = pd.to_numeric(da_values, errors='coerce')
                    sales_df["Set1_DA"] = np.where(
                        n_da.isna(), "Store Closed",
                        np.where(n_da < da_low,  f"<{int(da_low)}",
                        np.where(n_da >= da_high, f">={int(da_high)}",
                        f">={int(da_low)},<{int(da_high)}")))

                # Load or default store type values
                if st_type_col:
                    type_values = lookup_keys.map(st_df.set_index("M_STORE_JOIN")[st_type_col].to_dict())
                    type_values = type_values.fillna("Not Found")
                else:
                    type_values = lookup_keys.map(st_df.set_index("M_STORE_JOIN")[st_code_col].to_dict())
                    type_values = np.where(type_values.isna(), "Not Found", "Mall")
                    type_values = pd.Series(type_values, index=lookup_keys.index)

                # Clean and normalize to standard "Standalone" / "Mall" / "Store Closed"
                is_sa_val = type_values.astype(str).str.strip().str.upper().isin(["SALONE", "STANDALONE", "SA", "STAND ALONE"])
                is_closed_val = type_values.astype(str).str.strip().str.upper().isin(["NAN", "NONE", "", "STORE CLOSED", "CLOSED"])
                
                normalized_type = np.where(is_closed_val, "Store Closed", np.where(is_sa_val, "Standalone", "Mall"))
                
                sales_df["Store Type"] = normalized_type
                sales_df["Set2_StoreType"] = normalized_type

                if st_da_col:
                    s2_upper  = sales_df["Set2_StoreType"].astype(str).str.strip().str.upper()
                    n_da2     = pd.to_numeric(sales_df["Display Area"], errors='coerce')
                    is_closed = (s2_upper == "STORE CLOSED") | n_da2.isna()
                    is_mall   = s2_upper.str.contains("MALL",       na=False)
                    is_sa     = s2_upper.str.contains("STANDALONE", na=False)

                    sales_df["Set3_Type_DA"] = np.where(
                        is_closed, "Store Closed",
                        np.where(is_mall, "Mall",
                        np.where(is_sa & (n_da2 < sa_split),  f"SA (<{int(sa_split)})",
                        np.where(is_sa & (n_da2 >= sa_split), f"SA (>={int(sa_split)})",
                        sales_df["Set2_StoreType"].astype(str).str.strip()))))

                    sales_df["Set4_CurrentSetting"] = np.where(
                        is_closed, "Store Closed",
                        np.where(is_mall & (n_da2 >= da_slicer), f"Mall >= {int(da_slicer):,}",
                        np.where(is_sa   & (n_da2 >= da_slicer), f"SA >= {int(da_slicer):,}",
                        f"Mall+SA < {int(da_slicer):,}")))

                    sales_df["Set5_Mall_SA_Split"] = np.where(
                        is_closed, "Store Closed",
                        np.where(is_mall & (n_da2 < set5_mall), f"Mall < {int(set5_mall):,}",
                        np.where(is_mall & (n_da2 >= set5_mall), f"Mall >= {int(set5_mall):,}",
                        np.where(is_sa & (n_da2 < set5_sa), f"SA < {int(set5_sa):,}",
                        np.where(is_sa & (n_da2 >= set5_sa), f"SA >= {int(set5_sa):,}",
                        sales_df["Set2_StoreType"].astype(str).str.strip())))))
        except Exception as e:
            print(f"Store lookup load error: {e}")

    # Reorder columns so Set columns appear after Turnover
    cols = list(sales_df.columns)
    for c in target_cols:
        if c in cols: cols.remove(c)
    
    t_idx = cols.index("Turnover") if "Turnover" in cols else len(cols)-1
    for i, c in enumerate(target_cols):
        cols.insert(t_idx + 1 + i, c)
    
    return sales_df[cols]


def build_sales_performance_pivot(df, set_col, noise_values=None):
    """
    Builds a flat pivot DataFrame for a given reporting set column.
    Uses fuzzy column names for Sales and Turnover values.
    """
    if noise_values is None:
        noise_values = ["Store Closed", "Not Found", "Searching..."]
    
    # Fuzzy find the metric columns
    def find_col(kw):
        for c in df.columns:
            if all(k.lower() in c.lower() for k in kw): return c
        return None
    
    s_col = find_col(["sales", "avg"]) or "AVG_TOP2_TOTAL_AMT_SALES"
    t_col = "Turnover_Val" if "Turnover_Val" in df.columns else (find_col(["turnover"]) or "Turnover")
    d_col = find_col(["normalized", "dept"]) or "Normalized Dept"

    try:
        p = pd.pivot_table(
            df,
            values=[s_col, t_col],
            index=[d_col],
            columns=[set_col],
            aggfunc='mean'
        ).fillna(0)

        available_cats = df[set_col].unique()
        da_order = sorted([s for s in available_cats if s not in noise_values])

        headers = ["DEPARTMENT"]
        final_cols = []
        for da in da_order:
            final_cols.append((s_col, da))
            headers.append(f"{da} Sales")
            final_cols.append((t_col, da))
            headers.append(f"{da} Turnover")

        export_data = []
        for dept in p.index:
            row = [dept]
            for col_key in final_cols:
                if col_key in p.columns:
                    val = p.loc[dept, col_key]
                    if isinstance(val, pd.Series): val = val.iloc[0]
                    row.append(float(val))
                else:
                    row.append(0.0)
            export_data.append(row)

        return pd.DataFrame(export_data, columns=headers)
    except Exception as e:
        print(f"build_sales_performance_pivot failed for {set_col}: {e}")
        return None


def compute_workspace_pivot(df, row_c, col_c, val_items, filter_values):
    """
    Computes a pivot/crosstab result for the workspace pivot card.

    Parameters:
        df           : Source DataFrame
        row_c        : List of row field names
        col_c        : List of column field names
        val_items    : List of (field, agg_func, display_name) tuples
        filter_values: Dict {col: [selected_values]} for pre-filtering

    Returns:
        A pandas DataFrame (pivot result), or None if inputs are insufficient.
    """
    filtered_df = df.copy()

    # Apply filters
    for f_col, selected in filter_values.items():
        if f_col in row_c or f_col in col_c or f_col in [v[0] for v in val_items]:
            pass  # filter by the actual filter fields
        if selected and f_col in filtered_df.columns:
            filtered_df = filtered_df[filtered_df[f_col].astype(str).isin(selected)]

    if not row_c and not col_c:
        return None

    pivot_idx  = row_c  if row_c  else None
    pivot_cols = col_c  if col_c  else None

    if not val_items:
        if pivot_idx and pivot_cols:
            res = pd.crosstab(
                [filtered_df[x] for x in pivot_idx],
                [filtered_df[x] for x in pivot_cols],
                rownames=pivot_idx, colnames=pivot_cols,
                margins=True, margins_name='Grand Total')
        elif pivot_idx:
            res = pd.pivot_table(filtered_df, index=pivot_idx, values=filtered_df.columns[0],
                                 aggfunc='count', margins=True, margins_name='Grand Total')
        else:
            res = pd.pivot_table(filtered_df, index=pivot_cols, values=filtered_df.columns[0],
                                 aggfunc='count', margins=True, margins_name='Grand Total')
    else:
        agg_dict = {}
        for f, agg, _ in val_items:
            if f not in agg_dict:
                agg_dict[f] = []
            if agg not in agg_dict[f]:
                agg_dict[f].append(agg)
        for k in agg_dict:
            if len(agg_dict[k]) == 1:
                agg_dict[k] = agg_dict[k][0]
        pivot_vals = list(agg_dict.keys())

        if pivot_idx and pivot_cols:
            res = pd.pivot_table(filtered_df, index=pivot_idx, columns=pivot_cols, values=pivot_vals,
                                 aggfunc=agg_dict, margins=True, margins_name='Grand Total')
        elif pivot_idx:
            res = pd.pivot_table(filtered_df, index=pivot_idx, values=pivot_vals,
                                 aggfunc=agg_dict, margins=True, margins_name='Grand Total')
        else:
            res = pd.pivot_table(filtered_df, index=pivot_cols, values=pivot_vals,
                                 aggfunc=agg_dict, margins=True, margins_name='Grand Total')

    return res
